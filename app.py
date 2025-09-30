import os
import json
import shutil
import threading
import schedule
import time
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
from sqlalchemy import func # Added for func.date
from sqlalchemy import text

# Firebase Authentication
from firebase_config import (
    initialize_firebase, 
    verify_firebase_token, 
    get_user_role_from_email,
    FIREBASE_CONFIG
)

# ë°±ì—… ì‹œìŠ¤í…œì„ ìœ„í•œ import
try:
    import pandas as pd
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    BACKUP_EXCEL_AVAILABLE = True
except ImportError as e:
    print(f"âš ï¸ Excel ë°±ì—… ê¸°ëŠ¥ì„ ìœ„í•œ íŒ¨í‚¤ì§€ê°€ ì„¤ì¹˜ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤: {e}")
    print("   pip install pandas openpyxl ëª…ë ¹ì–´ë¡œ ì„¤ì¹˜í•˜ì„¸ìš”.")
    BACKUP_EXCEL_AVAILABLE = False

# í™˜ê²½ ë³€ìˆ˜ ë¡œë“œ
load_dotenv()

# Flask ì•± ìƒì„±
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production-firebase-auth')

# === ğŸ” ë³´ì•ˆ ì„¤ì • (2025-09-21 ì¶”ê°€) ===
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # 30ë¶„ ì„¸ì…˜ íƒ€ì„ì•„ì›ƒ
app.config['SESSION_COOKIE_HTTPONLY'] = True  # JavaScriptë¡œ ì¿ í‚¤ ì ‘ê·¼ ì°¨ë‹¨
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # ê¸°ë³¸ CSRF ê³µê²© ë°©ì§€
app.config['SESSION_COOKIE_SECURE'] = False  # ê°œë°œí™˜ê²½: False, í”„ë¡œë•ì…˜: True

# ë°ì´í„°ë² ì´ìŠ¤ ì„¤ì •
if os.environ.get('DATABASE_URL'):
    # Railway ë˜ëŠ” í”„ë¡œë•ì…˜ í™˜ê²½
    app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
    app.config['SESSION_COOKIE_SECURE'] = True  # í”„ë¡œë•ì…˜ì—ì„œëŠ” HTTPS ê°•ì œ
else:
    # ê°œë°œ í™˜ê²½ - SQLite ì‚¬ìš©
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///child_center.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# í™•ì¥ í”„ë¡œê·¸ë¨ ì´ˆê¸°í™”
db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'ë¡œê·¸ì¸ì´ í•„ìš”í•©ë‹ˆë‹¤.'

# === ğŸ›¡ï¸ ê³ ê¸‰ ë³´ì•ˆ í—¤ë” ì„¤ì • (2025-09-21 í™•ì¥) ===
@app.after_request
def set_security_headers(response):
    """ëª¨ë“  ì‘ë‹µì— ê°•í™”ëœ ë³´ì•ˆ í—¤ë” ì¶”ê°€"""
    
    # === ê¸°ë³¸ ë³´ì•ˆ í—¤ë” ===
    response.headers['X-Content-Type-Options'] = 'nosniff'  # MIME íƒ€ì… ìŠ¤ë‹ˆí•‘ ë°©ì§€
    response.headers['X-Frame-Options'] = 'DENY'  # í´ë¦­ì¬í‚¹ ë°©ì§€
    response.headers['X-XSS-Protection'] = '1; mode=block'  # XSS ê³µê²© ë°©ì§€
    
    # === Content Security Policy (CSP) ===
    # ëª¨ë“  í™˜ê²½ì—ì„œ ë™ì¼í•œ CSP ì ìš© (ë¡œì»¬=ë°°í¬ ì¼ê´€ì„±)
    csp_policy = (
        "default-src 'self'; "
        # JavaScript: Firebase SDK + Bootstrap + CDN
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' "
        "https://www.gstatic.com https://apis.google.com https://www.googleapis.com "
        "https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        # ìŠ¤íƒ€ì¼ì‹œíŠ¸: Bootstrap + Google Fonts + Firebase UI + CDN
        "style-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net https://fonts.googleapis.com https://cdnjs.cloudflare.com "
        "https://www.gstatic.com; "
        # í°íŠ¸: Bootstrap Icons + Google Fonts + CDN
        "font-src 'self' data: "
        "https://fonts.gstatic.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        # ì´ë¯¸ì§€: ëª¨ë“  HTTPS ì†ŒìŠ¤ í—ˆìš©
        "img-src 'self' data: https:; "
        # ì—°ê²°: Firebase ëª¨ë“  ì—”ë“œí¬ì¸íŠ¸ + CDN
        "connect-src 'self' "
        "https://identitytoolkit.googleapis.com https://securetoken.googleapis.com "
        "https://www.googleapis.com https://firebase.googleapis.com "
        "https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        # ë³´ì•ˆ ì •ì±…
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    response.headers['Content-Security-Policy'] = csp_policy
    
    # === HTTP Strict Transport Security (HSTS) ===
    # í”„ë¡œë•ì…˜ì—ì„œë§Œ HSTS ì ìš© (HTTPS í•„ìš”)
    if os.environ.get('DATABASE_URL'):  # í”„ë¡œë•ì…˜ í™˜ê²½ ê°ì§€
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'
    
    # === Permissions Policy ===
    # ë¶ˆí•„ìš”í•œ ë¸Œë¼ìš°ì € ê¸°ëŠ¥ ì°¨ë‹¨
    permissions_policy = (
        "accelerometer=(), "
        "camera=(), "
        "geolocation=(), "
        "gyroscope=(), "
        "magnetometer=(), "
        "microphone=(), "
        "payment=(), "
        "usb=()"
    )
    response.headers['Permissions-Policy'] = permissions_policy
    
    # === ì¶”ê°€ ë³´ì•ˆ í—¤ë” ===
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'  # ë¦¬í¼ëŸ¬ ì •ì±…
    response.headers['X-Permitted-Cross-Domain-Policies'] = 'none'  # Flash/PDF ì •ì±… ì°¨ë‹¨
    response.headers['Cross-Origin-Opener-Policy'] = 'same-origin'  # íŒì—… ë³´ì•ˆ
    response.headers['Cross-Origin-Resource-Policy'] = 'same-origin'  # ë¦¬ì†ŒìŠ¤ ê³µìœ  ì œí•œ
    
    return response

# === ìˆ˜ë™ í¬ì¸íŠ¸ ì•ˆì „ ê³„ì‚° í•¨ìˆ˜ (2024-09-28 ì¶”ê°€) ===
def get_manual_points_from_history(record):
    """manual_historyì—ì„œ ì•ˆì „í•˜ê²Œ manual_points ê³„ì‚° - ë°ì´í„° ì¼ê´€ì„± ë³´ì¥"""
    if not record or not record.manual_history:
        return 0
    
    try:
        history = json.loads(record.manual_history)
        if not isinstance(history, list):
            return 0
        
        total = 0
        for item in history:
            if isinstance(item, dict) and 'points' in item:
                total += item.get('points', 0)
        
        return total
    except Exception as e:
        print(f"âŒ manual_history íŒŒì‹± ì˜¤ë¥˜: {e}")
        return 0

# === â° ì„¸ì…˜ ì˜êµ¬í™” ===
@app.before_request
def make_session_permanent():
    """ëª¨ë“  ì„¸ì…˜ì„ ì˜êµ¬ ì„¸ì…˜ìœ¼ë¡œ ì„¤ì •í•˜ì—¬ íƒ€ì„ì•„ì›ƒ ì ìš©"""
    session.permanent = True

# === ğŸ›¡ï¸ ë¸Œë£¨íŠ¸í¬ìŠ¤ ê³µê²© ë°©ì§€ ì‹œìŠ¤í…œ ===
# IPë³„ ë¡œê·¸ì¸ ì‹œë„ ì¶”ì  (ë©”ëª¨ë¦¬ ê¸°ë°˜)
failed_login_attempts = {}
blocked_ips = {}

def get_client_ip():
    """í´ë¼ì´ì–¸íŠ¸ ì‹¤ì œ IP ì£¼ì†Œ ê°€ì ¸ì˜¤ê¸° (í”„ë¡ì‹œ ê³ ë ¤)"""
    if request.environ.get('HTTP_X_FORWARDED_FOR'):
        return request.environ['HTTP_X_FORWARDED_FOR'].split(',')[0].strip()
    elif request.environ.get('HTTP_X_REAL_IP'):
        return request.environ['HTTP_X_REAL_IP']
    else:
        return request.remote_addr

def is_ip_blocked(ip_address):
    """IPê°€ ì°¨ë‹¨ë˜ì—ˆëŠ”ì§€ í™•ì¸"""
    if ip_address in blocked_ips:
        block_time = blocked_ips[ip_address]
        # 30ë¶„(1800ì´ˆ) í›„ ìë™ í•´ì œ
        if datetime.utcnow() - block_time < timedelta(minutes=30):
            return True
        else:
            # ì°¨ë‹¨ í•´ì œ
            del blocked_ips[ip_address]
            if ip_address in failed_login_attempts:
                del failed_login_attempts[ip_address]
    return False

def record_failed_login(ip_address):
    """ë¡œê·¸ì¸ ì‹¤íŒ¨ ê¸°ë¡"""
    current_time = datetime.utcnow()
    
    if ip_address not in failed_login_attempts:
        failed_login_attempts[ip_address] = []
    
    # ìµœê·¼ 1ì‹œê°„ ë‚´ ì‹¤íŒ¨ ê¸°ë¡ë§Œ ìœ ì§€
    failed_login_attempts[ip_address] = [
        attempt_time for attempt_time in failed_login_attempts[ip_address]
        if current_time - attempt_time < timedelta(hours=1)
    ]
    
    # ìƒˆë¡œìš´ ì‹¤íŒ¨ ê¸°ë¡ ì¶”ê°€
    failed_login_attempts[ip_address].append(current_time)
    
    # 5íšŒ ì´ìƒ ì‹¤íŒ¨ ì‹œ IP ì°¨ë‹¨
    if len(failed_login_attempts[ip_address]) >= 5:
        blocked_ips[ip_address] = current_time
        print(f"ğŸš¨ IP {ip_address} ì°¨ë‹¨ë¨ (5íšŒ ì—°ì† ë¡œê·¸ì¸ ì‹¤íŒ¨)")
        return True
    
    return False

def clear_failed_login(ip_address):
    """ë¡œê·¸ì¸ ì„±ê³µ ì‹œ ì‹¤íŒ¨ ê¸°ë¡ ì´ˆê¸°í™”"""
    if ip_address in failed_login_attempts:
        del failed_login_attempts[ip_address]

# ì»¨í…ìŠ¤íŠ¸ í”„ë¡œì„¸ì„œ: ëª¨ë“  í…œí”Œë¦¿ì—ì„œ ì„¼í„° ì •ë³´ ì‚¬ìš© ê°€ëŠ¥
@app.context_processor
def inject_center_info():
    """ëª¨ë“  í…œí”Œë¦¿ì—ì„œ ì„¼í„° ì •ë³´ë¥¼ ì‚¬ìš©í•  ìˆ˜ ìˆë„ë¡ ì»¨í…ìŠ¤íŠ¸ì— ì¶”ê°€"""
    return {
        'center_name': os.environ.get('CENTER_NAME', 'ì§€ì—­ì•„ë™ì„¼í„°'),
        'center_description': os.environ.get('CENTER_DESCRIPTION', 'í•™ìŠµê´€ë¦¬ ì‹œìŠ¤í…œ'),
        'center_location': os.environ.get('CENTER_LOCATION', 'ì„œìš¸ì‹œ'),
        'theme_color': os.environ.get('THEME_COLOR', '#ff6b35'),
        'branch_indicator_enabled': os.environ.get('BRANCH_INDICATOR_ENABLED', 'true').lower() == 'true'
    }

# ë°ì´í„°ë² ì´ìŠ¤ ëª¨ë¸
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=True)  # Firebase ì‚¬ìš© ì‹œ nullable
    password_hash = db.Column(db.String(255), nullable=True)  # Firebase ì‚¬ìš© ì‹œ nullable
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    login_attempts = db.Column(db.Integer, default=0)
    last_attempt = db.Column(db.DateTime)
    is_locked = db.Column(db.Boolean, default=False)  # ê³„ì • ì ê¸ˆ ìƒíƒœ
    locked_until = db.Column(db.DateTime, nullable=True)  # ì ê¸ˆ í•´ì œ ì‹œê°„
    
    # Firebase Auth ì „ìš© í•„ë“œë“¤
    email = db.Column(db.String(120), unique=True, nullable=True)
    firebase_uid = db.Column(db.String(128), unique=True, nullable=True)
    
    def is_account_locked(self):
        """ê³„ì •ì´ ì ê²¨ìˆëŠ”ì§€ í™•ì¸"""
        if not self.is_locked:
            return False
        
        if self.locked_until and datetime.utcnow() > self.locked_until:
            # ì ê¸ˆ ì‹œê°„ì´ ì§€ë‚¬ìœ¼ë©´ ìë™ í•´ì œ
            self.is_locked = False
            self.locked_until = None
            self.login_attempts = 0
            db.session.commit()
            return False
        
        return self.is_locked
    
    def lock_account(self, minutes=30):
        """ê³„ì • ì ê¸ˆ"""
        self.is_locked = True
        self.locked_until = datetime.utcnow() + timedelta(minutes=minutes)
        self.login_attempts += 1
        db.session.commit()
    
    def unlock_account(self):
        """ê³„ì • ì ê¸ˆ í•´ì œ"""
        self.is_locked = False
        self.locked_until = None
        self.login_attempts = 0
        db.session.commit()

class Child(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    grade = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # ëˆ„ì  í¬ì¸íŠ¸ (ì „ì²´ ê³¼ëª© í•©ê³„)
    cumulative_points = db.Column(db.Integer, default=0)
    
    # ê´€ê³„ ì„¤ì •
    learning_records = db.relationship('LearningRecord', backref='child', lazy=True, cascade='all, delete-orphan')
    notes = db.relationship('ChildNote', backref='child', lazy=True, cascade='all, delete-orphan')
    daily_points = db.relationship('DailyPoints', backref='child_ref', lazy=True, cascade='all, delete-orphan')
    include_in_stats = db.Column(db.Boolean, default=True) # í†µê³„ì— í¬í•¨í• ì§€ ì—¬ë¶€

class LearningRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    child_id = db.Column(db.Integer, db.ForeignKey('child.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    
    # êµ­ì–´
    korean_problems_solved = db.Column(db.Integer, default=0)
    korean_problems_correct = db.Column(db.Integer, default=0)
    korean_score = db.Column(db.Float, default=0)
    korean_last_page = db.Column(db.Integer, default=0)
    
    # ìˆ ìˆ˜í•™
    math_problems_solved = db.Column(db.Integer, default=0)
    math_problems_correct = db.Column(db.Integer, default=0)
    math_score = db.Column(db.Float, default=0)
    math_last_page = db.Column(db.Integer, default=0)
    
    # ë…ì„œ
    reading_completed = db.Column(db.Boolean, default=False)
    reading_score = db.Column(db.Float, default=0)
    
    # ì´ì 
    total_score = db.Column(db.Float, default=0)
    
    # ë©”íƒ€ë°ì´í„°
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class ChildNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    child_id = db.Column(db.Integer, db.ForeignKey('child.id'), nullable=False)
    note = db.Column(db.Text, nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # ê´€ê³„ ì„¤ì •
    creator = db.relationship('User', backref='notes', lazy=True)

# ìƒˆë¡œìš´ í¬ì¸íŠ¸ ì‹œìŠ¤í…œì„ ìœ„í•œ í…Œì´ë¸”
class DailyPoints(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    child_id = db.Column(db.Integer, db.ForeignKey('child.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    
    # ê° ê³¼ëª©ë³„ í¬ì¸íŠ¸ (200 ë˜ëŠ” 100)
    korean_points = db.Column(db.Integer, default=0)
    math_points = db.Column(db.Integer, default=0)
    ssen_points = db.Column(db.Integer, default=0)
    reading_points = db.Column(db.Integer, default=0)
    
    # ìƒˆ ê³¼ëª©ë“¤ (2025-09-17 ì¶”ê°€)
    piano_points = db.Column(db.Integer, default=0)        # í”¼ì•„ë…¸
    english_points = db.Column(db.Integer, default=0)      # ì˜ì–´
    advanced_math_points = db.Column(db.Integer, default=0) # ê³ í•™ë…„ìˆ˜í•™
    writing_points = db.Column(db.Integer, default=0)      # ì“°ê¸°
    
    # ìˆ˜ë™ í¬ì¸íŠ¸ ê´€ë¦¬
    manual_points = db.Column(db.Integer, default=0)       # ìˆ˜ë™ ì¶”ê°€/ì°¨ê° í•©ê³„
    manual_history = db.Column(db.Text, default='[]')     # JSON í˜•íƒœ íˆìŠ¤í† ë¦¬
    
    # ì´ í¬ì¸íŠ¸
    total_points = db.Column(db.Integer, default=0)
    
    # ë©”íƒ€ë°ì´í„°
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # ê´€ê³„ ì„¤ì •
    child = db.relationship('Child', lazy=True)
    creator = db.relationship('User', backref='points_records', lazy=True)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ì ìˆ˜ ê³„ì‚° í•¨ìˆ˜
def calculate_score(correct, total):
    if total == 0:
        return 0
    return (correct / total) * 100

# ê¶Œí•œ í™•ì¸ í•¨ìˆ˜
def check_permission(required_roles=None, excluded_roles=None):
    """ê¶Œí•œ í™•ì¸ í•¨ìˆ˜"""
    if not current_user.is_authenticated:
        return False
    
    # ì œì™¸ëœ ì—­í•  í™•ì¸
    if excluded_roles and current_user.role in excluded_roles:
        return False
    
    # í•„ìš”í•œ ì—­í•  í™•ì¸
    if required_roles and current_user.role not in required_roles:
        return False
    
    return True

# ì‚¬ìš© ì˜ˆì‹œ:
# @app.route('/admin')
# @login_required
# def admin_page():
#     if not check_permission(required_roles=['ì„¼í„°ì¥', 'ê°œë°œì']):
#         abort(403)
#     return render_template('admin.html')

# ë°ì´í„°ë² ì´ìŠ¤ ì´ˆê¸°í™” í•¨ìˆ˜
def init_db():
    """âš ï¸ ì£¼ì˜: ì´ í•¨ìˆ˜ëŠ” ê°œë°œ/í…ŒìŠ¤íŠ¸ í™˜ê²½ì—ì„œë§Œ ì‚¬ìš©í•˜ì„¸ìš”!"""
    print("ê²½ê³ : init_db() í•¨ìˆ˜ê°€ í˜¸ì¶œë˜ì—ˆìŠµë‹ˆë‹¤!")
    print("í˜„ì¬ ë°ì´í„°ë² ì´ìŠ¤ì˜ ëª¨ë“  ë°ì´í„°ê°€ ì‚­ì œë©ë‹ˆë‹¤!")
    
    # ì‚¬ìš©ì í™•ì¸ (ì•ˆì „ì¥ì¹˜)
    confirm = input("ì •ë§ë¡œ ëª¨ë“  ë°ì´í„°ë¥¼ ì‚­ì œí•˜ì‹œê² ìŠµë‹ˆê¹Œ? (yes/no): ")
    if confirm.lower() != 'yes':
        print("ë°ì´í„° ì‚­ì œê°€ ì·¨ì†Œë˜ì—ˆìŠµë‹ˆë‹¤.")
        return
    
    with app.app_context():
        # ê¸°ì¡´ í…Œì´ë¸” ì‚­ì œ í›„ ì¬ìƒì„± (ìŠ¤í‚¤ë§ˆ ë³€ê²½ ë°˜ì˜)
        db.drop_all()
        db.create_all()
        
        # ê¸°ë³¸ ì‚¬ìš©ì ê³„ì • ìƒì„± (í™˜ê²½ë³€ìˆ˜ì—ì„œ ì½ì–´ì˜´)
        import os
        from dotenv import load_dotenv
        
        # í™˜ê²½ë³€ìˆ˜ ë¡œë“œ
        load_dotenv()
        
        # í™˜ê²½ë³€ìˆ˜ì—ì„œ ì‚¬ìš©ì ì •ë³´ ì½ê¸°
        usernames = os.environ.get('DEFAULT_USERS', 'developer,center_head,care_teacher').split(',')
        passwords = os.environ.get('DEFAULT_PASSWORDS', 'dev123,center123!,care123!').split(',')
        roles = os.environ.get('DEFAULT_USER_ROLES', 'ê°œë°œì,ì„¼í„°ì¥,ëŒë´„ì„ ìƒë‹˜').split(',')
        
        # ì‚¬ìš©ì ë°ì´í„° ìƒì„±
        default_users = []
        for i, username in enumerate(usernames):
            if i < len(passwords) and i < len(roles):
                default_users.append({
                    'username': username.strip(),
                    'name': roles[i].strip(),
                    'role': roles[i].strip(),
                    'password': passwords[i].strip()
                })
        
        # í™˜ê²½ë³€ìˆ˜ì—ì„œ ì½ì„ ìˆ˜ ì—†ëŠ” ê²½ìš° ê¸°ë³¸ê°’ ì‚¬ìš©
        if not default_users:
            print("âš ï¸ í™˜ê²½ë³€ìˆ˜ì—ì„œ ì‚¬ìš©ì ë°ì´í„°ë¥¼ ì½ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤. ê¸°ë³¸ ë°ì´í„°ë¥¼ ì‚¬ìš©í•©ë‹ˆë‹¤.")
        default_users = [
            {'username': 'developer', 'name': 'ê°œë°œì', 'role': 'ê°œë°œì', 'password': 'dev123'},
            {'username': 'center_head', 'name': 'ì„¼í„°ì¥', 'role': 'ì„¼í„°ì¥', 'password': 'center123!'},
                {'username': 'care_teacher', 'name': 'ëŒë´„ì„ ìƒë‹˜', 'role': 'ëŒë´„ì„ ìƒë‹˜', 'password': 'care123!'}
        ]
        
        for user_data in default_users:
                password_hash = generate_password_hash(user_data['password'])
                user = User(
                    username=user_data['username'],
                    password_hash=password_hash,
                    name=user_data['name'],
                    role=user_data['role']
                )
                db.session.add(user)
        
        # í…ŒìŠ¤íŠ¸ìš© ì•„ë™ ë°ì´í„° ì¶”ê°€ (í™˜ê²½ë³€ìˆ˜ì—ì„œ ì½ì–´ì˜´)
        test_children_data = []
        
        # í™˜ê²½ë³€ìˆ˜ì—ì„œ í…ŒìŠ¤íŠ¸ ì•„ë™ ë°ì´í„° ì½ê¸°
        test_children_count = int(os.environ.get('TEST_CHILDREN_COUNT', 4))
        
        # 1í•™ë…„ë¶€í„° ì‹œì‘í•´ì„œ í…ŒìŠ¤íŠ¸ ì•„ë™ ìƒì„±
        for i in range(test_children_count):
            grade = (i % 4) + 1  # 1-4í•™ë…„ ìˆœí™˜
            env_key = f'CHILDREN_GRADE{grade}'
            children_names = os.environ.get(env_key, '').split(',')
            
            if children_names and len(children_names) > 0:
                name = children_names[0].strip()  # ì²« ë²ˆì§¸ ì•„ë™ ì‚¬ìš©
                include_in_stats = (i < test_children_count - 1)  # ë§ˆì§€ë§‰ ì•„ë™ë§Œ í†µê³„ ì œì™¸
                
                test_children_data.append(Child(
                    name=name,
                    grade=grade,
                    include_in_stats=include_in_stats
                ))
        
        # í™˜ê²½ë³€ìˆ˜ì—ì„œ ì½ì„ ìˆ˜ ì—†ëŠ” ê²½ìš° ê¸°ë³¸ê°’ ì‚¬ìš©
        if not test_children_data:
            print("âš ï¸ í™˜ê²½ë³€ìˆ˜ì—ì„œ í…ŒìŠ¤íŠ¸ ì•„ë™ ë°ì´í„°ë¥¼ ì½ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤. ê¸°ë³¸ ë°ì´í„°ë¥¼ ì‚¬ìš©í•©ë‹ˆë‹¤.")
            test_children_data = [
            Child(name='ê¹€ì² ìˆ˜', grade=3, include_in_stats=True),
            Child(name='ë°•ì˜í¬', grade=3, include_in_stats=True),
            Child(name='ì´ë¯¼ìˆ˜', grade=4, include_in_stats=True),
            Child(name='ìµœì§€ì˜', grade=4, include_in_stats=False),  # í†µê³„ ì œì™¸ ì˜ˆì‹œ
        ]
        
        test_children = test_children_data
        
        for child in test_children:
            db.session.add(child)
        
        db.session.commit()  # ì•„ë™ ë¨¼ì € ì €ì¥
        
        # í…ŒìŠ¤íŠ¸ìš© ì ìˆ˜ ë°ì´í„° ì¶”ê°€
        from datetime import date, timedelta
        today = date.today()
        
        test_records = [
            # ê¹€ì² ìˆ˜ (3í•™ë…„) - ìµœê·¼ 3ì¼ ê¸°ë¡
            LearningRecord(
                child_id=1, date=today,
                korean_problems_solved=20, korean_problems_correct=18, korean_last_page=15,
                math_problems_solved=15, math_problems_correct=12, math_last_page=22,
                reading_completed=True, reading_score=200,
                korean_score=90, math_score=80, total_score=370, created_by=1
            ),
            LearningRecord(
                child_id=1, date=today - timedelta(days=1),
                korean_problems_solved=18, korean_problems_correct=15, korean_last_page=14,
                math_problems_solved=12, math_problems_correct=10, math_last_page=21,
                reading_completed=True, reading_score=100,
                korean_score=83.3, math_score=83.3, total_score=266.6, created_by=1
            ),
            
            # ë°•ì˜í¬ (3í•™ë…„) - ê°™ì€ í˜ì´ì§€ ë¹„êµìš©
            LearningRecord(
                child_id=2, date=today,
                korean_problems_solved=20, korean_problems_correct=19, korean_last_page=15,
                math_problems_solved=15, math_problems_correct=14, math_last_page=22,
                reading_completed=True, reading_score=200,
                korean_score=95, math_score=93.3, total_score=388.3, created_by=1
            ),
            LearningRecord(
                child_id=2, date=today - timedelta(days=2),
                korean_problems_solved=16, korean_problems_correct=14, korean_last_page=13,
                math_problems_solved=10, math_problems_correct=9, math_last_page=20,
                reading_completed=False, reading_score=100,
                korean_score=87.5, math_score=90, total_score=277.5, created_by=1
            ),
            
            # ì´ë¯¼ìˆ˜ (4í•™ë…„) - ë‹¤ë¥¸ í•™ë…„
            LearningRecord(
                child_id=3, date=today,
                korean_problems_solved=25, korean_problems_correct=20, korean_last_page=18,
                math_problems_solved=20, math_problems_correct=16, math_last_page=25,
                reading_completed=True, reading_score=200,
                korean_score=80, math_score=80, total_score=360, created_by=1
            ),
        ]
        
        for record in test_records:
            db.session.add(record)
        
        # DailyPoints í…ŒìŠ¤íŠ¸ ë°ì´í„° ì¶”ê°€ (ì‹œê°í™”ìš©)
        from datetime import date, timedelta
        today = date.today()
        
        # ìµœê·¼ 28ì¼ê°„ì˜ í¬ì¸íŠ¸ ë°ì´í„° ìƒì„±
        for i in range(28):
            current_date = today - timedelta(days=i)
            
            # ê° ì•„ë™ë³„ë¡œ ëœë¤í•œ í¬ì¸íŠ¸ ìƒì„±
            for child_id in [1, 2, 3]:  # ê¹€ì² ìˆ˜, ë°•ì˜í¬, ì´ë¯¼ìˆ˜
                # ëœë¤í•œ í¬ì¸íŠ¸ ìƒì„± (200, 100, 0 ì¤‘ì—ì„œ)
                import random
                korean_points = random.choice([0, 100, 200])
                math_points = random.choice([0, 100, 200])
                ssen_points = random.choice([0, 100, 200])
                reading_points = random.choice([0, 100, 200])
                total_points = korean_points + math_points + ssen_points + reading_points + piano_points + english_points + advanced_math_points + writing_points + manual_points
                
                # ì¼ë¶€ ë‚ ì§œëŠ” ê¸°ë¡ ì—†ìŒ (ë” í˜„ì‹¤ì ì¸ ë°ì´í„°)
                if random.random() > 0.3:  # 70% í™•ë¥ ë¡œ ê¸°ë¡ ìƒì„±
                    daily_point = DailyPoints(
                        child_id=child_id,
                        date=current_date,
                        korean_points=korean_points,
                        math_points=math_points,
                        ssen_points=ssen_points,
                        reading_points=reading_points,
                piano_points=piano_points,
                english_points=english_points,
                advanced_math_points=advanced_math_points,
                writing_points=writing_points,
                manual_points=manual_points,
                        total_points=total_points,
                        created_by=1
                    )
                    db.session.add(daily_point)
        
        db.session.commit()
        print("ë°ì´í„°ë² ì´ìŠ¤ê°€ ì´ˆê¸°í™”ë˜ì—ˆìŠµë‹ˆë‹¤. (í…ŒìŠ¤íŠ¸ ë°ì´í„° + ì ìˆ˜ ê¸°ë¡ í¬í•¨)")

# ë¼ìš°íŠ¸
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# === ì™„ì „ Firebase Auth ì‹œìŠ¤í…œ ===
@app.route('/login', methods=['GET', 'POST'])
def login():
    """ì™„ì „ Firebase Auth ê¸°ë°˜ ë¡œê·¸ì¸"""
    if request.method == 'POST':
        # === ğŸ›¡ï¸ ë¸Œë£¨íŠ¸í¬ìŠ¤ ê³µê²© ë°©ì§€ ì²´í¬ ===
        client_ip = get_client_ip()
        
        if is_ip_blocked(client_ip):
            flash('âš ï¸ ë³´ì•ˆìƒ ë¡œê·¸ì¸ì´ ì¼ì‹œì ìœ¼ë¡œ ì œí•œë˜ì—ˆìŠµë‹ˆë‹¤. 30ë¶„ í›„ ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”.', 'error')
            return render_template('login.html', firebase_config=FIREBASE_CONFIG)
        
        # Firebase í† í° ê²€ì¦
        token = request.json.get('token') if request.is_json else request.form.get('token')
        
        if not token:
            flash('ë¡œê·¸ì¸ í† í°ì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
            return render_template('login.html', firebase_config=FIREBASE_CONFIG)
        
        # Firebase í† í° ê²€ì¦
        decoded_token = verify_firebase_token(token)
        
        if decoded_token:
            # ì‚¬ìš©ì ì •ë³´ ì¶”ì¶œ
            firebase_uid = decoded_token['uid']
            email = decoded_token['email']
            name = decoded_token.get('name', email.split('@')[0])
            
            # Firebase ì‚¬ìš©ìë¡œ ë¡œê·¸ì¸ ì²˜ë¦¬ (firebase_uid ë˜ëŠ” emailë¡œ ì°¾ê¸°)
            user = User.query.filter(
                (User.firebase_uid == firebase_uid) | (User.email == email)
            ).first()
            
            if not user:
                # ìƒˆ Firebase ì‚¬ìš©ì ìƒì„±
                user = User(
                    firebase_uid=firebase_uid,
                    email=email,
                    name=name,
                    role=get_user_role_from_email(email),
                    username=email.split('@')[0],  # í˜¸í™˜ì„±ì„ ìœ„í•´
                    password_hash=''  # Firebase ì‚¬ìš©ìëŠ” ë¹„ë°€ë²ˆí˜¸ ì—†ìŒ
                )
                db.session.add(user)
                db.session.commit()
                print(f"âœ… ìƒˆ Firebase ì‚¬ìš©ì ìƒì„±: {email}")
            elif not user.firebase_uid:
                # ê¸°ì¡´ ì‚¬ìš©ìì— firebase_uid ì¶”ê°€
                user.firebase_uid = firebase_uid
                db.session.commit()
                print(f"âœ… ê¸°ì¡´ ì‚¬ìš©ì Firebase UID ì—…ë°ì´íŠ¸: {email}")
            
            # Firebase ì‚¬ìš©ìë¡œ ë¡œê·¸ì¸
                login_user(user)
            
            # === ğŸ›¡ï¸ ë¡œê·¸ì¸ ì„±ê³µ ì‹œ ì‹¤íŒ¨ ê¸°ë¡ ì´ˆê¸°í™” ===
            clear_failed_login(client_ip)
            
            flash(f'{user.name}ë‹˜, Firebase ì¸ì¦ìœ¼ë¡œ ë¡œê·¸ì¸ë˜ì—ˆìŠµë‹ˆë‹¤!', 'success')
            
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('dashboard')})
            else:
                return redirect(url_for('dashboard'))
        else:
            # === ğŸ›¡ï¸ ë¡œê·¸ì¸ ì‹¤íŒ¨ ì‹œ ì‹¤íŒ¨ ê¸°ë¡ ===
            is_now_blocked = record_failed_login(client_ip)
            if is_now_blocked:
                flash('âš ï¸ ì—°ì†ëœ ë¡œê·¸ì¸ ì‹¤íŒ¨ë¡œ ì¸í•´ 30ë¶„ê°„ ë¡œê·¸ì¸ì´ ì œí•œë©ë‹ˆë‹¤.', 'error')
            else:
                flash('Firebase ì¸ì¦ì— ì‹¤íŒ¨í–ˆìŠµë‹ˆë‹¤.', 'error')
            
            if request.is_json:
                return jsonify({'success': False, 'error': 'Invalid Firebase token'})
    
    # Firebase ì„¤ì • ì •ë³´ë¥¼ í…œí”Œë¦¿ì— ì „ë‹¬
    return render_template('login.html', firebase_config=FIREBASE_CONFIG)

@app.route('/firebase-login', methods=['POST'])
def firebase_login():
    """Firebase Auth API ì—”ë“œí¬ì¸íŠ¸"""
    try:
        # === ğŸ›¡ï¸ ë¸Œë£¨íŠ¸í¬ìŠ¤ ê³µê²© ë°©ì§€ ì²´í¬ ===
        client_ip = get_client_ip()
        
        if is_ip_blocked(client_ip):
            return jsonify({'success': False, 'error': 'ë³´ì•ˆìƒ ë¡œê·¸ì¸ì´ ì¼ì‹œì ìœ¼ë¡œ ì œí•œë˜ì—ˆìŠµë‹ˆë‹¤. 30ë¶„ í›„ ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”.'})
        
        data = request.get_json()
        token = data.get('token')
        
        if not token:
            return jsonify({'success': False, 'error': 'Token is required'})
        
        # Firebase í† í° ê²€ì¦
        decoded_token = verify_firebase_token(token)
        
        if decoded_token:
            # ì‚¬ìš©ì ì •ë³´ ì¶”ì¶œ
            firebase_uid = decoded_token['uid']
            email = decoded_token['email']
            name = decoded_token.get('name', email.split('@')[0])
            
            # Firebase ì‚¬ìš©ìë¡œ ë¡œê·¸ì¸ ì²˜ë¦¬ (firebase_uid ë˜ëŠ” emailë¡œ ì°¾ê¸°)
            user = User.query.filter(
                (User.firebase_uid == firebase_uid) | (User.email == email)
            ).first()
            
            if not user:
                # ìƒˆ Firebase ì‚¬ìš©ì ìƒì„±
                user = User(
                    firebase_uid=firebase_uid,
                    email=email,
                    name=name,
                    role=get_user_role_from_email(email),
                    username=email.split('@')[0],  # í˜¸í™˜ì„±ì„ ìœ„í•´
                    password_hash=''  # Firebase ì‚¬ìš©ìëŠ” ë¹„ë°€ë²ˆí˜¸ ì—†ìŒ
                )
                db.session.add(user)
                db.session.commit()
                print(f"âœ… ìƒˆ Firebase ì‚¬ìš©ì ìƒì„±: {email}")
            elif not user.firebase_uid:
                # ê¸°ì¡´ ì‚¬ìš©ìì— firebase_uid ì¶”ê°€
                user.firebase_uid = firebase_uid
                db.session.commit()
                print(f"âœ… ê¸°ì¡´ ì‚¬ìš©ì Firebase UID ì—…ë°ì´íŠ¸: {email}")
            
            # Firebase ì‚¬ìš©ìë¡œ ë¡œê·¸ì¸
            login_user(user)
            
            # === ğŸ›¡ï¸ ë¡œê·¸ì¸ ì„±ê³µ ì‹œ ì‹¤íŒ¨ ê¸°ë¡ ì´ˆê¸°í™” ===
            clear_failed_login(client_ip)
            
            return jsonify({
                'success': True, 
                'redirect': url_for('dashboard'),
                'user': {
                    'id': user.id,
                    'name': user.name,
                    'email': user.email,
                    'role': user.role,
                    'firebase_uid': user.firebase_uid
                }
            })
        else:
            # === ğŸ›¡ï¸ ë¡œê·¸ì¸ ì‹¤íŒ¨ ì‹œ ì‹¤íŒ¨ ê¸°ë¡ ===
            record_failed_login(client_ip)
            return jsonify({'success': False, 'error': 'Invalid Firebase token'})
    
    except Exception as e:
        print(f"Firebase login error: {e}")
        # === ğŸ›¡ï¸ ì˜¤ë¥˜ ì‹œì—ë„ ì‹¤íŒ¨ ê¸°ë¡ ===
        client_ip = get_client_ip()
        record_failed_login(client_ip)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('ë¡œê·¸ì•„ì›ƒë˜ì—ˆìŠµë‹ˆë‹¤.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    today = datetime.now(timezone.utc).date()
    
    # ====== [í¬ì¸íŠ¸ ì‹œìŠ¤í…œ í†µê³„ ê³„ì‚°] ======
    # ì˜¤ëŠ˜ í¬ì¸íŠ¸ë¥¼ ì…ë ¥í•œ ì•„ë™ ìˆ˜
    today_points_children = db.session.query(DailyPoints.child_id).filter_by(date=today).distinct().count()
    
    # ì „ì²´ ë“±ë¡ ì•„ë™ ìˆ˜
    total_children = Child.query.count()
    
    # ì´ë²ˆ ì£¼ í‰ê·  í¬ì¸íŠ¸ ê³„ì‚°
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    weekly_points = DailyPoints.query.filter(
        DailyPoints.date >= week_start,
        DailyPoints.date <= week_end
    ).all()
    
    if weekly_points:
        total_weekly_points = sum(record.total_points for record in weekly_points)
        weekly_avg_points = int(round(total_weekly_points / len(weekly_points), 0))
        weekly_points_count = len(weekly_points)
    else:
        weekly_avg_points = 0
        total_weekly_points = 0
        weekly_points_count = 0
    
    # ì´ë²ˆ ì£¼ í¬ì¸íŠ¸ ì°¸ì—¬ìœ¨ ê³„ì‚°
    weekly_participants = db.session.query(DailyPoints.child_id).filter(
        DailyPoints.date >= week_start,
        DailyPoints.date <= week_end
    ).distinct().count()
    
    if total_children > 0:
        participation_rate = int(round((weekly_participants / total_children) * 100, 0))
    else:
        participation_rate = 0
    
    # ìµœê·¼ í¬ì¸íŠ¸ ê¸°ë¡ (ìµœê·¼ 10ê°œ)
    recent_records = db.session.query(DailyPoints, Child).join(Child).order_by(DailyPoints.created_at.desc()).limit(10).all()
    
    # ====== [ê³¼ëª©ë³„ ì£¼ê°„ í‰ê·  í¬ì¸íŠ¸ ê³„ì‚°] ======
    weekly_korean_avg = 0
    weekly_math_avg = 0
    weekly_ssen_avg = 0
    weekly_reading_avg = 0
    weekly_total_points = 0
    
    if weekly_points:
        # ê³¼ëª©ë³„ í‰ê·  ê³„ì‚°
        korean_points = [record.korean_points for record in weekly_points if record.korean_points > 0]
        math_points = [record.math_points for record in weekly_points if record.math_points > 0]
        ssen_points = [record.ssen_points for record in weekly_points if record.ssen_points > 0]
        reading_points = [record.reading_points for record in weekly_points if record.reading_points > 0]
        
        weekly_korean_avg = round(sum(korean_points) / len(korean_points), 0) if korean_points else 0
        weekly_math_avg = round(sum(math_points) / len(math_points), 0) if math_points else 0
        weekly_ssen_avg = round(sum(ssen_points) / len(ssen_points), 0) if ssen_points else 0
        weekly_reading_avg = round(sum(reading_points) / len(reading_points), 0) if reading_points else 0
        
        # ì£¼ê°„ ì´ í¬ì¸íŠ¸
        weekly_total_points = sum(record.total_points for record in weekly_points)
    
    # ====== [ì•Œë¦¼ ì‹œìŠ¤í…œ í™œì„±í™”] ======
    notifications = get_user_notifications(current_user.id, limit=5)
    
    return render_template('dashboard.html', 
                         today_points_children=today_points_children,
                         total_children=total_children,
                         weekly_avg_points=weekly_avg_points,
                         participation_rate=participation_rate,
                         recent_records=recent_records,
                         notifications=notifications,
                         weekly_korean_avg=weekly_korean_avg,
                         weekly_math_avg=weekly_math_avg,
                         weekly_ssen_avg=weekly_ssen_avg,
                         weekly_reading_avg=weekly_reading_avg,
                         weekly_total_points=weekly_total_points,
                         weekly_points_count=weekly_points_count)

# ì•„ë™ ê´€ë¦¬ ë¼ìš°íŠ¸
@app.route('/children')
@login_required
def children_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    grade_filter = request.args.get('grade', '', type=str)
    
    query = Child.query
    
    # ê²€ìƒ‰ í•„í„°
    if search:
        query = query.filter(Child.name.contains(search))
    
    # í•™ë…„ í•„í„°
    if grade_filter:
        query = query.filter(Child.grade == int(grade_filter))
    
    # í˜ì´ì§€ë„¤ì´ì…˜
    children = query.order_by(Child.name).paginate(
        page=page, per_page=10, error_out=False
    )
    
    # í•™ë…„ ëª©ë¡ (í•„í„°ìš©)
    grades = db.session.query(Child.grade).distinct().order_by(Child.grade).all()
    grade_list = [g[0] for g in grades]
    
    return render_template('children/list.html', 
                         children=children, 
                         search=search,
                         grade_filter=grade_filter,
                         grade_list=grade_list)

@app.route('/children/add', methods=['GET', 'POST'])
@login_required
def add_child():
    if request.method == 'POST':
        name = request.form['name'].strip()
        grade = request.form['grade']
        
        # ìœ íš¨ì„± ê²€ì‚¬
        if not name:
            flash('ì´ë¦„ì„ ì…ë ¥í•´ì£¼ì„¸ìš”.', 'error')
            return render_template('children/form.html')
        
        if not grade or int(grade) < 1 or int(grade) > 6:
            flash('í•™ë…„ì„ ì˜¬ë°”ë¥´ê²Œ ì„ íƒí•´ì£¼ì„¸ìš”. (1-6í•™ë…„)', 'error')
            return render_template('children/form.html')
        
        # ì¤‘ë³µ ì´ë¦„ í™•ì¸
        existing_child = Child.query.filter_by(name=name).first()
        if existing_child:
            flash('ì´ë¯¸ ë“±ë¡ëœ ì´ë¦„ì…ë‹ˆë‹¤. ë‹¤ë¥¸ ì´ë¦„ì„ ì‚¬ìš©í•´ì£¼ì„¸ìš”.', 'error')
            return render_template('children/form.html')
        
        # ì•„ë™ ë“±ë¡
        child = Child(name=name, grade=int(grade))
        db.session.add(child)
        db.session.commit()
        
        flash(f'{name} ì•„ë™ì´ ì„±ê³µì ìœ¼ë¡œ ë“±ë¡ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
        return redirect(url_for('children_list'))
    
    return render_template('children/form.html')

@app.route('/children/<int:child_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_child(child_id):
    child = Child.query.get_or_404(child_id)
    
    if request.method == 'POST':
        name = request.form['name'].strip()
        grade = request.form['grade']
        
        # ìœ íš¨ì„± ê²€ì‚¬
        if not name:
            flash('ì´ë¦„ì„ ì…ë ¥í•´ì£¼ì„¸ìš”.', 'error')
            return render_template('children/form.html', child=child)
        
        if not grade or int(grade) < 1 or int(grade) > 6:
            flash('í•™ë…„ì„ ì˜¬ë°”ë¥´ê²Œ ì„ íƒí•´ì£¼ì„¸ìš”. (1-6í•™ë…„)', 'error')
            return render_template('children/form.html', child=child)
        
        # ì¤‘ë³µ ì´ë¦„ í™•ì¸ (ìê¸° ìì‹  ì œì™¸)
        existing_child = Child.query.filter(Child.name == name, Child.id != child_id).first()
        if existing_child:
            flash('ì´ë¯¸ ë“±ë¡ëœ ì´ë¦„ì…ë‹ˆë‹¤. ë‹¤ë¥¸ ì´ë¦„ì„ ì‚¬ìš©í•´ì£¼ì„¸ìš”.', 'error')
            return render_template('children/form.html', child=child)
        
        # ì•„ë™ ì •ë³´ ì—…ë°ì´íŠ¸
        child.name = name
        child.grade = int(grade)
        db.session.commit()
        
        flash(f'{name} ì•„ë™ ì •ë³´ê°€ ì„±ê³µì ìœ¼ë¡œ ìˆ˜ì •ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
        return redirect(url_for('children_list'))
    
    return render_template('children/form.html', child=child)

@app.route('/children/<int:child_id>/delete', methods=['POST'])
@login_required
def delete_child(child_id):
    child = Child.query.get_or_404(child_id)
    child_name = child.name
    
    # ê¶Œí•œ í™•ì¸ (ì„¼í„°ì¥ê³¼ ëŒë´„ì„ ìƒë‹˜ë§Œ ì‚­ì œ ê°€ëŠ¥)
    if current_user.role not in ['ê°œë°œì']:
        flash('ì•„ë™ ì‚­ì œ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('children_list'))
    
    try:
        # ê´€ë ¨ ê¸°ë¡ë“¤ë„ í•¨ê»˜ ì‚­ì œë¨ (cascade ì„¤ì •)
        db.session.delete(child)
        db.session.commit()
        
        flash(f'{child_name} ì•„ë™ê³¼ ê´€ë ¨ ê¸°ë¡ì´ ëª¨ë‘ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
    except Exception as e:
        db.session.rollback()
        print(f"ğŸ› DEBUG: ì•„ë™ ì‚­ì œ ì˜¤ë¥˜ - {str(e)}")
        print(f"ğŸ› DEBUG: child_id: {child_id}, child_name: {child_name}")
        flash(f'ì‚­ì œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}', 'error')
    
    return redirect(url_for('children_list'))

@app.route('/children/<int:child_id>')
@login_required
def child_detail(child_id):
    child = Child.query.get_or_404(child_id)
    
    # í˜ì´ì§€ë„¤ì´ì…˜ íŒŒë¼ë¯¸í„°
    page = request.args.get('page', 1, type=int)
    per_page = 20  # í•œ í˜ì´ì§€ë‹¹ 20ê°œ ê¸°ë¡
    
    # ìƒˆë¡œìš´ í¬ì¸íŠ¸ ì‹œìŠ¤í…œ ê¸°ë¡ë“¤ (ì¤‘ë³µ ì œê±°) - í˜ì´ì§€ë„¤ì´ì…˜ ì ìš©
    from sqlalchemy import text
    
    # ì „ì²´ ê¸°ë¡ ìˆ˜ ê³„ì‚°
    count_result = db.session.execute(text("""
        SELECT COUNT(DISTINCT date) as total_count
        FROM daily_points 
        WHERE child_id = :child_id
    """), {"child_id": child_id})
    total_records = count_result.fetchone()[0]
    
    # í˜ì´ì§€ë„¤ì´ì…˜ ê³„ì‚°
    offset = (page - 1) * per_page
    total_pages = (total_records + per_page - 1) // per_page
    
    # í˜ì´ì§€ë³„ ê¸°ë¡ ì¡°íšŒ
    result = db.session.execute(text("""
        SELECT id, date, korean_points, math_points, ssen_points, reading_points, total_points, created_at
        FROM daily_points 
        WHERE child_id = :child_id 
        AND id IN (
            SELECT MAX(id) 
            FROM daily_points 
            WHERE child_id = :child_id 
            GROUP BY date
        )
        ORDER BY date DESC
        LIMIT :per_page OFFSET :offset
    """), {"child_id": child_id, "per_page": per_page, "offset": offset})
    
    # DailyPoints ê°ì²´ë¡œ ë³€í™˜
    recent_records = []
    for row in result:
        # ë‚ ì§œ íƒ€ì… ë³€í™˜
        date_value = row[1]
        if isinstance(date_value, str):
            from datetime import datetime
            date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
        
        created_at_value = row[7]
        if isinstance(created_at_value, str):
            from datetime import datetime
            created_at_value = datetime.strptime(created_at_value, '%Y-%m-%d %H:%M:%S.%f')
        
        point_record = DailyPoints(
            id=row[0],
            date=date_value,
            korean_points=row[2],
            math_points=row[3],
            ssen_points=row[4],
            reading_points=row[5],
            total_points=row[6]
        )
        # created_atì„ ë³„ë„ë¡œ ì„¤ì • (í…œí”Œë¦¿ì—ì„œ ì‚¬ìš©í•  ìˆ˜ ìˆë„ë¡)
        point_record.created_at = created_at_value
        recent_records.append(point_record)
    
    # ìµœê·¼ íŠ¹ì´ì‚¬í•­ë“¤
    recent_notes = ChildNote.query.filter_by(child_id=child_id)\
                                  .order_by(ChildNote.created_at.desc())\
                                  .limit(5).all()
    
    # í†µê³„ ê³„ì‚° (ìƒˆë¡œìš´ í¬ì¸íŠ¸ ì‹œìŠ¤í…œ ê¸°ë°˜)
    if recent_records:
        # ìµœê·¼ 5ê°œ ê¸°ë¡ì˜ í‰ê· 
        recent_avg = sum(record.total_points for record in recent_records[:5]) / min(len(recent_records), 5)
        
        # ê°€ì¥ ìµœê·¼ ê¸°ë¡
        latest_record = recent_records[0] if recent_records else None
    else:
        recent_avg = 0
        latest_record = None
    
    # ì´ ëˆ„ì  í¬ì¸íŠ¸ (ì‹¤ì œ ì „ì²´ ëˆ„ì )
    total_points = child.cumulative_points
    
    return render_template('children/detail.html', 
                         child=child,
                         recent_records=recent_records,
                         recent_notes=recent_notes,
                         recent_avg=recent_avg,
                         latest_record=latest_record,
                         total_points=total_points,
                         # í˜ì´ì§€ë„¤ì´ì…˜ ì •ë³´
                         current_page=page,
                         total_pages=total_pages,
                         total_records=total_records,
                         per_page=per_page)

# ===== íŠ¹ì´ì‚¬í•­ ê´€ë¦¬ ë¼ìš°íŠ¸ =====

@app.route('/children/<int:child_id>/notes', methods=['POST'])
@login_required
def add_child_note(child_id):
    """ì•„ë™ íŠ¹ì´ì‚¬í•­ ì¶”ê°€"""
    child = Child.query.get_or_404(child_id)
    
    note_text = request.form.get('note', '').strip()
    if not note_text:
        flash('íŠ¹ì´ì‚¬í•­ì„ ì…ë ¥í•´ì£¼ì„¸ìš”.', 'error')
        return redirect(url_for('child_detail', child_id=child_id))
    
    try:
        new_note = ChildNote(
            child_id=child_id,
            note=note_text,
            created_by=current_user.id
        )
        
        db.session.add(new_note)
        db.session.commit()
        
        flash(f'âœ… {child.name} ì•„ë™ì˜ íŠ¹ì´ì‚¬í•­ì´ ì¶”ê°€ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
        
        # íŠ¹ì´ì‚¬í•­ ì¶”ê°€ ì•Œë¦¼ ìƒì„±
        print(f"DEBUG: íŠ¹ì´ì‚¬í•­ ì¶”ê°€ ì•Œë¦¼ ìƒì„± ì‹œë„ - {child.name}")
        notification = create_notification(
            title=f'ğŸ“ {child.name} íŠ¹ì´ì‚¬í•­ ì¶”ê°€',
            message=f'{current_user.name}ë‹˜ì´ {child.name} ì•„ë™ì˜ íŠ¹ì´ì‚¬í•­ì„ ì¶”ê°€í–ˆìŠµë‹ˆë‹¤.',
            notification_type='warning',
            child_id=child.id,
            target_role=None,  # ëª¨ë“  ì‚¬ìš©ìì—ê²Œ í‘œì‹œ
            priority=2,
            auto_expire=True,
            expire_days=3
        )
        print(f"DEBUG: ì•Œë¦¼ ìƒì„± ê²°ê³¼ - {notification}")
        
    except Exception as e:
        db.session.rollback()
        flash(f'âŒ íŠ¹ì´ì‚¬í•­ ì¶”ê°€ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}', 'error')
    
    return redirect(url_for('child_detail', child_id=child_id))

@app.route('/children/<int:child_id>/notes/<int:note_id>/edit', methods=['POST'])
@login_required  
def edit_child_note(child_id, note_id):
    """ì•„ë™ íŠ¹ì´ì‚¬í•­ ìˆ˜ì •"""
    child = Child.query.get_or_404(child_id)
    note = ChildNote.query.get_or_404(note_id)
    
    # ê¶Œí•œ í™•ì¸ (ì‘ì„±ì ë˜ëŠ” ê°œë°œìë§Œ ìˆ˜ì • ê°€ëŠ¥)
    # ê¶Œí•œ ì²´í¬ ì œê±° - ëª¨ë“  ì‚¬ìš©ìê°€ ìˆ˜ì • ê°€ëŠ¥
    
    note_text = request.form.get('note', '').strip()
    if not note_text:
        flash('íŠ¹ì´ì‚¬í•­ì„ ì…ë ¥í•´ì£¼ì„¸ìš”.', 'error')
        return redirect(url_for('child_detail', child_id=child_id))
    
    try:
        old_note = note.note
        note.note = note_text
        note.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'âœ… {child.name} ì•„ë™ì˜ íŠ¹ì´ì‚¬í•­ì´ ìˆ˜ì •ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
        
        # íŠ¹ì´ì‚¬í•­ ìˆ˜ì • ì•Œë¦¼ ìƒì„±
        print(f"DEBUG: íŠ¹ì´ì‚¬í•­ ìˆ˜ì • ì•Œë¦¼ ìƒì„± ì‹œë„ - {child.name}")
        notification = create_notification(
                title=f'ğŸ“ {child.name} íŠ¹ì´ì‚¬í•­ ìˆ˜ì •',
                message=f'{current_user.name}ë‹˜ì´ {child.name} ì•„ë™ì˜ íŠ¹ì´ì‚¬í•­ì„ ìˆ˜ì •í–ˆìŠµë‹ˆë‹¤.',
            notification_type='warning',
                child_id=child.id,
            target_role=None,  # ëª¨ë“  ì‚¬ìš©ìì—ê²Œ í‘œì‹œ
            priority=2,
            auto_expire=True,
            expire_days=3
        )
        print(f"DEBUG: ìˆ˜ì • ì•Œë¦¼ ìƒì„± ê²°ê³¼ - {notification}")
        
    except Exception as e:
        db.session.rollback()
        flash(f'âŒ íŠ¹ì´ì‚¬í•­ ìˆ˜ì • ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}', 'error')
    
    return redirect(url_for('child_detail', child_id=child_id))

@app.route('/children/<int:child_id>/notes/<int:note_id>/delete', methods=['POST'])
@login_required
def delete_child_note(child_id, note_id):
    """ì•„ë™ íŠ¹ì´ì‚¬í•­ ì‚­ì œ"""
    child = Child.query.get_or_404(child_id)
    note = ChildNote.query.get_or_404(note_id)
    
    # ê¶Œí•œ í™•ì¸ (ì‘ì„±ì ë˜ëŠ” ê°œë°œìë§Œ ì‚­ì œ ê°€ëŠ¥)
    # ê¶Œí•œ ì²´í¬ ì œê±° - ëª¨ë“  ì‚¬ìš©ìê°€ ì‚­ì œ ê°€ëŠ¥
    
    try:
        db.session.delete(note)
        db.session.commit()
        
        flash(f'âœ… {child.name} ì•„ë™ì˜ íŠ¹ì´ì‚¬í•­ì´ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
        
        # íŠ¹ì´ì‚¬í•­ ì‚­ì œ ì•Œë¦¼ ìƒì„±
        create_notification(
            title=f'ğŸ—‘ï¸ {child.name} íŠ¹ì´ì‚¬í•­ ì‚­ì œ',
            message=f'{current_user.name}ë‹˜ì´ {child.name} ì•„ë™ì˜ íŠ¹ì´ì‚¬í•­ì„ ì‚­ì œí–ˆìŠµë‹ˆë‹¤.',
            notification_type='warning',
            child_id=child.id,
            target_role=None,  # ëª¨ë“  ì‚¬ìš©ìì—ê²Œ í‘œì‹œ
            priority=2,
            auto_expire=True,
            expire_days=3
        )
        
    except Exception as e:
        db.session.rollback()
        flash(f'âŒ íŠ¹ì´ì‚¬í•­ ì‚­ì œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}', 'error')
    
    return redirect(url_for('child_detail', child_id=child_id))

@app.route('/children/<int:child_id>/notes/all')
@login_required
def view_all_child_notes(child_id):
    """ì•„ë™ íŠ¹ì´ì‚¬í•­ ì „ì²´ ë³´ê¸°"""
    child = Child.query.get_or_404(child_id)
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    notes = ChildNote.query.filter_by(child_id=child_id)\
                          .order_by(ChildNote.created_at.desc())\
                          .paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template('children/notes.html', child=child, notes=notes)

# ì ìˆ˜ ì…ë ¥ ë¼ìš°íŠ¸
@app.route('/scores')
@login_required
def scores_list():
    # ìµœê·¼ ì…ë ¥ëœ ì ìˆ˜ë“¤
    recent_records = db.session.query(LearningRecord, Child)\
                              .join(Child)\
                              .order_by(LearningRecord.date.desc())\
                              .limit(20).all()
    
    return render_template('scores/list.html', recent_records=recent_records)

@app.route('/scores/add', methods=['GET', 'POST'])
@login_required
def add_score():
    # URL íŒŒë¼ë¯¸í„°ì—ì„œ child_id ê°€ì ¸ì˜¤ê¸°
    preselected_child_id = request.args.get('child_id', type=int)
    
    if request.method == 'POST':
        try:
            # í¼ ë°ì´í„° ë°›ê¸°
            child_id = request.form['child_id']
            date_str = request.form['date']
            
            # êµ­ì–´ ë°ì´í„°
            korean_problems_solved = int(request.form.get('korean_problems_solved', 0))
            korean_problems_correct = int(request.form.get('korean_problems_correct', 0))
            korean_last_page = int(request.form.get('korean_last_page', 0))
            
            # ìˆ˜í•™ ë°ì´í„°  
            math_problems_solved = int(request.form.get('math_problems_correct', 0))
            math_problems_correct = int(request.form.get('math_problems_correct', 0))
            math_last_page = int(request.form.get('math_last_page', 0))
            
            # ë…ì„œ ë°ì´í„°
            reading_completed = 'reading_completed' in request.form
            reading_score = float(request.form.get('reading_score', 0))
            
            # ìœ íš¨ì„± ê²€ì‚¬
            if not child_id:
                flash('ì•„ë™ì„ ì„ íƒí•´ì£¼ì„¸ìš”.', 'error')
                return render_template('scores/form.html', children=Child.query.all())
            
            if not date_str:
                flash('ë‚ ì§œë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.', 'error')
                return render_template('scores/form.html', children=Child.query.all())
            
            # ë‚ ì§œ ë³€í™˜
            from datetime import datetime
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # ì´ë¯¸ í•´ë‹¹ ë‚ ì§œì— ê¸°ë¡ì´ ìˆëŠ”ì§€ í™•ì¸
            existing_record = LearningRecord.query.filter_by(
                child_id=child_id, 
                date=date
            ).first()
            
            if existing_record:
                flash('í•´ë‹¹ ë‚ ì§œì— ì´ë¯¸ ê¸°ë¡ì´ ìˆìŠµë‹ˆë‹¤. ìˆ˜ì •í•˜ì‹œê² ìŠµë‹ˆê¹Œ?', 'error')
                return redirect(url_for('edit_score', record_id=existing_record.id))
            
            # ì ìˆ˜ ê³„ì‚°
            korean_score = calculate_score(korean_problems_correct, korean_problems_solved) if korean_problems_solved > 0 else 0
            math_score = calculate_score(math_problems_correct, math_problems_solved) if math_problems_solved > 0 else 0
            
            # ì´ì  ê³„ì‚° (êµ­ì–´ + ìˆ˜í•™ + ë…ì„œ)
            total_score = korean_score + math_score + reading_score
            
            # ìƒˆ ê¸°ë¡ ìƒì„±
            new_record = LearningRecord(
                child_id=child_id,
                date=date,
                korean_problems_solved=korean_problems_solved,
                korean_problems_correct=korean_problems_correct,
                korean_score=korean_score,
                korean_last_page=korean_last_page,
                math_problems_solved=math_problems_solved,
                math_problems_correct=math_problems_correct,
                math_score=math_score,
                math_last_page=math_last_page,
                reading_completed=reading_completed,
                reading_score=reading_score,
                total_score=total_score,
                created_by=current_user.id
            )
            
            db.session.add(new_record)
            db.session.commit()
            
            child = Child.query.get(child_id)
            flash(f'{child.name} ì•„ë™ì˜ {date_str} í•™ìŠµ ê¸°ë¡ì´ ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
            return redirect(url_for('child_detail', child_id=child_id))
            
        except ValueError as e:
            flash('ì…ë ¥ê°’ì„ ë‹¤ì‹œ í™•ì¸í•´ì£¼ì„¸ìš”.', 'error')
            return render_template('scores/form.html', children=Child.query.all())
        except Exception as e:
            db.session.rollback()
            flash('ì €ì¥ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤. ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”.', 'error')
            return render_template('scores/form.html', children=Child.query.all())
    
    # GET ìš”ì²­ ì‹œ í¼ í‘œì‹œ
    children = Child.query.order_by(Child.name).all()
    return render_template('scores/form.html', children=children, preselected_child_id=preselected_child_id)

@app.route('/scores/<int:record_id>/edit', methods=['GET', 'POST'])
@login_required  
def edit_score(record_id):
    record = LearningRecord.query.get_or_404(record_id)
    
    if request.method == 'POST':
        try:
            # í¼ ë°ì´í„° ë°›ê¸°
            date_str = request.form['date']
            
            # êµ­ì–´ ë°ì´í„°
            korean_problems_solved = int(request.form.get('korean_problems_correct', 0))
            korean_problems_correct = int(request.form.get('korean_problems_correct', 0))
            korean_last_page = int(request.form.get('korean_last_page', 0))
            
            # ìˆ˜í•™ ë°ì´í„°
            math_problems_solved = int(request.form.get('math_problems_correct', 0))
            math_problems_correct = int(request.form.get('math_problems_correct', 0))
            math_last_page = int(request.form.get('math_last_page', 0))
            
            # ë…ì„œ ë°ì´í„°
            reading_completed = 'reading_completed' in request.form
            reading_score = float(request.form.get('reading_score', 0))
            
            # ë‚ ì§œ ë³€í™˜
            from datetime import datetime
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # ì ìˆ˜ ê³„ì‚°
            korean_score = calculate_score(korean_problems_correct, korean_problems_solved) if korean_problems_solved > 0 else 0
            math_score = calculate_score(math_problems_correct, math_problems_solved) if math_problems_solved > 0 else 0
            
            # ì´ì  ê³„ì‚°
            total_score = korean_score + math_score + reading_score
            
            # ê¸°ë¡ ì—…ë°ì´íŠ¸
            record.date = date
            record.korean_problems_solved = korean_problems_solved
            record.korean_problems_correct = korean_problems_correct
            record.korean_score = korean_score
            record.korean_last_page = korean_last_page
            record.math_problems_solved = math_problems_solved
            record.math_problems_correct = math_problems_correct
            record.math_score = math_score
            record.math_last_page = math_last_page
            record.reading_completed = reading_completed
            record.reading_score = reading_score
            record.total_score = total_score
            record.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'{record.child.name} ì•„ë™ì˜ í•™ìŠµ ê¸°ë¡ì´ ìˆ˜ì •ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
            return redirect(url_for('child_detail', child_id=record.child_id))
            
        except ValueError:
            flash('ì…ë ¥ê°’ì„ ë‹¤ì‹œ í™•ì¸í•´ì£¼ì„¸ìš”.', 'error')
        except Exception as e:
            db.session.rollback()
            flash('ìˆ˜ì • ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤. ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”.', 'error')
    
    return render_template('scores/form.html', record=record, children=Child.query.all())

@app.route('/scores/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_score(record_id):
    record = LearningRecord.query.get_or_404(record_id)
    child_name = record.child.name
    child_id = record.child_id
    
    # ê¶Œí•œ í™•ì¸ (ì„¼í„°ì¥ê³¼ ëŒë´„ì„ ìƒë‹˜ë§Œ ì‚­ì œ ê°€ëŠ¥)
    if current_user.role not in ['ê°œë°œì']:
        flash('ì ìˆ˜ ê¸°ë¡ ì‚­ì œ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('child_detail', child_id=child_id))
    
    try:
        db.session.delete(record)
        db.session.commit()
        flash(f'{child_name} ì•„ë™ì˜ {record.date.strftime("%Y-%m-%d")} í•™ìŠµ ê¸°ë¡ì´ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
    except Exception as e:
        db.session.rollback()
        print(f"ğŸ› DEBUG: í•™ìŠµ ê¸°ë¡ ì‚­ì œ ì˜¤ë¥˜ - {str(e)}")
        print(f"ğŸ› DEBUG: record_id: {record_id}, child_id: {child_id}")
        flash(f'ì‚­ì œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}', 'error')
    
    return redirect(url_for('child_detail', child_id=child_id))

# í†µê³„ í¬í•¨/ì œì™¸ í† ê¸€
@app.route('/children/<int:child_id>/toggle_stats', methods=['POST'])
@login_required
def toggle_child_stats(child_id):
    child = Child.query.get_or_404(child_id)
    print(f"DEBUG: {child.name} - ì´ì „ ìƒíƒœ: {child.include_in_stats}")
    
    # ëª…í™•í•œ í† ê¸€ ë¡œì§
    if child.include_in_stats:
        child.include_in_stats = False
    else:
        child.include_in_stats = True
    
    print(f"DEBUG: {child.name} - ë³€ê²½ í›„ ìƒíƒœ: {child.include_in_stats}")
    db.session.commit()
    
    status = "í¬í•¨" if child.include_in_stats else "ì œì™¸"
    flash(f'{child.name} ì•„ì´ê°€ í†µê³„ì—ì„œ {status}ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
    return redirect(url_for('children_list'))

# ë…ì„œ ê¸°ë¡ ë¼ìš°íŠ¸
@app.route('/reading')
@login_required
def reading_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    grade_filter = request.args.get('grade', '', type=str)
    
    # ë…ì„œ ê¸°ë¡ ì¡°íšŒ (reading_scoreê°€ ìˆëŠ” ê¸°ë¡ë§Œ)
    query = db.session.query(LearningRecord, Child).join(Child).filter(LearningRecord.reading_score.isnot(None))
    
    # ê²€ìƒ‰ í•„í„°
    if search:
        query = query.filter(Child.name.contains(search))
    
    # í•™ë…„ í•„í„°
    if grade_filter:
        query = query.filter(Child.grade == int(grade_filter))
    
    # ìµœì‹ ìˆœ ì •ë ¬
    query = query.order_by(LearningRecord.date.desc(), LearningRecord.created_at.desc())
    
    # í˜ì´ì§€ë„¤ì´ì…˜
    records = query.paginate(
        page=page, per_page=15, error_out=False
    )
    
    # í•™ë…„ ëª©ë¡ (í•„í„°ìš©)
    grades = db.session.query(Child.grade).distinct().order_by(Child.grade).all()
    grade_list = [g[0] for g in grades]
    
    return render_template('reading/list.html', 
                         records=records, 
                         search=search,
                         grade_filter=grade_filter,
                         grade_list=grade_list)

# ê³¼ëª©ë³„ ë¹„êµ í†µê³„ í˜ì´ì§€
@app.route('/statistics')
@login_required
def statistics_overview():
    # í•™ë…„ë³„ í˜„ì¬ ì§„ë„ í˜„í™©
    grade_progress = {}
    
    for grade in range(1, 7):
        children = Child.query.filter_by(grade=grade, include_in_stats=True).all()
        if not children:
            continue
            
        # ê° ê³¼ëª©ë³„ ìµœì‹  í˜ì´ì§€ ì¡°íšŒ
        grade_progress[grade] = {
            'total_children': len(children),
            'korean_pages': {},
            'math_pages': {},
        }
        
        for child in children:
            latest_record = LearningRecord.query.filter_by(child_id=child.id).order_by(LearningRecord.date.desc()).first()
            if latest_record:
                # êµ­ì–´ í˜ì´ì§€ë³„ ì•„ì´ë“¤ ê·¸ë£¹í™”
                if latest_record.korean_last_page:
                    page = latest_record.korean_last_page
                    if page not in grade_progress[grade]['korean_pages']:
                        grade_progress[grade]['korean_pages'][page] = []
                    grade_progress[grade]['korean_pages'][page].append(child)
                
                # ìˆ˜í•™ í˜ì´ì§€ë³„ ì•„ì´ë“¤ ê·¸ë£¹í™”  
                if latest_record.math_last_page:
                    page = latest_record.math_last_page
                    if page not in grade_progress[grade]['math_pages']:
                        grade_progress[grade]['math_pages'][page] = []
                    grade_progress[grade]['math_pages'][page].append(child)
    
    return render_template('statistics/overview.html', grade_progress=grade_progress)

# íŠ¹ì • í˜ì´ì§€ë³„ ìƒì„¸ í†µê³„
@app.route('/statistics/<int:grade>/<subject>/<int:page>')
@login_required
def page_statistics(grade, subject, page):
    # í•´ë‹¹ í•™ë…„, ê³¼ëª©, í˜ì´ì§€ì˜ ëª¨ë“  ê¸°ë¡ ì¡°íšŒ
    children_in_grade = Child.query.filter_by(grade=grade, include_in_stats=True).all()
    child_ids = [child.id for child in children_in_grade]
    
    if subject == 'korean':
        records = LearningRecord.query.filter(
            LearningRecord.child_id.in_(child_ids),
            LearningRecord.korean_last_page == page
        ).order_by(LearningRecord.date.desc()).all()
        
        # êµ­ì–´ ì ìˆ˜ ê¸°ì¤€ìœ¼ë¡œ ì •ë ¬
        records_with_scores = []
        for record in records:
            score = calculate_score(record.korean_problems_correct, record.korean_problems_solved)
            records_with_scores.append({
                'record': record,
                'score': score,
                'child_name': record.child.name
            })
        records_with_scores.sort(key=lambda x: x['score'], reverse=True)
        
    elif subject == 'math':
        records = LearningRecord.query.filter(
            LearningRecord.child_id.in_(child_ids),
            LearningRecord.math_last_page == page
        ).order_by(LearningRecord.date.desc()).all()
        
        # ìˆ˜í•™ ì ìˆ˜ ê¸°ì¤€ìœ¼ë¡œ ì •ë ¬
        records_with_scores = []
        for record in records:
            score = calculate_score(record.math_problems_correct, record.math_problems_solved)
            records_with_scores.append({
                'record': record,
                'score': score,
                'child_name': record.child.name
            })
        records_with_scores.sort(key=lambda x: x['score'], reverse=True)
    
    # ì•„ì§ í•´ë‹¹ í˜ì´ì§€ë¥¼ í’€ì§€ ì•Šì€ ì•„ì´ë“¤
    completed_child_ids = [r['record'].child_id for r in records_with_scores]
    pending_children = [child for child in children_in_grade if child.id not in completed_child_ids]
    
    return render_template('statistics/page_detail.html', 
                         grade=grade, 
                         subject=subject, 
                         page=page,
                         records_with_scores=records_with_scores,
                         pending_children=pending_children,
                         total_children=len(children_in_grade))

# ì‹œê°í™” í†µê³„ í˜ì´ì§€ (ì§„ë„ ë° ì„±ì  ë¹„êµ)
@app.route('/statistics/charts')
@login_required
def statistics_charts():
    # ì˜¤ëŠ˜ ë‚ ì§œ
    today = datetime.utcnow().date()
    
    # ì˜¤ëŠ˜ í•™ìŠµ ê¸°ë¡ ê°€ì ¸ì˜¤ê¸°
    today_records = LearningRecord.query.filter(
        func.date(LearningRecord.created_at) == today
    ).order_by(LearningRecord.created_at.desc()).all()
    
    # í•™ë…„ë³„ë¡œ ì•„ì´ë“¤ ê·¸ë£¹í™”
    grade_progress_data = {}
    page_comparison_data = {}
    grade_average_progress = {}
    
    for grade in range(1, 7):
        children = Child.query.filter_by(grade=grade, include_in_stats=True).all()
        if not children:
            continue
        
        # ê° ì•„ì´ì˜ ìµœì‹  ì§„ë„ ì •ë³´ ê°€ì ¸ì˜¤ê¸°
        grade_students = []
        for child in children:
            # ìµœì‹  ê¸°ë¡ ê°€ì ¸ì˜¤ê¸°
            latest_record = LearningRecord.query.filter_by(
                child_id=child.id
            ).order_by(LearningRecord.date.desc()).first()
            
            korean_page = latest_record.korean_last_page if latest_record else 0
            math_page = latest_record.math_last_page if latest_record else 0

            grade_students.append({
                'name': child.name,
                'korean_page': korean_page,
                'math_page': math_page,
                'total_pages': korean_page + math_page
            })
        
        # ì´ ì§„ë„ìˆœìœ¼ë¡œ ì •ë ¬
        grade_students.sort(key=lambda x: x['total_pages'], reverse=True)
        grade_progress_data[str(grade)] = grade_students
        
        # ê°™ì€ í˜ì´ì§€ ì„±ì  ë¹„êµ ë°ì´í„°
        page_comparison_data[str(grade)] = {'korean': {}, 'math': {}}
        
        # êµ­ì–´ í˜ì´ì§€ë³„ ë¹„êµ
        korean_pages = {}
        for child in children:
            records = LearningRecord.query.filter_by(child_id=child.id).all()
            for record in records:
                if record.korean_last_page:
                    page = record.korean_last_page
                    if page not in korean_pages:
                        korean_pages[page] = []
                    korean_pages[page].append({
                        'name': child.name,
                        'score': record.korean_score,
                        'correct': record.korean_problems_correct,
                        'solved': record.korean_problems_solved
                    })
        
        # 2ëª… ì´ìƒì¸ í˜ì´ì§€ë§Œ í•„í„°ë§í•˜ê³  ì ìˆ˜ìˆœ ì •ë ¬
        for page, students in korean_pages.items():
            if len(students) >= 2:
                students.sort(key=lambda x: x['score'], reverse=True)
                page_comparison_data[str(grade)]['korean'][page] = students

        # ìˆ˜í•™ í˜ì´ì§€ë³„ ë¹„êµ
        math_pages = {}
        for child in children:
            records = LearningRecord.query.filter_by(child_id=child.id).all()
            for record in records:
                if record.math_last_page:
                    page = record.math_last_page
                    if page not in math_pages:
                        math_pages[page] = []
                    math_pages[page].append({
                        'name': child.name,
                        'score': record.math_score,
                        'correct': record.math_problems_correct,
                        'solved': record.math_problems_solved
                    })
        
        # 2ëª… ì´ìƒì¸ í˜ì´ì§€ë§Œ í•„í„°ë§í•˜ê³  ì ìˆ˜ìˆœ ì •ë ¬
        for page, students in math_pages.items():
            if len(students) >= 2:
                students.sort(key=lambda x: x['score'], reverse=True)
                page_comparison_data[str(grade)]['math'][page] = students

        # í•™ë…„ë³„ í‰ê·  ì§„ë„ ê³„ì‚°
        if grade_students:
            korean_avg_page = sum(s['korean_page'] for s in grade_students) / len(grade_students)
            math_avg_page = sum(s['math_page'] for s in grade_students) / len(grade_students)
            
            # í‰ê·  ì ìˆ˜ ê³„ì‚°
            korean_scores = []
            math_scores = []
            reading_scores = []
            
            for child in children:
                records = LearningRecord.query.filter_by(child_id=child.id).all()
                for record in records:
                    if record.korean_score > 0:
                        korean_scores.append(record.korean_score)
                    if record.math_score > 0:
                        math_scores.append(record.math_score)
                    if record.reading_score > 0:
                        reading_scores.append(record.reading_score)
            
            grade_average_progress[str(grade)] = {
                'korean_avg_page': round(korean_avg_page, 1),
                'math_avg_page': round(math_avg_page, 1),
                'korean_avg_score': round(sum(korean_scores) / len(korean_scores), 1) if korean_scores else 0,
                'math_avg_score': round(sum(math_scores) / len(math_scores), 1) if math_scores else 0,
                'reading_avg_score': round(sum(reading_scores) / len(reading_scores), 1) if reading_scores else 0
            }
    
    # ì „ì²´ ì§„ë„ ë¦¬ë”ë³´ë“œ (ìƒìœ„ 10ëª…)
    all_children = Child.query.filter_by(include_in_stats=True).all()
    all_students = []
    
    for child in all_children:
        latest_record = LearningRecord.query.filter_by(
            child_id=child.id
        ).order_by(LearningRecord.date.desc()).first()
        
        korean_page = latest_record.korean_last_page if latest_record else 0
        math_page = latest_record.math_last_page if latest_record else 0
        
        # í‰ê·  ì ìˆ˜ ê³„ì‚°
        all_records = LearningRecord.query.filter_by(child_id=child.id).all()
        total_score = 0
        record_count = 0
        for record in all_records:
            if record.total_score > 0:
                total_score += record.total_score
                record_count += 1
        avg_score = round(total_score / record_count, 1) if record_count > 0 else 0
        
        # ìµœê·¼ í•™ìŠµì¼
        last_study = latest_record.date.strftime('%m/%d') if latest_record else '-'

        all_students.append({
            'name': child.name,
            'grade': child.grade,
            'korean_page': korean_page,
            'math_page': math_page,
            'total_pages': korean_page + math_page,
            'avg_score': avg_score,
            'last_study': last_study
        })
    
    # ì´ ì§„ë„ìˆœìœ¼ë¡œ ì •ë ¬í•˜ê³  ìƒìœ„ 10ëª…ë§Œ
    all_students.sort(key=lambda x: x['total_pages'], reverse=True)
    progress_leaderboard = all_students[:10]

    return render_template('statistics/charts.html',
                         grade_progress_data=json.dumps(grade_progress_data),
                         page_comparison_data=json.dumps(page_comparison_data),
                         grade_average_progress=json.dumps(grade_average_progress),
                         progress_leaderboard=progress_leaderboard,
                         today_records=today_records)

# ë¦¬í¬íŠ¸ ë¼ìš°íŠ¸ë“¤
@app.route('/reports')
@login_required
def reports_overview():
    """ë¦¬í¬íŠ¸ ë©”ì¸ í˜ì´ì§€"""
    # í…ŒìŠ¤íŠ¸ì‚¬ìš©ìëŠ” ì ‘ê·¼ ë¶ˆê°€
    if current_user.role == 'í…ŒìŠ¤íŠ¸ì‚¬ìš©ì':
        flash('ë¦¬í¬íŠ¸ í˜ì´ì§€ì— ì ‘ê·¼í•  ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('dashboard'))
    from datetime import datetime, timedelta
    today = datetime.utcnow().date()
    return render_template('reports/overview.html', today=today, timedelta=timedelta)

@app.route('/reports/child/<int:child_id>')
@login_required
def child_report(child_id):
    """ê°œë³„ ì•„ë™ ë¦¬í¬íŠ¸"""
    child = Child.query.get_or_404(child_id)
    
    # ìµœê·¼ 30ì¼ê°„ì˜ í•™ìŠµ ê¸°ë¡
    thirty_days_ago = datetime.utcnow().date() - timedelta(days=30)
    recent_records = LearningRecord.query.filter_by(child_id=child_id).filter(
        LearningRecord.date >= thirty_days_ago
    ).order_by(LearningRecord.date.desc()).all()
    
    # í†µê³„ ê³„ì‚°
    total_records = len(recent_records)
    korean_records = [r for r in recent_records if r.korean_score > 0]
    math_records = [r for r in recent_records if r.math_score > 0]
    reading_records = [r for r in recent_records if r.reading_score > 0]
    
    # í‰ê·  ì ìˆ˜
    avg_korean = sum(r.korean_score for r in korean_records) / len(korean_records) if korean_records else 0
    avg_math = sum(r.math_score for r in math_records) / len(math_records) if math_records else 0
    avg_reading = sum(r.reading_score for r in reading_records) / len(reading_records) if reading_records else 0
    
    # ìµœì‹  ì§„ë„
    latest_record = recent_records[0] if recent_records else None
    current_korean_page = latest_record.korean_last_page if latest_record else 0
    current_math_page = latest_record.math_last_page if latest_record else 0
    
    # ì›”ë³„ í•™ìŠµ ì¼ìˆ˜
    monthly_activity = {}
    for record in recent_records:
        month_key = record.date.strftime('%Y-%m')
        if month_key not in monthly_activity:
            monthly_activity[month_key] = 0
        monthly_activity[month_key] += 1
    
    return render_template('reports/child_report.html',
                         child=child,
                         recent_records=recent_records,
                         total_records=total_records,
                         avg_korean=round(avg_korean, 1),
                         avg_math=round(avg_math, 1),
                         avg_reading=round(avg_reading, 1),
                         current_korean_page=current_korean_page,
                         current_math_page=current_math_page,
                         monthly_activity=monthly_activity)

@app.route('/reports/grade/<int:grade>')
@login_required
def grade_report(grade):
    """í•™ë…„ë³„ ë¦¬í¬íŠ¸"""
    children = Child.query.filter_by(grade=grade, include_in_stats=True).all()
    
    if not children:
        flash(f'{grade}í•™ë…„ì— ë“±ë¡ëœ ì•„ë™ì´ ì—†ìŠµë‹ˆë‹¤.', 'warning')
        return redirect(url_for('reports_overview'))
    
    # í•™ë…„ í†µê³„ ê³„ì‚°
    grade_stats = {
        'total_children': len(children),
        'avg_korean_page': 0,
        'avg_math_page': 0,
        'avg_korean_score': 0,
        'avg_math_score': 0,
        'avg_reading_score': 0,
        'children_data': []
    }
    
    korean_pages = []
    math_pages = []
    korean_scores = []
    math_scores = []
    reading_scores = []
    
    for child in children:
        latest_record = LearningRecord.query.filter_by(child_id=child.id).order_by(LearningRecord.date.desc()).first()
        
        if latest_record:
            korean_pages.append(latest_record.korean_last_page)
            math_pages.append(latest_record.math_last_page)
            
            if latest_record.korean_score > 0:
                korean_scores.append(latest_record.korean_score)
            if latest_record.math_score > 0:
                math_scores.append(latest_record.math_score)
            if latest_record.reading_score > 0:
                reading_scores.append(latest_record.reading_score)
        
        # ì•„ë™ë³„ ë°ì´í„°
        child_data = {
            'id': child.id,
            'name': child.name,
            'korean_page': latest_record.korean_last_page if latest_record else 0,
            'math_page': latest_record.math_last_page if latest_record else 0,
            'korean_score': latest_record.korean_score if latest_record else 0,
            'math_score': latest_record.math_score if latest_record else 0,
            'reading_score': latest_record.reading_score if latest_record else 0,
            'last_study': latest_record.date.strftime('%m/%d') if latest_record else '-'
        }
        grade_stats['children_data'].append(child_data)
    
    # í‰ê·  ê³„ì‚°
    if korean_pages:
        grade_stats['avg_korean_page'] = round(sum(korean_pages) / len(korean_pages), 1)
    if math_pages:
        grade_stats['avg_math_page'] = round(sum(math_pages) / len(math_pages), 1)
    if korean_scores:
        grade_stats['avg_korean_score'] = round(sum(korean_scores) / len(korean_scores), 1)
    if math_scores:
        grade_stats['avg_math_score'] = round(sum(math_scores) / len(math_scores), 1)
    if reading_scores:
        grade_stats['avg_reading_score'] = round(sum(reading_scores) / len(reading_scores), 1)
    
    return render_template('reports/grade_report.html',
                         grade=grade,
                         grade_stats=grade_stats)

@app.route('/reports/period')
@login_required
def period_report():
    """ê¸°ê°„ë³„ ë¦¬í¬íŠ¸"""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    if not start_date or not end_date:
        # ê¸°ë³¸ê°’: ì´ë²ˆ ë‹¬
        today = datetime.utcnow().date()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    
    # ë‚ ì§œ ë³€í™˜
    start = datetime.strptime(start_date, '%Y-%m-%d').date()
    end = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # ê¸°ê°„ ë‚´ ëª¨ë“  ê¸°ë¡ ì¡°íšŒ
    records = LearningRecord.query.filter(
        LearningRecord.date >= start,
        LearningRecord.date <= end
    ).order_by(LearningRecord.date.desc()).all()
    
    # í†µê³„ ê³„ì‚°
    total_records = len(records)
    total_children = Child.query.filter_by(include_in_stats=True).count()
    
    # ê³¼ëª©ë³„ í†µê³„
    korean_records = [r for r in records if r.korean_score > 0]
    math_records = [r for r in records if r.math_score > 0]
    reading_records = [r for r in records if r.reading_score > 0]
    
    # ë…ì„œ ì°¸ì—¬ ì•„ë™ ìˆ˜ (ì¤‘ë³µ ì œê±°)
    reading_children = len(set(r.child_id for r in reading_records))
    
    # ë””ë²„ê¹…ìš© ë¡œê·¸
    print(f"DEBUG: total_children = {total_children}")
    print(f"DEBUG: reading_children = {reading_children}")
    print(f"DEBUG: reading_records count = {len(reading_records)}")
    print(f"DEBUG: all_children = {Child.query.count()}")
    print(f"DEBUG: include_in_stats_children = {Child.query.filter_by(include_in_stats=True).count()}")
    
    period_stats = {
        'total_records': total_records,
        'total_children': total_children,
        'korean_count': len(korean_records),
        'math_count': len(math_records),
        'reading_count': reading_children,  # ë…ì„œ ì°¸ì—¬ ì•„ë™ ìˆ˜ë¡œ ë³€ê²½
        'avg_korean_score': round(sum(r.korean_score for r in korean_records) / len(korean_records), 1) if korean_records else 0,
        'avg_math_score': round(sum(r.math_score for r in math_records) / len(math_records), 1) if math_records else 0,
        'avg_reading_score': round(sum(r.reading_score for r in reading_records) / len(reading_records), 1) if reading_records else 0
    }
    
    return render_template('reports/period_report.html',
                         start_date=start_date,
                         end_date=end_date,
                         period_stats=period_stats,
                         records=records)


# ìƒˆë¡œìš´ í¬ì¸íŠ¸ ì‹œìŠ¤í…œ ë¼ìš°íŠ¸ë“¤
@app.route('/points')
@login_required
def points_list():
    """í¬ì¸íŠ¸ ê¸°ë¡ ëª©ë¡"""
    # ìµœê·¼ ì…ë ¥ëœ í¬ì¸íŠ¸ë“¤ (ì…ë ¥ ì‹œê°„ ê¸°ì¤€ìœ¼ë¡œ ì •ë ¬)
    points_records = DailyPoints.query.order_by(DailyPoints.created_at.desc()).limit(20).all()
    return render_template('points/list.html', points_records=points_records)

@app.route('/points/input/select')
@login_required
def points_input_select():
    """í¬ì¸íŠ¸ ì…ë ¥í•  ì•„ë™ ì„ íƒ í˜ì´ì§€"""
    return render_template('points/select.html')

@app.route('/points/input/<int:child_id>', methods=['GET', 'POST'])
@login_required
def points_input(child_id):
    """í¬ì¸íŠ¸ ì…ë ¥ í˜ì´ì§€"""
    child = Child.query.get_or_404(child_id)
    
    if request.method == 'POST':
        # ì˜¤ëŠ˜ ë‚ ì§œ
        today = datetime.utcnow().date()
        
        # ê¸°ì¡´ ê¸°ë¡ì´ ìˆëŠ”ì§€ í™•ì¸
        existing_record = DailyPoints.query.filter_by(
            child_id=child_id, 
            date=today
        ).first()
        
        try:
            # í¬ì¸íŠ¸ ê°’ ê°€ì ¸ì˜¤ê¸° ë° ê²€ì¦
            korean_points = int(request.form.get('korean_points', 0))
            math_points = int(request.form.get('math_points', 0))
            ssen_points = int(request.form.get('ssen_points', 0))
            reading_points = int(request.form.get('reading_points', 0))
        
            # ìƒˆ ê³¼ëª©ë“¤ (2025-09-17 ì¶”ê°€)
            piano_points = int(request.form.get('piano_points', 0))
            english_points = int(request.form.get('english_points', 0))
            advanced_math_points = int(request.form.get('advanced_math_points', 0))
            writing_points = int(request.form.get('writing_points', 0))

            # ìˆ˜ë™ í¬ì¸íŠ¸ (manual_historyì—ì„œ ì‹¤ì‹œê°„ ê³„ì‚°)
            manual_points = get_manual_points_from_history(existing_record)
        
            # ê°’ ê²€ì¦: ìŒìˆ˜ ë°©ì§€ë§Œ ë°©ì§€
            if any(points < 0 for points in [korean_points, math_points, ssen_points, reading_points, piano_points, english_points, advanced_math_points, writing_points]):
                flash('âŒ í¬ì¸íŠ¸ëŠ” ìŒìˆ˜ì¼ ìˆ˜ ì—†ìŠµë‹ˆë‹¤. 0 ì´ìƒì˜ ê°’ì„ ì…ë ¥í•´ì£¼ì„¸ìš”.', 'error')
                return redirect(url_for('points_input', child_id=child_id))
            
            # ì´ í¬ì¸íŠ¸ ê³„ì‚° (ê²€ì¦ëœ ê°’ìœ¼ë¡œ)
            total_points = korean_points + math_points + ssen_points + reading_points + piano_points + english_points + advanced_math_points + writing_points + manual_points
            
            # ê³„ì‚° ê²°ê³¼ ê²€ì¦
            expected_total = sum([korean_points, math_points, ssen_points, reading_points, piano_points, english_points, advanced_math_points, writing_points, manual_points])
            if total_points != expected_total:
                flash(f'âŒ í¬ì¸íŠ¸ ê³„ì‚° ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤. ì˜ˆìƒ: {expected_total}, ê³„ì‚°: {total_points}', 'error')
                return redirect(url_for('points_input', child_id=child_id))
            
            if existing_record:
                # ê¸°ì¡´ ê¸°ë¡ ì—…ë°ì´íŠ¸ (ë³€ê²½ ì´ë ¥ ê¸°ë¡)
                old_total = existing_record.total_points
                old_korean = existing_record.korean_points
                old_math = existing_record.math_points
                old_ssen = existing_record.ssen_points
                old_reading = existing_record.reading_points
                old_piano = existing_record.piano_points
                old_english = existing_record.english_points
                old_advanced_math = existing_record.advanced_math_points
                old_writing = existing_record.writing_points
                old_manual = existing_record.manual_points
                
                # ê¸°ì¡´ ê¸°ë¡ ì—…ë°ì´íŠ¸
                existing_record.korean_points = korean_points
                existing_record.math_points = math_points
                existing_record.ssen_points = ssen_points
                existing_record.reading_points = reading_points
                existing_record.piano_points = piano_points
                existing_record.english_points = english_points
                existing_record.advanced_math_points = advanced_math_points
                existing_record.writing_points = writing_points
                # existing_record.manual_points = manual_points  # ì œê±°: manual_historyì—ì„œ ì‹¤ì‹œê°„ ê³„ì‚°
                existing_record.total_points = total_points
                existing_record.updated_at = datetime.utcnow()
                
                # ë³€ê²½ ì´ë ¥ ê¸°ë¡ (PointsHistory í…Œì´ë¸”) - ë³€ê²½ì‚¬í•­ì´ ìˆì„ ë•Œë§Œ
                if (old_korean != korean_points or old_math != math_points or 
                    old_ssen != ssen_points or old_reading != reading_points or old_piano != piano_points or old_english != english_points or old_advanced_math != advanced_math_points or old_writing != writing_points or old_manual != manual_points):
                    
                    history_record = PointsHistory(
                        child_id=child_id,
                        date=today,
                        old_korean_points=old_korean,
                        old_math_points=old_math,
                        old_ssen_points=old_ssen,
                        old_reading_points=old_reading,
                        old_total_points=old_total,
                        new_korean_points=korean_points,
                        new_math_points=math_points,
                        new_ssen_points=ssen_points,
                        new_reading_points=reading_points,
                        new_total_points=total_points,
                        change_type='update',
                        changed_by=current_user.id,
                        change_reason='ì›¹ UIë¥¼ í†µí•œ í¬ì¸íŠ¸ ìˆ˜ì •'
                    )
                    db.session.add(history_record)
                    
                    # ë³€ê²½ ì´ë ¥ ê¸°ë¡ (ê°„ë‹¨í•œ ë¡œê·¸)
                    print(f"ğŸ“ í¬ì¸íŠ¸ ë³€ê²½ ì´ë ¥ - {child.name}({child.grade}í•™ë…„) - {today}")
                    print(f"  êµ­ì–´: {old_korean} â†’ {korean_points}")
                    print(f"  ìˆ˜í•™: {old_math} â†’ {math_points}")
                    print(f"  ìˆìˆ˜í•™: {old_ssen} â†’ {ssen_points}")
                    print(f"  ë…ì„œ: {old_reading} â†’ {reading_points}")
                    print(f"  ì´ì : {old_total} â†’ {total_points}")
                    print(f"  ë³€ê²½ì: {current_user.username}")
                
                # ëˆ„ì  í¬ì¸íŠ¸ ìë™ ì—…ë°ì´íŠ¸ (ì»¤ë°‹ ì—†ì´)
                update_cumulative_points(child_id, commit=False)
                
                # ëª¨ë“  ë³€ê²½ì‚¬í•­ì„ í•œ ë²ˆì— ì»¤ë°‹
                db.session.commit()
                
                # ì‹¤ì‹œê°„ ë°±ì—… í˜¸ì¶œ (ë°±ì—… ì‹¤íŒ¨ê°€ í¬ì¸íŠ¸ ì…ë ¥ì— ì˜í–¥ ì£¼ì§€ ì•Šë„ë¡)
                try:
                    realtime_backup(child_id, "update")
                except Exception as backup_error:
                    print(f"ë°±ì—… ì‹¤íŒ¨: {backup_error}")
                    # ë°±ì—… ì‹¤íŒ¨ëŠ” í¬ì¸íŠ¸ ì…ë ¥ ì„±ê³µì— ì˜í–¥ì„ ì£¼ì§€ ì•ŠìŒ
                
                flash(f'âœ… {child.name} ì•„ì´ì˜ í¬ì¸íŠ¸ê°€ ìˆ˜ì •ë˜ì—ˆìŠµë‹ˆë‹¤. (ì´ì : {total_points}ì )', 'success')
                return redirect(url_for('points_list'))
            else:
                # ìƒˆ ê¸°ë¡ ìƒì„± (ìƒì„± ì´ë ¥ ê¸°ë¡)
                history_record = PointsHistory(
                    child_id=child_id,
                    date=today,
                    old_korean_points=0,
                    old_math_points=0,
                    old_ssen_points=0,
                    old_reading_points=0,
                    old_total_points=0,
                    new_korean_points=korean_points,
                    new_math_points=math_points,
                    new_ssen_points=ssen_points,
                    new_reading_points=reading_points,
                    new_total_points=total_points,
                    change_type='create',
                    changed_by=current_user.id,
                    change_reason='ì›¹ UIë¥¼ í†µí•œ í¬ì¸íŠ¸ ì‹ ê·œ ì…ë ¥'
                )
                db.session.add(history_record)
                
                # ìƒˆ ê¸°ë¡ ìƒì„±
                new_record = DailyPoints(
                    child_id=child_id,
                    date=today,
                    korean_points=korean_points,
                    math_points=math_points,
                    ssen_points=ssen_points,
                    reading_points=reading_points,
                piano_points=piano_points,
                english_points=english_points,
                advanced_math_points=advanced_math_points,
                writing_points=writing_points,
                    total_points=total_points,
                    created_by=current_user.id
                )
                db.session.add(new_record)
                
                # ëˆ„ì  í¬ì¸íŠ¸ ìë™ ì—…ë°ì´íŠ¸ (ì»¤ë°‹ ì—†ì´)
                update_cumulative_points(child_id, commit=False)
                
                # ëª¨ë“  ë³€ê²½ì‚¬í•­ì„ í•œ ë²ˆì— ì»¤ë°‹
                db.session.commit()
                
                # ì‹¤ì‹œê°„ ë°±ì—… í˜¸ì¶œ (ë°±ì—… ì‹¤íŒ¨ê°€ í¬ì¸íŠ¸ ì…ë ¥ì— ì˜í–¥ ì£¼ì§€ ì•Šë„ë¡)
                try:
                    realtime_backup(child_id, "create")
                except Exception as backup_error:
                    print(f"ë°±ì—… ì‹¤íŒ¨: {backup_error}")
                    # ë°±ì—… ì‹¤íŒ¨ëŠ” í¬ì¸íŠ¸ ì…ë ¥ ì„±ê³µì— ì˜í–¥ì„ ì£¼ì§€ ì•ŠìŒ
                
                flash(f'âœ… {child.name} ì•„ì´ì˜ í¬ì¸íŠ¸ê°€ ì €ì¥ë˜ì—ˆìŠµë‹ˆë‹¤. (ì´ì : {total_points}ì )', 'success')
                return redirect(url_for('points_list'))
            
        except ValueError as e:
            flash('âŒ ì˜ëª»ëœ í¬ì¸íŠ¸ ê°’ì´ ì…ë ¥ë˜ì—ˆìŠµë‹ˆë‹¤. ìˆ«ìë§Œ ì…ë ¥í•´ì£¼ì„¸ìš”.', 'error')
            return redirect(url_for('points_input', child_id=child_id))
        except Exception as e:
            db.session.rollback()
            flash(f'âŒ í¬ì¸íŠ¸ ì €ì¥ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}', 'error')
            return redirect(url_for('points_input', child_id=child_id))
    
    # ì˜¤ëŠ˜ ê¸°ë¡ ê°€ì ¸ì˜¤ê¸°
    today = datetime.utcnow().date()
    today_record = DailyPoints.query.filter_by(
        child_id=child_id, 
        date=today
    ).first()
    
    # ì˜¤ëŠ˜ ë‚ ì§œ ë¬¸ìì—´ ê³„ì‚°
    today_date = datetime.utcnow().strftime('%Yë…„ %mì›” %dì¼')
    
    return render_template('points/input.html', child=child, today_record=today_record, today_date=today_date)

def update_cumulative_points(child_id, commit=True):
    """ì•„ë™ì˜ ëˆ„ì  í¬ì¸íŠ¸ë¥¼ ìë™ìœ¼ë¡œ ì—…ë°ì´íŠ¸"""
    try:
        # í•´ë‹¹ ì•„ë™ì˜ ëª¨ë“  ì¼ì¼ í¬ì¸íŠ¸ í•©ê³„ ê³„ì‚°
        total_cumulative = db.session.query(
            db.func.sum(DailyPoints.total_points)
        ).filter_by(child_id=child_id).scalar() or 0
        
        # Child ëª¨ë¸ì˜ cumulative_points ì—…ë°ì´íŠ¸
        child = Child.query.get(child_id)
        if child:
            child.cumulative_points = total_cumulative
            if commit:
                db.session.commit()
            print(f"ğŸ“Š {child.name}ì˜ ëˆ„ì  í¬ì¸íŠ¸ ì—…ë°ì´íŠ¸: {total_cumulative}ì ")
            return total_cumulative
            
    except Exception as e:
        print(f"âŒ ëˆ„ì  í¬ì¸íŠ¸ ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
        if commit:
            db.session.rollback()
        raise e

@app.route('/points/statistics')
@login_required
def points_statistics():
    """í¬ì¸íŠ¸ í†µê³„ í˜ì´ì§€"""
    # ì˜¤ëŠ˜ ë‚ ì§œ
    today = datetime.utcnow().date()
    
    # í•™ë…„ë³„ í¬ì¸íŠ¸ í†µê³„
    grade_stats = {}
    for grade in range(1, 7):  # 1í•™ë…„~6í•™ë…„
        children = Child.query.filter_by(grade=grade, include_in_stats=True).all()
        if not children:
            continue
            
        grade_points = []
        for child in children:
            # ì˜¤ëŠ˜ í¬ì¸íŠ¸ ê¸°ë¡
            today_points = DailyPoints.query.filter_by(
                child_id=child.id, 
                date=today
            ).first()
            
            if today_points:
                grade_points.append(today_points.total_points)
        
        if grade_points:
            grade_stats[grade] = {
                'avg_points': round(sum(grade_points) / len(grade_points), 1),
                'max_points': max(grade_points),
                'min_points': min(grade_points),
                'total_children': len(children),
                'participated_children': len(grade_points)
            }
    
    return render_template('points/statistics.html', grade_stats=grade_stats, today=today)

@app.route('/points/analysis')
@login_required
def points_analysis():
    """í¬ì¸íŠ¸ ë¶„ì„ í˜ì´ì§€ - ì•„ë™ë³„ ìƒì„¸ ë¶„ì„"""
    # ì•„ë™ ì„ íƒ íŒŒë¼ë¯¸í„°
    child_id = request.args.get('child_id', type=int)
    
    if child_id:
        # íŠ¹ì • ì•„ë™ ë¶„ì„
        child = Child.query.get_or_404(child_id)
        
        # í•´ë‹¹ ì•„ë™ì˜ ì „ì²´ í¬ì¸íŠ¸ ê¸°ë¡ (ì¤‘ë³µ ì œê±° í›„)
        # ë‚ ì§œë³„ë¡œ í•˜ë‚˜ì˜ ê¸°ë¡ë§Œ ê°€ì ¸ì˜¤ê¸°
        result = db.session.execute(text("""
            SELECT id, date, korean_points, math_points, ssen_points, reading_points, total_points
            FROM daily_points 
            WHERE child_id = :child_id 
            AND id IN (
                SELECT MAX(id) 
                FROM daily_points 
                WHERE child_id = :child_id 
                GROUP BY date 
            )
            ORDER BY date DESC
        """), {"child_id": child_id})
        
        # ì‹¤ì œ DailyPoints ê°ì²´ë¡œ ë³€í™˜
        child_points = []
        for row in result:
            # ë‚ ì§œ íƒ€ì… ë³€í™˜ (ë¬¸ìì—´ì¼ ê²½ìš° datetime.dateë¡œ ë³€í™˜)
            date_value = row[1]
            if isinstance(date_value, str):
                from datetime import datetime
                date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
            
            # DailyPoints ê°ì²´ ìƒì„±
            point_record = DailyPoints(
                id=row[0],
                date=date_value,
                korean_points=row[2],
                math_points=row[3],
                ssen_points=row[4],
                reading_points=row[5],
                total_points=row[6]
            )
            child_points.append(point_record)
        
        # ì´ í¬ì¸íŠ¸ ê³„ì‚° (ì¤‘ë³µ ì œê±°ëœ ë°ì´í„°ë¡œ)
        total_points = sum(record.total_points for record in child_points)
        
        # ë””ë²„ê¹…: ì‹¤ì œ ë°ì´í„° í™•ì¸
        print(f"=== {child.name} í¬ì¸íŠ¸ ë¶„ì„ ===")
        print(f"ì•„ë™ ID: {child_id}")
        print(f"ì•„ë™ ì´ë¦„: {child.name}")
        print(f"ì´ ê¸°ë¡ ìˆ˜: {len(child_points)}")
        print(f"ê³„ì‚°ëœ ì´ í¬ì¸íŠ¸: {total_points}")
        print(f"Child.cumulative_points: {child.cumulative_points}")
        print("================================")
        
        # ê°™ì€ í•™ë…„ ì•„ë™ë“¤ì˜ í¬ì¸íŠ¸ ë¹„êµ (ì¤‘ë³µ ì œê±° í›„)
        same_grade_children = Child.query.filter_by(grade=child.grade, include_in_stats=True).all()
        grade_comparison = []
        
        for grade_child in same_grade_children:
            if grade_child.id != child_id:  # ìê¸° ìì‹  ì œì™¸
                # ì¤‘ë³µ ì œê±°ëœ í¬ì¸íŠ¸ ê³„ì‚°
                result = db.session.execute(text("""
                    SELECT SUM(total_points) as total, COUNT(*) as count
                    FROM daily_points 
                    WHERE child_id = :child_id 
                    AND id IN (
                        SELECT MAX(id) 
                        FROM daily_points 
                        WHERE child_id = :child_id 
                        GROUP BY date
                    )
                """), {"child_id": grade_child.id})
                
                row = result.fetchone()
                grade_child_total = row[0] or 0
                record_count = row[1] or 0
                
                grade_comparison.append({
                    'id': grade_child.id,
                    'name': grade_child.name,
                    'total_points': grade_child_total,
                    'record_count': record_count
                })
        
        # í•™ë…„ ë‚´ ìˆœìœ„ ê³„ì‚°
        grade_comparison.append({
            'id': child.id,
            'name': child.name,
            'total_points': total_points,
            'record_count': len(child_points)
        })
        grade_comparison.sort(key=lambda x: x['total_points'], reverse=True)
        
        # ì „ì²´ í•™ë…„ ìˆœìœ„ (ì¤‘ë³µ ì œê±° í›„)
        all_children = Child.query.filter_by(include_in_stats=True).all()
        overall_ranking = []
        
        for all_child in all_children:
            # ì¤‘ë³µ ì œê±°ëœ í¬ì¸íŠ¸ ê³„ì‚°
            result = db.session.execute(text("""
                SELECT SUM(total_points) as total, COUNT(*) as count
                FROM daily_points 
                WHERE child_id = :child_id 
                AND id IN (
                    SELECT MAX(id) 
                    FROM daily_points 
                    WHERE child_id = :child_id 
                    GROUP BY date
                )
            """), {"child_id": all_child.id})
            
            row = result.fetchone()
            all_child_total = row[0] or 0
            record_count = row[1] or 0
            
            overall_ranking.append({
                'id': all_child.id,
                'name': all_child.name,
                'grade': all_child.grade,
                'total_points': all_child_total,
                'record_count': record_count
            })
        
        overall_ranking.sort(key=lambda x: x['total_points'], reverse=True)
        
        return render_template('points/analysis.html', 
                             child=child,
                             child_points=child_points,
                             total_points=total_points,
                             grade_comparison=grade_comparison,
                             overall_ranking=overall_ranking)
    else:
        # ì•„ë™ ëª©ë¡ í‘œì‹œ
        children = Child.query.filter_by(include_in_stats=True).order_by(Child.grade, Child.name).all()
        return render_template('points/analysis.html', children=children)

@app.route('/points/visualization')
@login_required
def points_visualization():
    """í¬ì¸íŠ¸ ì‹œê°í™” í˜ì´ì§€"""
    from datetime import datetime, timedelta
    import calendar
    
    today = datetime.utcnow().date()
    
    # 1. ì£¼ê°„ íŠ¸ë Œë“œ (ìµœê·¼ 4ì£¼) - ë°°ì¹˜ ì¿¼ë¦¬ë¡œ ìµœì í™”
    date_range = [today - timedelta(days=i) for i in range(28, -1, -1)]
    weekly_points = DailyPoints.query.filter(
        DailyPoints.date.in_(date_range)
    ).all()
    
    weekly_data = []
    for date in date_range:
        day_points = [p for p in weekly_points if p.date == date]
        total_points = sum(record.total_points for record in day_points)
        weekly_data.append({
            'date': date.strftime('%m/%d'),
            'points': total_points
        })
    
    # 2. ì›”ë³„ í•©ê³„ (ì˜¬í•´ ì „ì²´) - ë°°ì¹˜ ì¿¼ë¦¬ë¡œ ìµœì í™”
    year_start = datetime(today.year, 1, 1).date()
    year_end = datetime(today.year, 12, 31).date()
    all_year_points = DailyPoints.query.filter(
        DailyPoints.date >= year_start,
        DailyPoints.date <= year_end
    ).all()
    
    monthly_data = []
    for month in range(1, 13):
        month_start = datetime(today.year, month, 1).date()
        if month == 12:
            month_end = datetime(today.year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(today.year, month + 1, 1).date() - timedelta(days=1)
        
        month_points = [p for p in all_year_points 
                       if month_start <= p.date <= month_end]
        total_month_points = sum(record.total_points for record in month_points)
        
        monthly_data.append({
            'month': f'{month}ì›”',
            'points': total_month_points
        })
    
    # 3. ê³¼ëª©ë³„ ë¶„í¬ (ì „ì²´ ê¸°ê°„) - ê¸°ì¡´ all_points ì¬ì‚¬ìš©
    subject_totals = {
        'êµ­ì–´': sum(record.korean_points for record in all_year_points),
        'ìˆ˜í•™': sum(record.math_points for record in all_year_points),
        'ìˆìˆ˜í•™': sum(record.ssen_points for record in all_year_points),
        'ë…ì„œ': sum(record.reading_points for record in all_year_points)
    }
    
    # 4. í•™ë…„ë³„ í‰ê·  - ë°°ì¹˜ ì¿¼ë¦¬ë¡œ ìµœì í™”
    all_children = Child.query.filter_by(include_in_stats=True).all()
    child_ids = [child.id for child in all_children]
    all_children_points = DailyPoints.query.filter(
        DailyPoints.child_id.in_(child_ids)
    ).all() if child_ids else []
    
    grade_averages = {}
    for grade_num in [1, 2, 3, 4, 5, 6]:
        grade_str = f'{grade_num}í•™ë…„'
        grade_children = [child for child in all_children if child.grade == grade_num]
        
        if grade_children:
            grade_child_ids = [child.id for child in grade_children]
            grade_points = [p for p in all_children_points if p.child_id in grade_child_ids]
            
            grade_total_points = sum(record.total_points for record in grade_points)
            grade_total_records = len(grade_points)
            
            if grade_total_records > 0:
                # í‰ê·  = ì´ í¬ì¸íŠ¸ / ì´ ê¸°ë¡ ìˆ˜ (ê° ê¸°ë¡ë‹¹ í‰ê·  í¬ì¸íŠ¸)
                grade_averages[grade_str] = round(grade_total_points / grade_total_records, 1)
                print(f"DEBUG: {grade_str} - ì´ í¬ì¸íŠ¸: {grade_total_points}, ì´ ê¸°ë¡: {grade_total_records}, í‰ê· : {grade_averages[grade_str]}")
            else:
                grade_averages[grade_str] = 0
                print(f"DEBUG: {grade_str} - ê¸°ë¡ ì—†ìŒ")
        else:
            grade_averages[grade_str] = 0
            print(f"DEBUG: {grade_str} - ì•„ë™ ì—†ìŒ")
    
    return render_template('points/visualization.html', 
                         weekly_data=weekly_data,
                         monthly_data=monthly_data,
                         subject_totals=subject_totals,
                         grade_averages=grade_averages,
                         today=today,
                         # JSON í˜•íƒœë¡œ ë¯¸ë¦¬ ë³€í™˜
                         weekly_labels=[item['date'] for item in weekly_data],
                         weekly_points=[item['points'] for item in weekly_data],
                         monthly_labels=[item['month'] for item in monthly_data],
                         monthly_points=[item['points'] for item in monthly_data],
                         subject_labels=list(subject_totals.keys()),
                         subject_values=list(subject_totals.values()),
                         grade_labels=list(grade_averages.keys()),
                         grade_values=list(grade_averages.values()))

@app.route('/points/child/<int:child_id>')
@login_required
def child_point_analysis(child_id):
    """ê°œë³„ ì•„ë™ í¬ì¸íŠ¸ ë¶„ì„ í˜ì´ì§€"""
    from datetime import datetime, timedelta
    import calendar
    
    child = Child.query.get_or_404(child_id)
    today = datetime.utcnow().date()
    
    # 1. ì£¼ê°„ í¬ì¸íŠ¸ íŠ¸ë Œë“œ (ìµœê·¼ 8ì£¼)
    weekly_data = []
    for i in range(56, -1, -1):  # ìµœê·¼ 8ì£¼ (56ì¼)
        date = today - timedelta(days=i)
        daily_record = DailyPoints.query.filter_by(child_id=child_id, date=date).first()
        points = daily_record.total_points if daily_record else 0
        weekly_data.append({
            'date': date.strftime('%m/%d'),
            'points': points
        })
    
    # 2. ì›”ê°„ í¬ì¸íŠ¸ í•©ê³„ (ìµœê·¼ 6ê°œì›”)
    monthly_data = []
    for i in range(6):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        if month_date.month == 12:
            month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
        
        month_points = DailyPoints.query.filter(
            DailyPoints.child_id == child_id,
            DailyPoints.date >= month_start,
            DailyPoints.date <= month_end
        ).all()
        total_month_points = sum(record.total_points for record in month_points)
        
        monthly_data.append({
            'month': month_date.strftime('%Yë…„ %mì›”'),
            'points': total_month_points
        })
    
    # 3. ì¦ê°ë¥  ê³„ì‚°
    # ì´ë²ˆ ì£¼ vs ì§€ë‚œ ì£¼
    this_week_start = today - timedelta(days=today.weekday())
    last_week_start = this_week_start - timedelta(days=7)
    
    this_week_points = DailyPoints.query.filter(
        DailyPoints.child_id == child_id,
        DailyPoints.date >= this_week_start,
        DailyPoints.date <= today
    ).all()
    this_week_total = sum(record.total_points for record in this_week_points)
    
    last_week_points = DailyPoints.query.filter(
        DailyPoints.child_id == child_id,
        DailyPoints.date >= last_week_start,
        DailyPoints.date < this_week_start
    ).all()
    last_week_total = sum(record.total_points for record in last_week_points)
    
    weekly_change = 0
    if last_week_total > 0:
        weekly_change = round(((this_week_total - last_week_total) / last_week_total) * 100, 1)
    
    # ì´ë²ˆ ë‹¬ vs ì§€ë‚œ ë‹¬
    this_month_start = today.replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    
    this_month_points = DailyPoints.query.filter(
        DailyPoints.child_id == child_id,
        DailyPoints.date >= this_month_start,
        DailyPoints.date <= today
    ).all()
    this_month_total = sum(record.total_points for record in this_month_points)
    
    last_month_points = DailyPoints.query.filter(
        DailyPoints.child_id == child_id,
        DailyPoints.date >= last_month_start,
        DailyPoints.date < this_month_start
    ).all()
    last_month_total = sum(record.total_points for record in last_month_points)
    
    monthly_change = 0
    if last_month_total > 0:
        monthly_change = round(((this_month_total - last_month_total) / last_month_total) * 100, 1)
    
    # 4. ê°™ì€ í•™ë…„ ë¹„êµ
    grade_children = Child.query.filter_by(grade=child.grade, include_in_stats=True).all()
    grade_comparison = []
    
    for grade_child in grade_children:
        child_total_points = DailyPoints.query.filter_by(child_id=grade_child.id).all()
        total_points = sum(record.total_points for record in child_total_points)
        record_count = len(child_total_points)
        
        grade_comparison.append({
            'id': grade_child.id,
            'name': grade_child.name,
            'total_points': total_points,
            'record_count': record_count,
            'avg_points': round(total_points / record_count, 0) if record_count > 0 else 0
        })
    
    # í¬ì¸íŠ¸ ìˆœìœ¼ë¡œ ì •ë ¬
    grade_comparison.sort(key=lambda x: x['total_points'], reverse=True)
    
    # í˜„ì¬ ì•„ë™ì˜ ìˆœìœ„ ì°¾ê¸°
    current_rank = 1
    for i, comp_child in enumerate(grade_comparison):
        if comp_child['id'] == child_id:
            current_rank = i + 1
            break
    
    return render_template('points/child_analysis.html',
                         child=child,
                         weekly_data=weekly_data,
                         monthly_data=monthly_data,
                         weekly_change=weekly_change,
                         monthly_change=monthly_change,
                         this_week_total=this_week_total,
                         last_week_total=last_week_total,
                         this_month_total=this_month_total,
                         last_month_total=last_month_total,
                         grade_comparison=grade_comparison,
                         current_rank=current_rank,
                         total_children_in_grade=len(grade_comparison))

@app.route('/points/grade-comparison/<int:grade>')
@login_required
def grade_point_comparison(grade):
    """í•™ë…„ë³„ í¬ì¸íŠ¸ ë¹„êµ ì‹œê°í™”"""
    from datetime import datetime, timedelta
    
    today = datetime.utcnow().date()
    
    # í•´ë‹¹ í•™ë…„ì˜ ëª¨ë“  ì•„ë™ ì¡°íšŒ
    grade_children = Child.query.filter_by(grade=grade, include_in_stats=True).all()
    
    if not grade_children:
        flash(f'{grade}í•™ë…„ì— ì•„ë™ì´ ì—†ìŠµë‹ˆë‹¤.', 'warning')
        return redirect(url_for('points_visualization'))
    
    # ê° ì•„ë™ì˜ í¬ì¸íŠ¸ ë°ì´í„° ìˆ˜ì§‘
    children_data = []
    for child in grade_children:
        # ì „ì²´ í¬ì¸íŠ¸
        all_points = DailyPoints.query.filter_by(child_id=child.id).all()
        total_points = sum(record.total_points for record in all_points)
        
        # ì´ë²ˆ ì£¼ í¬ì¸íŠ¸
        this_week_start = today - timedelta(days=today.weekday())
        this_week_points = DailyPoints.query.filter(
            DailyPoints.child_id == child.id,
            DailyPoints.date >= this_week_start,
            DailyPoints.date <= today
        ).all()
        this_week_total = sum(record.total_points for record in this_week_points)
        
        # ì´ë²ˆ ë‹¬ í¬ì¸íŠ¸
        this_month_start = today.replace(day=1)
        this_month_points = DailyPoints.query.filter(
            DailyPoints.child_id == child.id,
            DailyPoints.date >= this_month_start,
            DailyPoints.date <= today
        ).all()
        this_month_total = sum(record.total_points for record in this_month_points)
        
        # í‰ê·  í¬ì¸íŠ¸
        avg_points = round(total_points / len(all_points), 1) if all_points else 0
        
        children_data.append({
            'id': child.id,
            'name': child.name,
            'total_points': total_points,
            'this_week': this_week_total,
            'this_month': this_month_total,
            'avg_points': avg_points,
            'record_count': len(all_points)
        })
    
    # ì´ í¬ì¸íŠ¸ ìˆœìœ¼ë¡œ ì •ë ¬
    children_data.sort(key=lambda x: x['total_points'], reverse=True)
    
    # ì°¨íŠ¸ ë°ì´í„° ì¤€ë¹„
    chart_labels = [child['name'] for child in children_data]
    chart_total_points = [child['total_points'] for child in children_data]
    chart_weekly_points = [child['this_week'] for child in children_data]
    chart_monthly_points = [child['this_month'] for child in children_data]
    chart_avg_points = [child['avg_points'] for child in children_data]
    
    return render_template('points/grade_comparison.html',
                         grade=grade,
                         children_data=children_data,
                         chart_labels=chart_labels,
                         chart_total_points=chart_total_points,
                         chart_weekly_points=chart_weekly_points,
                         chart_monthly_points=chart_monthly_points,
                         chart_avg_points=chart_avg_points)

# ì„¤ì • ë¼ìš°íŠ¸ë“¤
@app.route('/settings')
@login_required
def settings():
    """ì„¤ì • ë©”ì¸ í˜ì´ì§€"""
    # í…ŒìŠ¤íŠ¸ì‚¬ìš©ìëŠ” ì ‘ê·¼ ë¶ˆê°€
    if current_user.role == 'í…ŒìŠ¤íŠ¸ì‚¬ìš©ì':
        flash('ì„¤ì • í˜ì´ì§€ì— ì ‘ê·¼í•  ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('settings/index.html')

@app.route('/settings/users', methods=['GET', 'POST'])
@login_required
def settings_users():
    """ì‚¬ìš©ì ê´€ë¦¬ í˜ì´ì§€"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'update_info':
            # ì‚¬ìš©ì ì •ë³´ ì—…ë°ì´íŠ¸
            new_username = request.form.get('username')
            new_name = request.form.get('name')
            
            # ì¤‘ë³µ í™•ì¸
            existing_user = User.query.filter_by(username=new_username).first()
            if existing_user and existing_user.id != current_user.id:
                flash('ì´ë¯¸ ì‚¬ìš© ì¤‘ì¸ ì•„ì´ë””ì…ë‹ˆë‹¤.', 'error')
                return redirect(url_for('settings_users'))
            
            current_user.username = new_username
            current_user.name = new_name
            db.session.commit()
            flash('ì‚¬ìš©ì ì •ë³´ê°€ ì—…ë°ì´íŠ¸ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
            
        elif action == 'change_password':
            # ë¹„ë°€ë²ˆí˜¸ ë³€ê²½
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            if not check_password_hash(current_user.password_hash, current_password):
                flash('í˜„ì¬ ë¹„ë°€ë²ˆí˜¸ê°€ ì˜¬ë°”ë¥´ì§€ ì•ŠìŠµë‹ˆë‹¤.', 'error')
                return redirect(url_for('settings_users'))
            
            if new_password != confirm_password:
                flash('ìƒˆ ë¹„ë°€ë²ˆí˜¸ê°€ ì¼ì¹˜í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤.', 'error')
                return redirect(url_for('settings_users'))
            
            current_user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            flash('ë¹„ë°€ë²ˆí˜¸ê°€ ë³€ê²½ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
            
        elif action == 'add_user':
            # ìƒˆ ì‚¬ìš©ì ì¶”ê°€ (ê°œë°œìë§Œ ê°€ëŠ¥)
            if current_user.role != 'ê°œë°œì':
                flash('ìƒˆ ì‚¬ìš©ì ì¶”ê°€ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
                return redirect(url_for('settings_users'))
            
            new_username = request.form.get('new_username')
            new_name = request.form.get('new_name')
            new_role = request.form.get('new_role')
            new_password = request.form.get('new_password')
            
            # ì¤‘ë³µ í™•ì¸
            if User.query.filter_by(username=new_username).first():
                flash('ì´ë¯¸ ì‚¬ìš© ì¤‘ì¸ ì•„ì´ë””ì…ë‹ˆë‹¤.', 'error')
                return redirect(url_for('settings_users'))
            
            new_user = User(
                username=new_username,
                name=new_name,
                role=new_role,
                password_hash=generate_password_hash(new_password)
            )
            db.session.add(new_user)
            db.session.commit()
            flash(f'{new_name} ì‚¬ìš©ìê°€ ì¶”ê°€ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
    
    return render_template('settings/users.html')

@app.route('/settings/points')
@login_required
def settings_points():
    """ìˆ˜ë™ í¬ì¸íŠ¸ ê´€ë¦¬ í˜ì´ì§€"""
    children = Child.query.filter_by(include_in_stats=True).order_by(Child.grade, Child.name).all()
    return render_template('settings/points.html', children=children)

@app.route('/api/children/by-grade')
@login_required
def get_children_by_grade():
    """í•™ë…„ë³„ ì•„ë™ ëª©ë¡ ì¡°íšŒ API"""
    try:
        grade = request.args.get('grade', type=int)
        
        # í†µê³„ì— í¬í•¨ëœ ì•„ë™ë“¤ë§Œ ì¡°íšŒ
        query = Child.query.filter_by(include_in_stats=True)
        
        # í•™ë…„ í•„í„° ì ìš© (ì„ íƒì‚¬í•­)
        if grade:
            query = query.filter_by(grade=grade)
        
        # í•™ë…„, ì´ë¦„ ìˆœìœ¼ë¡œ ì •ë ¬
        children = query.order_by(Child.grade, Child.name).all()
        
        # JSON í˜•íƒœë¡œ ë°˜í™˜
        children_data = []
        for child in children:
            children_data.append({
                'id': child.id,
                'name': child.name,
                'grade': child.grade,
                'display_name': f"{child.name} ({child.grade}í•™ë…„)"
            })
        
        return jsonify({
            'success': True,
            'children': children_data,
            'total': len(children_data)
        })
        
    except Exception as e:
        print(f"âŒ ì•„ë™ ëª©ë¡ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'children': []
        })

# ìˆ˜ë™ í¬ì¸íŠ¸ ê´€ë¦¬ API
@app.route('/api/manual-points', methods=['POST'])
@login_required
def add_manual_points():
    """ìˆ˜ë™ í¬ì¸íŠ¸ ì¶”ê°€ API"""
    try:
        data = request.get_json()
        child_id = data.get('child_id')
        subject = data.get('subject')
        points = data.get('points')
        reason = data.get('reason')
        
        # ì…ë ¥ ê²€ì¦
        if not all([child_id, subject, points is not None, reason]):
            return jsonify({'success': False, 'error': 'ëª¨ë“  í•„ë“œë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.'})
        
        # ì•„ë™ í™•ì¸
        child = Child.query.get(child_id)
        if not child:
            return jsonify({'success': False, 'error': 'ì•„ë™ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.'})
        
        # ì˜¤ëŠ˜ ë‚ ì§œì˜ ê¸°ë¡ ì°¾ê¸° ë˜ëŠ” ìƒì„±
        today = datetime.now().date()
        daily_record = DailyPoints.query.filter_by(child_id=child_id, date=today).first()
        
        if not daily_record:
            # ìƒˆ ê¸°ë¡ ìƒì„±
            daily_record = DailyPoints(
                child_id=child_id,
                date=today,
                korean_points=0,
                math_points=0,
                ssen_points=0,
                reading_points=0,
                piano_points=0,
                english_points=0,
                advanced_math_points=0,
                writing_points=0,
                manual_points=0,
                manual_history='[]',
                total_points=0,
                created_by=current_user.id
            )
            db.session.add(daily_record)
        
        # ìˆ˜ë™ íˆìŠ¤í† ë¦¬ ì—…ë°ì´íŠ¸
        import json
        try:
            history = json.loads(daily_record.manual_history) if daily_record.manual_history else []
        except:
            history = []
        
        # ìƒˆ íˆìŠ¤í† ë¦¬ í•­ëª© ì¶”ê°€
        new_history_item = {
            'id': len(history) + 1,
            'subject': subject,
            'points': points,
            'reason': reason,
            'created_by': current_user.name or current_user.username,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        history.append(new_history_item)
        
        # ìˆ˜ë™ í¬ì¸íŠ¸ ì´í•© ê³„ì‚°
        manual_total = sum(item['points'] for item in history)
        
        # ê¸°ë¡ ì—…ë°ì´íŠ¸
        daily_record.manual_history = json.dumps(history, ensure_ascii=False)
        daily_record.manual_points = manual_total
        
        # ì´ í¬ì¸íŠ¸ ì¬ê³„ì‚° (manual_historyì—ì„œ ì‹¤ì‹œê°„ ê³„ì‚°)
        manual_points_calculated = get_manual_points_from_history(daily_record)
        daily_record.total_points = (
            daily_record.korean_points + daily_record.math_points + 
            daily_record.ssen_points + daily_record.reading_points +
            daily_record.piano_points + daily_record.english_points +
            daily_record.advanced_math_points + daily_record.writing_points +
            manual_points_calculated
        )
        
        # í¬ì¸íŠ¸ íˆìŠ¤í† ë¦¬ì—ë„ ê¸°ë¡ (ë³€ê²½ ì´ë ¥ í˜ì´ì§€ìš©)
        change_type = 'ì¶”ê°€' if points > 0 else 'ì°¨ê°'
        points_history = PointsHistory(
            child_id=child_id,
            date=today,
            old_korean_points=0, old_math_points=0, old_ssen_points=0, old_reading_points=0, old_total_points=daily_record.total_points - points,
            new_korean_points=0, new_math_points=0, new_ssen_points=0, new_reading_points=0, new_total_points=daily_record.total_points,
            change_type=change_type,
            changed_by=current_user.id,
            change_reason=f'ìˆ˜ë™ {change_type}: {subject} ({reason})'
        )
        db.session.add(points_history)
        
        # ëˆ„ì  í¬ì¸íŠ¸ ìë™ ì—…ë°ì´íŠ¸
        update_cumulative_points(child_id, commit=False)
        
        db.session.commit()
        
        # ì‹¤ì‹œê°„ ë°±ì—… í˜¸ì¶œ
        try:
            realtime_backup(child_id, "manual_update")
        except Exception as backup_error:
            print(f"ë°±ì—… ì‹¤íŒ¨: {backup_error}")
        
        return jsonify({'success': True, 'message': f'ìˆ˜ë™ í¬ì¸íŠ¸ê°€ {change_type}ë˜ì—ˆìŠµë‹ˆë‹¤.'})
        
    except Exception as e:
        db.session.rollback()
        print(f"ìˆ˜ë™ í¬ì¸íŠ¸ ì¶”ê°€ ì˜¤ë¥˜: {str(e)}")
        return jsonify({'success': False, 'error': f'ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

@app.route('/api/manual-points/recent')
@login_required
def get_recent_manual_points():
    """ìµœê·¼ ìˆ˜ë™ í¬ì¸íŠ¸ ë‚´ì—­ ì¡°íšŒ API"""
    try:
        import json
        
        # ìµœê·¼ 20ê°œ ê¸°ë¡ ì¡°íšŒ
        recent_records = DailyPoints.query.filter(
            DailyPoints.manual_history != '[]',
            DailyPoints.manual_history.isnot(None)
        ).order_by(DailyPoints.date.desc()).limit(20).all()
        
        history_items = []
        
        for record in recent_records:
            try:
                history = json.loads(record.manual_history) if record.manual_history else []
                child = Child.query.get(record.child_id)
                
                for item in reversed(history):  # ìµœì‹ ìˆœìœ¼ë¡œ
                    history_items.append({
                        'id': f"{record.id}_{item['id']}",  # ê³ ìœ  ID
                        'child_name': child.name if child else 'ì•Œ ìˆ˜ ì—†ìŒ',
                        'subject': item['subject'],
                        'points': item['points'],
                        'reason': item['reason'],
                        'created_by': item['created_by'],
                        'created_at': item['created_at']
                    })
            except:
                continue
        
        # ë‚ ì§œìˆœìœ¼ë¡œ ì •ë ¬ í›„ 20ê°œë§Œ
        history_items.sort(key=lambda x: x['created_at'], reverse=True)
        history_items = history_items[:20]
        
        return jsonify({'success': True, 'history': history_items})
        
    except Exception as e:
        print(f"ìˆ˜ë™ í¬ì¸íŠ¸ ì¡°íšŒ ì˜¤ë¥˜: {str(e)}")
        return jsonify({'success': False, 'error': f'ì¡°íšŒ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

@app.route('/api/manual-points/<item_id>', methods=['DELETE'])
@login_required
def delete_manual_point(item_id):
    """ìˆ˜ë™ í¬ì¸íŠ¸ ì‚­ì œ API"""
    try:
        # item_idëŠ” "record_id_history_id" í˜•íƒœ
        record_id, history_id = item_id.split('_')
        record_id = int(record_id)
        history_id = int(history_id)
        
        daily_record = DailyPoints.query.get(record_id)
        if not daily_record:
            return jsonify({'success': False, 'error': 'ê¸°ë¡ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.'})
        
        import json
        try:
            history = json.loads(daily_record.manual_history) if daily_record.manual_history else []
        except:
            return jsonify({'success': False, 'error': 'íˆìŠ¤í† ë¦¬ ë°ì´í„° ì˜¤ë¥˜'})
        
        # í•´ë‹¹ í•­ëª© ì‚­ì œ
        history = [item for item in history if item['id'] != history_id]
        
        # ìˆ˜ë™ í¬ì¸íŠ¸ ì´í•© ì¬ê³„ì‚°
        manual_total = sum(item['points'] for item in history)
        
        # ê¸°ë¡ ì—…ë°ì´íŠ¸
        daily_record.manual_history = json.dumps(history, ensure_ascii=False)
        daily_record.manual_points = manual_total
        
        # ì´ í¬ì¸íŠ¸ ì¬ê³„ì‚° (manual_historyì—ì„œ ì‹¤ì‹œê°„ ê³„ì‚°)
        manual_points_calculated = get_manual_points_from_history(daily_record)
        daily_record.total_points = (
            daily_record.korean_points + daily_record.math_points + 
            daily_record.ssen_points + daily_record.reading_points +
            daily_record.piano_points + daily_record.english_points +
            daily_record.advanced_math_points + daily_record.writing_points +
            manual_points_calculated
        )
        
        # ëˆ„ì  í¬ì¸íŠ¸ ìë™ ì—…ë°ì´íŠ¸
        update_cumulative_points(daily_record.child_id, commit=False)
        
        db.session.commit()
        
        # ì‹¤ì‹œê°„ ë°±ì—… í˜¸ì¶œ
        try:
            realtime_backup(daily_record.child_id, "manual_delete")
        except Exception as backup_error:
            print(f"ë°±ì—… ì‹¤íŒ¨: {backup_error}")
        
        return jsonify({'success': True, 'message': 'ìˆ˜ë™ í¬ì¸íŠ¸ê°€ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤.'})
        
    except Exception as e:
        db.session.rollback()
        print(f"ìˆ˜ë™ í¬ì¸íŠ¸ ì‚­ì œ ì˜¤ë¥˜: {str(e)}")
        return jsonify({'success': False, 'error': f'ì‚­ì œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

@app.route('/settings/data')
@login_required
def settings_data():
    """ë°ì´í„° ê´€ë¦¬ í˜ì´ì§€"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'seed_data':
            # ì‹œë“œ ë°ì´í„° ì‹¤í–‰
            try:
                from scripts.seed_data import main as seed_main
                seed_main()
                flash('ê¸°ë³¸ ì‹œë“œ ë°ì´í„°ê°€ ì„±ê³µì ìœ¼ë¡œ ì‹¤í–‰ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
            except Exception as e:
                flash(f'ì‹œë“œ ë°ì´í„° ì‹¤í–‰ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {e}', 'error')
        
        elif action == 'reset_data':
            # ë°ì´í„° ì´ˆê¸°í™” (ê°œë°œìë§Œ)
            if current_user.role != 'ê°œë°œì':
                flash('ë°ì´í„° ì´ˆê¸°í™” ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
                return redirect(url_for('settings_data'))
            
            try:
                # ëª¨ë“  ë°ì´í„° ì‚­ì œ
                DailyPoints.query.delete()
                LearningRecord.query.delete()
                Child.query.delete()
                User.query.delete()
                db.session.commit()
                flash('ëª¨ë“  ë°ì´í„°ê°€ ì´ˆê¸°í™”ë˜ì—ˆìŠµë‹ˆë‹¤.', 'success')
            except Exception as e:
                flash(f'ë°ì´í„° ì´ˆê¸°í™” ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {e}', 'error')
        
        elif action == 'export_data':
            # ë°ì´í„° ë‚´ë³´ë‚´ê¸° (ê°œë°œìë§Œ)
            if current_user.role != 'ê°œë°œì':
                flash('ë°ì´í„° ë‚´ë³´ë‚´ê¸° ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
                return redirect(url_for('settings_data'))
            
            try:
                # ê°„ë‹¨í•œ ë°ì´í„° ìš”ì•½ ë‚´ë³´ë‚´ê¸°
                children_count = Child.query.count()
                users_count = User.query.count()
                records_count = LearningRecord.query.count()
                points_count = DailyPoints.query.count()
                
                export_data = {
                    'children_count': children_count,
                    'users_count': users_count,
                    'records_count': records_count,
                    'points_count': points_count,
                    'export_date': datetime.now().isoformat()
                }
                
                # JSON íŒŒì¼ë¡œ ë‹¤ìš´ë¡œë“œ
                response = jsonify(export_data)
                response.headers['Content-Disposition'] = 'attachment; filename=data_export.json'
                return response
                
            except Exception as e:
                flash(f'ë°ì´í„° ë‚´ë³´ë‚´ê¸° ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {e}', 'error')
    
    # í˜„ì¬ ë°ì´í„°ë² ì´ìŠ¤ í˜„í™©
    children_count = Child.query.count()
    users_count = User.query.count()
    records_count = LearningRecord.query.count()
    points_count = DailyPoints.query.count()
    
    return render_template('settings/data.html', 
                         children_count=children_count,
                         users_count=users_count,
                         records_count=records_count,
                         points_count=points_count)

@app.route('/settings/ui')
@login_required
def settings_ui():
    """UI/UX ì„¤ì • í˜ì´ì§€"""
    return render_template('settings/ui.html')

@app.route('/profile')
@login_required
def profile():
    """í”„ë¡œí•„ í˜ì´ì§€"""
    # í…ŒìŠ¤íŠ¸ì‚¬ìš©ìëŠ” ì ‘ê·¼ ë¶ˆê°€
    if current_user.role == 'í…ŒìŠ¤íŠ¸ì‚¬ìš©ì':
        flash('í”„ë¡œí•„ í˜ì´ì§€ì— ì ‘ê·¼í•  ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('profile.html')

@app.route('/privacy-policy')
def privacy_policy():
    """ê°œì¸ì •ë³´ë³´í˜¸ ë° ì‹œìŠ¤í…œ ë³´ì•ˆ ì •ì±… í˜ì´ì§€"""
    return render_template('privacy_policy.html')

@app.route('/settings/system')
@login_required
def settings_system():
    """ì‹œìŠ¤í…œ ì •ë³´ í˜ì´ì§€"""
    return render_template('settings/system.html')

@app.route('/settings/security')
@login_required
def settings_security():
    """ë³´ì•ˆ ì„¤ì • ì§„ë‹¨ í˜ì´ì§€ (ê°œë°œì ì „ìš©)"""
    if current_user.role != 'ê°œë°œì':
        flash('ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('settings'))
    
    # í˜„ì¬ ë³´ì•ˆ ì„¤ì • ìƒíƒœ ì²´í¬
    security_status = {
        'session_timeout': app.config.get('PERMANENT_SESSION_LIFETIME'),
        'secure_cookies': {
            'httponly': app.config.get('SESSION_COOKIE_HTTPONLY'),
            'samesite': app.config.get('SESSION_COOKIE_SAMESITE'),
            'secure': app.config.get('SESSION_COOKIE_SECURE')
        },
        'brute_force_protection': {
            'ip_tracking': len(failed_login_attempts),
            'blocked_ips': len(blocked_ips),
            'protection_enabled': True
        },
        'security_headers': {
            'csp_enabled': True,
            'hsts_enabled': bool(os.environ.get('DATABASE_URL')),
            'permissions_policy': True,
            'xss_protection': True
        },
        'environment': 'Production' if os.environ.get('DATABASE_URL') else 'Development'
    }
    
    return render_template('settings/security.html', security_status=security_status)

@app.route('/api/security/test-headers')
@login_required
def test_security_headers():
    """ë³´ì•ˆ í—¤ë” í…ŒìŠ¤íŠ¸ API (ê°œë°œì ì „ìš©)"""
    if current_user.role != 'ê°œë°œì':
        return jsonify({'error': 'ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤.'}), 403
    
    # í˜„ì¬ ì‘ë‹µì— ì ìš©ëœ í—¤ë”ë“¤ ë°˜í™˜
    test_response = {
        'timestamp': datetime.utcnow().isoformat(),
        'security_headers_test': 'OK',
        'csp_policy': 'Active',
        'hsts_status': 'Active' if os.environ.get('DATABASE_URL') else 'Development Mode',
        'brute_force_stats': {
            'failed_attempts_tracked': len(failed_login_attempts),
            'currently_blocked_ips': len(blocked_ips)
        }
    }
    
    return jsonify(test_response)

@app.route('/cumulative-points')
@login_required
def cumulative_points():
    """ëˆ„ì  í¬ì¸íŠ¸ ì…ë ¥ ë° ê´€ë¦¬ í˜ì´ì§€"""
    children = Child.query.order_by(Child.grade, Child.name).all()
    return render_template('cumulative_points.html', children=children)

@app.route('/cumulative-points/input', methods=['POST'])
@login_required
def input_cumulative_points():
    """ëˆ„ì  í¬ì¸íŠ¸ ì…ë ¥ ì²˜ë¦¬"""
    try:
        data = request.get_json()
        child_id = data.get('child_id')
        cumulative_points = data.get('cumulative_points')
        
        if not child_id or cumulative_points is None:
            return jsonify({'success': False, 'message': 'í•„ìˆ˜ ì •ë³´ê°€ ëˆ„ë½ë˜ì—ˆìŠµë‹ˆë‹¤.'}), 400
        
        # í¬ì¸íŠ¸ ê°’ ê²€ì¦
        try:
            cumulative_points = int(cumulative_points)
            if cumulative_points < 0:
                return jsonify({'success': False, 'message': 'í¬ì¸íŠ¸ëŠ” 0 ì´ìƒì´ì–´ì•¼ í•©ë‹ˆë‹¤.'}), 400
        except ValueError:
            return jsonify({'success': False, 'message': 'ì˜¬ë°”ë¥¸ ìˆ«ìë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.'}), 400
        
        # ì•„ë™ ì •ë³´ ì—…ë°ì´íŠ¸
        child = Child.query.get(child_id)
        if not child:
            return jsonify({'success': False, 'message': 'ì•„ë™ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.'}), 404
        
        child.cumulative_points = cumulative_points
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'{child.name}ì˜ ëˆ„ì  í¬ì¸íŠ¸ê°€ {cumulative_points}ì ìœ¼ë¡œ ì„¤ì •ë˜ì—ˆìŠµë‹ˆë‹¤.',
            'child_name': child.name,
            'cumulative_points': cumulative_points
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'}), 500

@app.route('/cumulative-points/bulk-input', methods=['POST'])
@login_required
def bulk_input_cumulative_points():
    """ì¼ê´„ ëˆ„ì  í¬ì¸íŠ¸ ì…ë ¥ ì²˜ë¦¬"""
    try:
        data = request.get_json()
        points_data = data.get('points_data', [])
        
        if not points_data:
            return jsonify({'success': False, 'message': 'ì…ë ¥í•  ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.'}), 400
        
        updated_count = 0
        errors = []
        
        for item in points_data:
            child_id = item.get('child_id')
            cumulative_points = item.get('cumulative_points')
            
            if not child_id or cumulative_points is None:
                errors.append(f'ì•„ë™ ID {child_id}: í¬ì¸íŠ¸ ì •ë³´ ëˆ„ë½')
                continue
            
            try:
                cumulative_points = int(cumulative_points)
                if cumulative_points < 0:
                    errors.append(f'ì•„ë™ ID {child_id}: í¬ì¸íŠ¸ëŠ” 0 ì´ìƒì´ì–´ì•¼ í•©ë‹ˆë‹¤.')
                    continue
            except ValueError:
                errors.append(f'ì•„ë™ ID {child_id}: ì˜¬ë°”ë¥¸ ìˆ«ìê°€ ì•„ë‹™ë‹ˆë‹¤.')
                continue
            
            child = Child.query.get(child_id)
            if not child:
                errors.append(f'ì•„ë™ ID {child_id}: ì•„ë™ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.')
                continue
            
            child.cumulative_points = cumulative_points
            updated_count += 1
        
        if errors:
            db.session.rollback()
            return jsonify({
                'success': False, 
                'message': f'{len(errors)}ê±´ì˜ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤.',
                'errors': errors
            }), 400
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{updated_count}ëª…ì˜ ì•„ë™ ëˆ„ì  í¬ì¸íŠ¸ê°€ ì„±ê³µì ìœ¼ë¡œ ì—…ë°ì´íŠ¸ë˜ì—ˆìŠµë‹ˆë‹¤.',
            'updated_count': updated_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'}), 500

# í¬ì¸íŠ¸ ë³€ê²½ ì´ë ¥ í…Œì´ë¸”
class PointsHistory(db.Model):
    """í¬ì¸íŠ¸ ë³€ê²½ ì´ë ¥ ê¸°ë¡"""
    id = db.Column(db.Integer, primary_key=True)
    child_id = db.Column(db.Integer, db.ForeignKey('child.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    
    # ë³€ê²½ ì „ í¬ì¸íŠ¸
    old_korean_points = db.Column(db.Integer, default=0)
    old_math_points = db.Column(db.Integer, default=0)
    old_ssen_points = db.Column(db.Integer, default=0)
    old_reading_points = db.Column(db.Integer, default=0)
    old_total_points = db.Column(db.Integer, default=0)
    
    old_piano_points = db.Column(db.Integer, default=0)
    old_english_points = db.Column(db.Integer, default=0) 
    old_advanced_math_points = db.Column(db.Integer, default=0)
    old_writing_points = db.Column(db.Integer, default=0)
    old_total_points = db.Column(db.Integer, default=0)
    
    # ë³€ê²½ í›„ í¬ì¸íŠ¸
    new_korean_points = db.Column(db.Integer, default=0)
    new_math_points = db.Column(db.Integer, default=0)
    new_ssen_points = db.Column(db.Integer, default=0)
    new_reading_points = db.Column(db.Integer, default=0)
    new_total_points = db.Column(db.Integer, default=0)

    new_piano_points = db.Column(db.Integer, default=0)
    new_english_points = db.Column(db.Integer, default=0)
    new_advanced_math_points = db.Column(db.Integer, default=0)
    new_writing_points = db.Column(db.Integer, default=0)
    new_total_points = db.Column(db.Integer, default=0)
    
    # ë³€ê²½ ì •ë³´
    change_type = db.Column(db.String(20), default='update')  # 'create', 'update', 'delete'
    changed_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    change_reason = db.Column(db.String(200))  # ë³€ê²½ ì‚¬ìœ  (ì„ íƒì‚¬í•­)
    
    # ê´€ê³„ ì„¤ì •
    child = db.relationship('Child', backref='points_history', lazy=True)
    user = db.relationship('User', backref='points_changes', lazy=True)
    
    def __repr__(self):
        return f'<PointsHistory {self.child.name} {self.date} {self.change_type}>'

class Notification(db.Model):
    """ì•Œë¦¼ ì‹œìŠ¤í…œ"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    message = db.Column(db.Text, nullable=False)
    
    # ì•Œë¦¼ íƒ€ì… ë° ìš°ì„ ìˆœìœ„
    type = db.Column(db.String(30), default='info')  # 'info', 'success', 'warning', 'danger'
    priority = db.Column(db.Integer, default=1)  # 1=ë‚®ìŒ, 2=ë³´í†µ, 3=ë†’ìŒ, 4=ê¸´ê¸‰
    
    # ëŒ€ìƒ ë° ì¡°ê±´
    target_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # nullì´ë©´ ì „ì²´ ê³µì§€
    target_role = db.Column(db.String(30), nullable=True)  # íŠ¹ì • ì—­í• ì—ë§Œ í‘œì‹œ
    child_id = db.Column(db.Integer, db.ForeignKey('child.id'), nullable=True)  # íŠ¹ì • ì•„ë™ ê´€ë ¨ ì•Œë¦¼
    
    # ìƒíƒœ ê´€ë¦¬
    is_read = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    # is_deleted = db.Column(db.Boolean, default=False)  # ì†Œí”„íŠ¸ ì‚­ì œ í”Œë˜ê·¸
    auto_expire = db.Column(db.Boolean, default=False)  # ìë™ ë§Œë£Œ ì—¬ë¶€
    expire_date = db.Column(db.DateTime, nullable=True)  # ë§Œë£Œ ì¼ì‹œ
    
    # ë©”íƒ€ë°ì´í„°
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    read_at = db.Column(db.DateTime, nullable=True)
    
    # ê´€ê³„ ì„¤ì •
    target_user = db.relationship('User', foreign_keys=[target_user_id], backref='received_notifications', lazy=True)
    creator = db.relationship('User', foreign_keys=[created_by], backref='created_notifications', lazy=True)
    child = db.relationship('Child', backref='notifications', lazy=True)
    
    def __repr__(self):
        return f'<Notification {self.title} ({self.type})>'
    
    @property
    def icon(self):
        """ì•Œë¦¼ íƒ€ì…ì— ë”°ë¥¸ ì•„ì´ì½˜ ë°˜í™˜"""
        icons = {
            'info': 'info-circle',
            'success': 'check-circle',
            'warning': 'exclamation-triangle',
            'danger': 'x-circle',
            'reminder': 'clock',
            'system': 'gear',
            'backup_success': 'cloud-check',
            'backup_failed': 'cloud-x',
            'restore_success': 'arrow-clockwise',
            'restore_failed': 'exclamation-triangle'
        }
        return icons.get(self.type, 'bell')
    
    @property
    def color(self):
        """ì•Œë¦¼ íƒ€ì…ì— ë”°ë¥¸ ìƒ‰ìƒ ë°˜í™˜"""
        colors = {
            'info': 'primary',
            'success': 'success',
            'warning': 'warning',
            'danger': 'danger',
            'reminder': 'info',
            'system': 'secondary',
            'backup_success': 'success',
            'backup_failed': 'danger',
            'restore_success': 'success',
            'restore_failed': 'danger'
        }
        return colors.get(self.type, 'primary')

def create_backup_notification(backup_type, status, message, target_role='ê°œë°œì'):
    """ë°±ì—… ê´€ë ¨ ì•Œë¦¼ ìƒì„±"""
    try:
        if status == 'success':
            notification_type = 'backup_success'
            title = f"{backup_type} ë°±ì—… ì™„ë£Œ"
            priority = 2  # ë³´í†µ ìš°ì„ ìˆœìœ„
        else:
            notification_type = 'backup_failed'
            title = f"{backup_type} ë°±ì—… ì‹¤íŒ¨"
            priority = 4  # ê¸´ê¸‰ ìš°ì„ ìˆœìœ„
        
        notification = Notification(
            title=title,
            message=message,
            type=notification_type,
            target_role=target_role,
            priority=priority,
            auto_expire=True,
            expire_date=datetime.utcnow() + timedelta(days=7),  # 7ì¼ í›„ ìë™ ë§Œë£Œ
            created_by=1  # ì‹œìŠ¤í…œ ìƒì„±
        )
        
        db.session.add(notification)
        db.session.commit()
        print(f"âœ… ë°±ì—… ì•Œë¦¼ ìƒì„±: {title}")
        return notification
        
    except Exception as e:
        db.session.rollback()
        print(f"âŒ ë°±ì—… ì•Œë¦¼ ìƒì„± ì‹¤íŒ¨: {e}")
        return None

def create_restore_notification(status, message, target_role='ê°œë°œì'):
    """ë³µì› ê´€ë ¨ ì•Œë¦¼ ìƒì„±"""
    try:
        if status == 'success':
            notification_type = 'restore_success'
            title = "ë°ì´í„°ë² ì´ìŠ¤ ë³µì› ì™„ë£Œ"
            priority = 3  # ë†’ì€ ìš°ì„ ìˆœìœ„
        else:
            notification_type = 'restore_failed'
            title = "ë°ì´í„°ë² ì´ìŠ¤ ë³µì› ì‹¤íŒ¨"
            priority = 4  # ê¸´ê¸‰ ìš°ì„ ìˆœìœ„
        
        notification = Notification(
            title=title,
            message=message,
            type=notification_type,
            target_role=target_role,
            priority=priority,
            auto_expire=True,
            expire_date=datetime.utcnow() + timedelta(days=7),  # 7ì¼ í›„ ìë™ ë§Œë£Œ
            created_by=1  # ì‹œìŠ¤í…œ ìƒì„±
        )
        
        db.session.add(notification)
        db.session.commit()
        print(f"âœ… ë³µì› ì•Œë¦¼ ìƒì„±: {title}")
        return notification
        
    except Exception as e:
        db.session.rollback()
        print(f"âŒ ë³µì› ì•Œë¦¼ ìƒì„± ì‹¤íŒ¨: {e}")
        return None

# ===== ì•Œë¦¼ ì‹œìŠ¤í…œ í—¬í¼ í•¨ìˆ˜ë“¤ =====

def create_notification(title, message, notification_type='info', target_user_id=None, target_role=None, 
                       child_id=None, priority=1, auto_expire=False, expire_days=None):
    """ìƒˆ ì•Œë¦¼ ìƒì„±"""
    try:
        print(f"DEBUG: create_notification í˜¸ì¶œë¨ - {title}")
        print(f"DEBUG: current_user.is_authenticated = {current_user.is_authenticated}")
        print(f"DEBUG: current_user.id = {current_user.id if current_user.is_authenticated else 'None'}")
        
        expire_date = None
        if auto_expire and expire_days:
            expire_date = datetime.utcnow() + timedelta(days=expire_days)
        
        notification = Notification(
            title=title,
            message=message,
            type=notification_type,
            target_user_id=target_user_id,
            target_role=target_role,
            child_id=child_id,
            priority=priority,
            auto_expire=auto_expire,
            expire_date=expire_date,
            created_by=current_user.id if current_user.is_authenticated else 1
        )
        
        print(f"DEBUG: Notification ê°ì²´ ìƒì„± ì™„ë£Œ")
        db.session.add(notification)
        print(f"DEBUG: DBì— ì¶”ê°€ ì™„ë£Œ")
        db.session.commit()
        print(f"DEBUG: DB ì»¤ë°‹ ì™„ë£Œ - ì•Œë¦¼ ID: {notification.id}")
        return notification
    except Exception as e:
        db.session.rollback()
        print(f"ì•Œë¦¼ ìƒì„± ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_user_notifications(user_id, limit=10, unread_only=False):
    """ì‚¬ìš©ìë³„ ì•Œë¦¼ ì¡°íšŒ"""
    user = User.query.get(user_id)
    if not user:
        return []
    
    # ì¡°ê±´ë“¤ì„ ë¦¬ìŠ¤íŠ¸ë¡œ êµ¬ì„±
    conditions = []
    
    # 1. ê°œì¸ ì•Œë¦¼ (target_user_idê°€ í˜„ì¬ ì‚¬ìš©ì)
    conditions.append(Notification.target_user_id == user_id)
    
    # 2. ì „ì²´ ê³µì§€ (target_user_id=None, target_role=None)
    conditions.append(
        db.and_(
            Notification.target_user_id.is_(None),
            Notification.target_role.is_(None)
        )
    )
    
    # 3. ì—­í• ë³„ ê³µì§€ (target_user_id=None, target_role=ì‚¬ìš©ì ì—­í• )
    conditions.append(
        db.and_(
            Notification.target_user_id.is_(None),
            Notification.target_role == user.role
        )
    )
    
    # 4. ì•„ë™ ê´€ë ¨ ì•Œë¦¼ (child_idê°€ ìˆëŠ” ì•Œë¦¼ - ëª¨ë“  ì‚¬ìš©ìì—ê²Œ í‘œì‹œ)
    conditions.append(Notification.child_id.isnot(None))
    
    # ê¸°ë³¸ ì¿¼ë¦¬: ìœ„ì˜ ì¡°ê±´ë“¤ ì¤‘ í•˜ë‚˜ë¼ë„ ë§Œì¡±í•˜ëŠ” ì•Œë¦¼
    query = Notification.query.filter(db.or_(*conditions))
    
    # ë§Œë£Œ ì¡°ê±´ ì ìš©
    query = query.filter(
        db.or_(
            Notification.expire_date.is_(None),
            Notification.expire_date > datetime.utcnow()
        )
    )
    
    # ì½ì§€ ì•Šì€ ì•Œë¦¼ë§Œ í•„í„°ë§ (í•„ìš”ì‹œ)
    if unread_only:
        query = query.filter(Notification.is_read == False)
    
    # ì •ë ¬
    query = query.order_by(Notification.priority.desc(), Notification.created_at.desc())
    
    # limit ì ìš©
    if limit is not None:
        query = query.limit(limit)
    
    return query.all()

def mark_notification_read(notification_id, user_id):
    """ì•Œë¦¼ì„ ì½ìŒìœ¼ë¡œ í‘œì‹œ"""
    notification = Notification.query.filter_by(id=notification_id).first()
    if notification and (notification.target_user_id == user_id or notification.target_user_id is None):
        notification.is_read = True
        notification.read_at = datetime.utcnow()
        db.session.commit()
        return True
    return False

def delete_notification(notification_id, user_id):
    """ì•Œë¦¼ ì†Œí”„íŠ¸ ì‚­ì œ (ê°œë°œìë§Œ ê°€ëŠ¥)"""
    try:
        # ì‚¬ìš©ì ê¶Œí•œ í™•ì¸
        user = User.query.get(user_id)
        if not user or user.role != 'ê°œë°œì':
            return False, "ê°œë°œì ê¶Œí•œì´ í•„ìš”í•©ë‹ˆë‹¤."
        
        # ì•Œë¦¼ ì¡°íšŒ
        notification = Notification.query.filter_by(id=notification_id).first()
        if not notification:
            return False, "ì•Œë¦¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤."
        
        # ì†Œí”„íŠ¸ ì‚­ì œ ì²˜ë¦¬
        db.session.delete(notification)
        db.session.commit()
        
        return True, "ì•Œë¦¼ì´ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤."
        
    except Exception as e:
        db.session.rollback()
        return False, f"ì‚­ì œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}"

def delete_multiple_notifications(notification_ids, user_id):
    """ì—¬ëŸ¬ ì•Œë¦¼ ì¼ê´„ ì‚­ì œ (ê°œë°œìë§Œ ê°€ëŠ¥)"""
    try:
        # ì‚¬ìš©ì ê¶Œí•œ í™•ì¸
        user = User.query.get(user_id)
        if not user or user.role != 'ê°œë°œì':
            return False, "ê°œë°œì ê¶Œí•œì´ í•„ìš”í•©ë‹ˆë‹¤."
        
        # ì•Œë¦¼ë“¤ ì¡°íšŒ ë° ì‚­ì œ
        notifications = Notification.query.filter(
            Notification.id.in_(notification_ids)
        ).all()
        
        if not notifications:
            return False, "ì‚­ì œí•  ì•Œë¦¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤."
        
        # ì†Œí”„íŠ¸ ì‚­ì œ ì²˜ë¦¬
        for notification in notifications:
            db.session.delete(notification)
        
        db.session.commit()
        
        return True, f"{len(notifications)}ê°œì˜ ì•Œë¦¼ì´ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤."
        
    except Exception as e:
        db.session.rollback()
        return False, f"ì‚­ì œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}"

def create_system_notification(title, message, target_role=None, priority=1):
    """ì‹œìŠ¤í…œ ì•Œë¦¼ ìƒì„±"""
    return create_notification(
        title=title,
        message=message,
        notification_type='system',
        target_role=target_role,
        priority=priority
    )

@app.route('/notifications')
@login_required
def notifications():
    """ì•Œë¦¼ ëª©ë¡ í˜ì´ì§€"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # ëª¨ë“  ì•Œë¦¼ ì¡°íšŒ
    all_notifications = get_user_notifications(current_user.id, limit=None)
    
    # í˜ì´ì§€ë„¤ì´ì…˜
    total = len(all_notifications)
    start = (page - 1) * per_page
    end = start + per_page
    notifications_page = all_notifications[start:end]
    
    # ì½ì§€ ì•Šì€ ì•Œë¦¼ ìˆ˜
    unread_count = len([n for n in all_notifications if not n.is_read])
    
    return render_template('notifications/list.html',
                         notifications=notifications_page,
                         page=page,
                         per_page=per_page,
                         total=total,
                         unread_count=unread_count)

@app.route('/notifications/<int:notification_id>/read', methods=['POST'])
@login_required
def mark_notification_as_read(notification_id):
    """ì•Œë¦¼ ì½ìŒ ì²˜ë¦¬"""
    success = mark_notification_read(notification_id, current_user.id)
    if success:
        return jsonify({'success': True})
    return jsonify({'success': False}), 400

@app.route('/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    """ëª¨ë“  ì•Œë¦¼ ì½ìŒ ì²˜ë¦¬"""
    notifications = get_user_notifications(current_user.id, limit=None, unread_only=True)
    
    for notification in notifications:
        mark_notification_read(notification.id, current_user.id)
    
    return jsonify({'success': True, 'count': len(notifications)})

@app.route('/notifications/<int:notification_id>/delete', methods=['POST'])
@login_required
def delete_single_notification(notification_id):
    """ê°œë³„ ì•Œë¦¼ ì‚­ì œ"""
    success, message = delete_notification(notification_id, current_user.id)
    if success:
        return jsonify({'success': True, 'message': message})
    return jsonify({'success': False, 'message': message}), 400

@app.route('/notifications/delete-multiple', methods=['POST'])
@login_required
def delete_multiple_notifications_route():
    """ì—¬ëŸ¬ ì•Œë¦¼ ì¼ê´„ ì‚­ì œ"""
    try:
        data = request.get_json()
        notification_ids = data.get('notification_ids', [])
        
        if not notification_ids:
            return jsonify({'success': False, 'message': 'ì‚­ì œí•  ì•Œë¦¼ì„ ì„ íƒí•´ì£¼ì„¸ìš”.'}), 400
        
        success, message = delete_multiple_notifications(notification_ids, current_user.id)
        if success:
            return jsonify({'success': True, 'message': message})
        return jsonify({'success': False, 'message': message}), 400
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'}), 500

@app.route('/notifications/test')
@login_required
def test_notifications():
    """í…ŒìŠ¤íŠ¸ ì•Œë¦¼ ìƒì„± (ê°œë°œìš©)"""
    if not current_user.role == 'ê°œë°œì':
        return redirect(url_for('dashboard'))
    
    # í…ŒìŠ¤íŠ¸ ì•Œë¦¼ë“¤ ìƒì„±
    create_notification(
        title="ì‹œìŠ¤í…œ ì—…ë°ì´íŠ¸ ì™„ë£Œ",
        message="í¬ì¸íŠ¸ ì‹œìŠ¤í…œì´ ì„±ê³µì ìœ¼ë¡œ ì—…ë°ì´íŠ¸ë˜ì—ˆìŠµë‹ˆë‹¤.",
        notification_type='success',
        priority=2
    )
    
    create_notification(
        title="ì£¼ê°„ ë³´ê³ ì„œ ì¤€ë¹„",
        message="ì´ë²ˆ ì£¼ ì•„ë™ë“¤ì˜ í•™ìŠµ ì„±ê³¼ ë³´ê³ ì„œë¥¼ í™•ì¸í•´ì£¼ì„¸ìš”.",
        notification_type='info',
        target_role='ì„¼í„°ì¥',
        priority=1
    )
    
    create_notification(
        title="ë°ì´í„° ë°±ì—… í•„ìš”",
        message="ì •ê¸° ë°ì´í„° ë°±ì—…ì„ ì§„í–‰í•´ì£¼ì„¸ìš”.",
        notification_type='warning',
        priority=3,
        auto_expire=True,
        expire_days=3
    )
    
    return redirect(url_for('notifications'))

@app.route('/points/history/<int:child_id>')
@login_required
def points_history(child_id):
    """ì•„ë™ë³„ í¬ì¸íŠ¸ ë³€ê²½ ì´ë ¥ ì¡°íšŒ"""
    child = Child.query.get_or_404(child_id)
    
    # ìµœê·¼ 30ì¼ê°„ì˜ ë³€ê²½ ì´ë ¥ ì¡°íšŒ
    from datetime import datetime, timedelta
    thirty_days_ago = datetime.utcnow().date() - timedelta(days=30)
    
    history_records = PointsHistory.query.filter(
        PointsHistory.child_id == child_id,
        PointsHistory.changed_at >= thirty_days_ago
    ).order_by(PointsHistory.changed_at.desc()).all()
    
    return render_template('points/history.html', 
                         child=child, 
                         history_records=history_records)

@app.route('/points/history')
@login_required
def all_points_history():
    """ì „ì²´ í¬ì¸íŠ¸ ë³€ê²½ ì´ë ¥ ì¡°íšŒ (ê´€ë¦¬ììš©)"""
    # ìµœê·¼ 100ê±´ì˜ ë³€ê²½ ì´ë ¥ ì¡°íšŒ
    history_records = PointsHistory.query.order_by(PointsHistory.changed_at.desc()).limit(100).all()
    
    return render_template('points/all_history.html', history_records=history_records)

def check_duplicate_daily_points():
    """ì¤‘ë³µ ì¼ì¼ í¬ì¸íŠ¸ ê¸°ë¡ ê²€ì‚¬ ë° ì •ë¦¬"""
    try:
        print("ğŸ” ì¤‘ë³µ ì¼ì¼ í¬ì¸íŠ¸ ê¸°ë¡ ê²€ì‚¬ ì‹œì‘...")
        from sqlalchemy import text
        
        result = db.session.execute(text("""
            SELECT child_id, date, COUNT(*) as count
            FROM daily_points 
            GROUP BY child_id, date 
            HAVING COUNT(*) > 1
        """))
        duplicates = result.fetchall()
        
        if not duplicates:
            print("âœ… ì¤‘ë³µëœ ì¼ì¼ í¬ì¸íŠ¸ ê¸°ë¡ì´ ì—†ìŠµë‹ˆë‹¤.")
            return
        
        print(f"âš ï¸ {len(duplicates)}ê°œì˜ ì¤‘ë³µ ê¸°ë¡ ë°œê²¬")
        
        for duplicate in duplicates:
            child_id = duplicate[0]
            date = duplicate[1]
            child = Child.query.get(child_id)
            print(f"  {child.name} - {date}: {duplicate[2]}ê°œ ê¸°ë¡")
            
            # í•´ë‹¹ ë‚ ì§œì˜ ëª¨ë“  ê¸°ë¡ì„ ID ìˆœìœ¼ë¡œ ì •ë ¬í•˜ì—¬ ì²« ë²ˆì§¸ë§Œ ë‚¨ê¸°ê³  ë‚˜ë¨¸ì§€ ì‚­ì œ
            records = DailyPoints.query.filter_by(
                child_id=child_id, 
                date=date
            ).order_by(DailyPoints.id.asc()).all()
            
            for record in records[1:]:  # ì²« ë²ˆì§¸ ì œì™¸í•˜ê³  ëª¨ë‘ ì‚­ì œ
                print(f"    ì‚­ì œ: ID {record.id} (ì´ì : {record.total_points})")
                db.session.delete(record)
            
            # ëˆ„ì  í¬ì¸íŠ¸ ì¬ê³„ì‚°
            update_cumulative_points(child_id)
        
        db.session.commit()
        print("âœ… ì¤‘ë³µ ê¸°ë¡ ì •ë¦¬ ì™„ë£Œ")
        
    except Exception as e:
        print(f"âŒ ì¤‘ë³µ ê¸°ë¡ ê²€ì‚¬ ì˜¤ë¥˜: {e}")
        db.session.rollback()

def validate_points_integrity():
    """í¬ì¸íŠ¸ ë°ì´í„° ë¬´ê²°ì„± ê²€ì¦ ë° ìë™ ìˆ˜ì •"""
    try:
        print("ğŸ” í¬ì¸íŠ¸ ë°ì´í„° ë¬´ê²°ì„± ê²€ì¦ ì‹œì‘...")
        children = Child.query.all()
        fixed_count = 0
        
        for child in children:
            # í•´ë‹¹ ì•„ë™ì˜ ëª¨ë“  ì¼ì¼ í¬ì¸íŠ¸ í•©ê³„ ê³„ì‚°
            calculated_total = db.session.query(
                db.func.sum(DailyPoints.total_points)
            ).filter_by(child_id=child.id).scalar() or 0
            
            if child.cumulative_points != calculated_total:
                print(f"âš ï¸ {child.name}ì˜ ëˆ„ì  í¬ì¸íŠ¸ ë¶ˆì¼ì¹˜ ë°œê²¬")
                print(f"  DB: {child.cumulative_points}, ê³„ì‚°: {calculated_total}")
                child.cumulative_points = calculated_total
                fixed_count += 1
        
        if fixed_count > 0:
            db.session.commit()
            print(f"ğŸ”§ ì´ {fixed_count}ëª…ì˜ ëˆ„ì  í¬ì¸íŠ¸ê°€ ìë™ìœ¼ë¡œ ìˆ˜ì •ë˜ì—ˆìŠµë‹ˆë‹¤.")
        else:
            print("âœ… ëª¨ë“  í¬ì¸íŠ¸ ë°ì´í„°ê°€ ì •ìƒì…ë‹ˆë‹¤.")
            
    except Exception as e:
        print(f"âŒ í¬ì¸íŠ¸ ë¬´ê²°ì„± ê²€ì¦ ì˜¤ë¥˜: {e}")
        db.session.rollback()

# ==================== ë°±ì—… ì‹œìŠ¤í…œ í•¨ìˆ˜ë“¤ ====================

def create_backup_directory():
    """ë°±ì—… ë””ë ‰í† ë¦¬ ìƒì„±"""
    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    # í•˜ìœ„ ë””ë ‰í† ë¦¬ë“¤ ìƒì„±
    subdirs = ['daily', 'monthly', 'realtime', 'database']
    for subdir in subdirs:
        subdir_path = os.path.join(backup_dir, subdir)
        if not os.path.exists(subdir_path):
            os.makedirs(subdir_path)
    
    return backup_dir

def get_backup_data():
    """ë°±ì—…í•  ë°ì´í„° ìˆ˜ì§‘"""
    try:
        # ì•„ë™ ì •ë³´
        children = Child.query.all()
        children_data = []
        for child in children:
            child_dict = {
                'id': child.id,
                'name': child.name,
                'grade': child.grade,
                'cumulative_points': child.cumulative_points,
                'created_at': child.created_at.isoformat() if child.created_at else None
            }
            children_data.append(child_dict)
        
        # ì¼ì¼ í¬ì¸íŠ¸ ê¸°ë¡
        daily_points = DailyPoints.query.all()
        daily_points_data = []
        for point in daily_points:
            point_dict = {
                'id': point.id,
                'child_id': point.child_id,
                'date': point.date.isoformat() if point.date else None,
                'korean_points': point.korean_points,
                'math_points': point.math_points,
                'ssen_points': point.ssen_points,
                'reading_points': point.reading_points,
                'piano_points': point.piano_points,
                'english_points': point.english_points,
                'advanced_math_points': point.advanced_math_points,
                'writing_points': point.writing_points,
                'manual_points': point.manual_points,
                'manual_history': point.manual_history,
                'total_points': point.total_points,
                'created_by': point.created_by,
                'created_at': point.created_at.isoformat() if point.created_at else None,
                'updated_at': point.updated_at.isoformat() if point.updated_at else None
            }
            daily_points_data.append(point_dict)
        
        # í¬ì¸íŠ¸ íˆìŠ¤í† ë¦¬
        points_history = PointsHistory.query.all()
        history_data = []
        for history in points_history:
            history_dict = {
                'id': history.id,
                'child_id': history.child_id,
                'date': history.date.isoformat() if history.date else None,
                'old_korean_points': history.old_korean_points,
                'old_math_points': history.old_math_points,
                'old_ssen_points': history.old_ssen_points,
                'old_reading_points': history.old_reading_points,
                'old_total_points': history.old_total_points,
                'new_korean_points': history.new_korean_points,
                'new_math_points': history.new_math_points,
                'new_ssen_points': history.new_ssen_points,
                'new_reading_points': history.new_reading_points,
                'new_total_points': history.new_total_points,
                'change_type': history.change_type,
                'changed_by': history.changed_by,
                'changed_at': history.changed_at.isoformat() if history.changed_at else None,
                'change_reason': history.change_reason
            }
            history_data.append(history_dict)
        
        # ì‚¬ìš©ì ì •ë³´
        users = User.query.all()
        users_data = []
        for user in users:
            user_dict = {
                'id': user.id,
                'username': user.username,
                'name': user.name,
                'role': user.role,
                'created_at': user.created_at.isoformat() if user.created_at else None
            }
            users_data.append(user_dict)
        
        backup_data = {
            'backup_metadata': {
                'backup_id': datetime.now().strftime('%Y-%m-%d_%H-%M-%S'),
                'backup_type': 'manual',
                'timestamp': datetime.now().isoformat(),
                'data_version': '1.0.0',
                'records_count': {
                    'children': len(children_data),
                    'daily_points': len(daily_points_data),
                    'points_history': len(history_data),
                    'users': len(users_data)
                }
            },
            'children': children_data,
            'daily_points': daily_points_data,
            'points_history': history_data,
            'users': users_data
        }
        
        return backup_data, None
        
    except Exception as e:
        return None, str(e)

def create_json_backup(backup_data, backup_dir, backup_type='manual'):
    """JSON í˜•íƒœë¡œ ë°±ì—… ìƒì„±"""
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        if backup_type == 'daily':
            filename = f"{datetime.now().strftime('%Y-%m-%d')}_{timestamp.split('_')[1]}.json"
            filepath = os.path.join(backup_dir, 'daily', filename)
        elif backup_type == 'monthly':
            filename = f"{datetime.now().strftime('%Y-%m')}_archive.json"
            filepath = os.path.join(backup_dir, 'monthly', filename)
        else:
            filename = f"{timestamp}.json"
            filepath = os.path.join(backup_dir, 'realtime', filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        return filepath, None
        
    except Exception as e:
        return None, str(e)

def create_excel_backup(backup_data, backup_dir, backup_type='manual'):
    """Excel í˜•íƒœë¡œ ë°±ì—… ìƒì„±"""
    if not BACKUP_EXCEL_AVAILABLE:
        print("âŒ Excel ë°±ì—…ì„ ìœ„í•œ íŒ¨í‚¤ì§€ê°€ ì„¤ì¹˜ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.")
        return None, "pandas ë˜ëŠ” openpyxl íŒ¨í‚¤ì§€ê°€ ì„¤ì¹˜ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤."
    
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        if backup_type == 'daily':
            filename = f"{datetime.now().strftime('%Y-%m-%d')}_{timestamp.split('_')[1]}.xlsx"
            filepath = os.path.join(backup_dir, 'daily', filename)
        elif backup_type == 'monthly':
            filename = f"{datetime.now().strftime('%Y-%m')}_archive.xlsx"
            filepath = os.path.join(backup_dir, 'monthly', filename)
        else:
            filename = f"{timestamp}.xlsx"
            filepath = os.path.join(backup_dir, 'realtime', filename)
        
        # Excel ì›Œí¬ë¶ ìƒì„±
        wb = Workbook()
        
        # ì•„ë™ ì •ë³´ ì‹œíŠ¸
        ws_children = wb.active
        ws_children.title = "ì•„ë™ì •ë³´"
        ws_children.append(['ID', 'ì´ë¦„', 'í•™ë…„', 'ëˆ„ì í¬ì¸íŠ¸', 'ìƒì„±ì¼'])
        
        for child in backup_data['children']:
            ws_children.append([
                child['id'],
                child['name'],
                child['grade'],
                child['cumulative_points'],
                child['created_at']
            ])
        
        # í¬ì¸íŠ¸ ê¸°ë¡ ì‹œíŠ¸
        ws_points = wb.create_sheet("í¬ì¸íŠ¸ê¸°ë¡")
        ws_points.append(['ID', 'ì•„ë™ID', 'ë‚ ì§œ', 'êµ­ì–´', 'ìˆ˜í•™', 'ìˆìˆ˜í•™', 'ë…ì„œ', 'í”¼ì•„ë…¸', 'ì˜ì–´', 'ê³ í•™ë…„ìˆ˜í•™', 'ì“°ê¸°', 'ìˆ˜ë™í¬ì¸íŠ¸', 'ìˆ˜ë™íˆìŠ¤í† ë¦¬', 'ì´ì ', 'ì…ë ¥ì', 'ìƒì„±ì¼'])
        
        for point in backup_data['daily_points']:
            ws_points.append([
                point['id'],
                point['child_id'],
                point['date'],
                point['korean_points'],
                point['math_points'],
                point['ssen_points'],
                point['reading_points'],
                point.get('piano_points', 0),
                point.get('english_points', 0),
                point.get('advanced_math_points', 0),
                point.get('writing_points', 0),
                point.get('manual_points', 0),
                point.get('manual_history', '[]'),
                point['total_points'],
                point['created_by'],
                point['created_at']
            ])
        
        # í¬ì¸íŠ¸ íˆìŠ¤í† ë¦¬ ì‹œíŠ¸
        ws_history = wb.create_sheet("í¬ì¸íŠ¸ë³€ê²½ì´ë ¥")
        ws_history.append(['ID', 'ì•„ë™ID', 'ë‚ ì§œ', 'ë³€ê²½íƒ€ì…', 'ë³€ê²½ì', 'ë³€ê²½ì¼', 'ë³€ê²½ì‚¬ìœ '])
        
        for history in backup_data['points_history']:
            ws_history.append([
                history['id'],
                history['child_id'],
                history['date'],
                history['change_type'],
                history['changed_by'],
                history['changed_at'],
                history['change_reason']
            ])
        
        # ì‚¬ìš©ì ì •ë³´ ì‹œíŠ¸
        ws_users = wb.create_sheet("ì‚¬ìš©ìì •ë³´")
        ws_users.append(['ID', 'ì‚¬ìš©ìëª…', 'ì´ë¦„', 'ê¶Œí•œ', 'ìƒì„±ì¼'])
        
        for user in backup_data['users']:
            ws_users.append([
                user['id'],
                user['username'],
                user['name'],
                user['role'],
                user['created_at']
            ])
        
        # ë©”íƒ€ë°ì´í„° ì‹œíŠ¸
        ws_meta = wb.create_sheet("ë°±ì—…ë©”íƒ€ë°ì´í„°")
        meta = backup_data['backup_metadata']
        ws_meta.append(['ë°±ì—…ID', meta['backup_id']])
        ws_meta.append(['ë°±ì—…íƒ€ì…', meta['backup_type']])
        ws_meta.append(['ë°±ì—…ì‹œê°„', meta['timestamp']])
        ws_meta.append(['ë°ì´í„°ë²„ì „', meta['data_version']])
        ws_meta.append(['ì•„ë™ìˆ˜', meta['records_count']['children']])
        ws_meta.append(['í¬ì¸íŠ¸ê¸°ë¡ìˆ˜', meta['records_count']['daily_points']])
        ws_meta.append(['ë³€ê²½ì´ë ¥ìˆ˜', meta['records_count']['points_history']])
        ws_meta.append(['ì‚¬ìš©ììˆ˜', meta['records_count']['users']])
        
        # ìŠ¤íƒ€ì¼ ì ìš©
        for ws in [ws_children, ws_points, ws_history, ws_users, ws_meta]:
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
        
        # íŒŒì¼ ì €ì¥
        wb.save(filepath)
        
        return filepath, None
        
    except Exception as e:
        return None, str(e)

def realtime_backup(child_id, action_type):
    """ì‹¤ì‹œê°„ ë°±ì—… ì‹¤í–‰ (í¬ì¸íŠ¸ ì…ë ¥ ì‹œ)"""
    try:
        # ë°±ì—… ë””ë ‰í† ë¦¬ ìƒì„±
        backup_dir = create_backup_directory()
        
        # ë°±ì—… ë°ì´í„° ìˆ˜ì§‘
        backup_data, error = get_backup_data()
        if error:
            error_msg = f"ì‹¤ì‹œê°„ ë°±ì—… ë°ì´í„° ìˆ˜ì§‘ ì‹¤íŒ¨: {error}"
            print(f"âŒ {error_msg}")
            create_backup_notification('ì‹¤ì‹œê°„', 'failed', error_msg)
            return False
        
        # JSON ë°±ì—… ìƒì„±
        json_path, error = create_json_backup(backup_data, backup_dir, 'realtime')
        if error:
            error_msg = f"ì‹¤ì‹œê°„ JSON ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
            print(f"âŒ {error_msg}")
            create_backup_notification('ì‹¤ì‹œê°„', 'failed', error_msg)
            return False
        
        # Excel ë°±ì—… ìƒì„±
        excel_path, error = create_excel_backup(backup_data, backup_dir, 'realtime')
        if error:
            error_msg = f"ì‹¤ì‹œê°„ Excel ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
            print(f"âŒ {error_msg}")
            create_backup_notification('ì‹¤ì‹œê°„', 'failed', error_msg)
            return False
        
        success_msg = f"ì‹¤ì‹œê°„ ë°±ì—… ì™„ë£Œ - {action_type}: {os.path.basename(json_path)}, {os.path.basename(excel_path)}"
        print(f"âœ… {success_msg}")
        create_backup_notification('ì‹¤ì‹œê°„', 'success', success_msg)
        return True
        
    except Exception as e:
        error_msg = f"ì‹¤ì‹œê°„ ë°±ì—… ì‹¤í–‰ ì¤‘ ì˜¤ë¥˜: {str(e)}"
        print(f"âŒ {error_msg}")
        create_backup_notification('ì‹¤ì‹œê°„', 'failed', error_msg)
        return False

def create_database_backup(backup_dir, backup_type='manual'):
    """ë°ì´í„°ë² ì´ìŠ¤ íŒŒì¼ ë°±ì—…"""
    try:
        # í˜„ì¬ DB íŒŒì¼ ê²½ë¡œ
        db_path = os.path.join(os.path.dirname(__file__), 'instance', 'child_center.db')
        
        if not os.path.exists(db_path):
            return None, "ë°ì´í„°ë² ì´ìŠ¤ íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤"
        
        # ë°±ì—… íŒŒì¼ëª…
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        backup_filename = f"{datetime.now().strftime('%Y-%m-%d')}_{timestamp.split('_')[1]}_{backup_type}.db"
        backup_path = os.path.join(backup_dir, 'database', backup_filename)
        
        # íŒŒì¼ ë³µì‚¬
        shutil.copy2(db_path, backup_path)
        
        return backup_path, None
        
    except Exception as e:
        return None, str(e)

# ìŠ¤ì¼€ì¤„ ë°±ì—… ì‹œìŠ¤í…œ
def daily_backup():
    """ì¼ì¼ ë°±ì—… ì‹¤í–‰ (ë§¤ì¼ 22ì‹œ)"""
    try:
        print("ğŸ”„ ì¼ì¼ ë°±ì—… ì‹œì‘...")
        
        # Flask ì•± ì»¨í…ìŠ¤íŠ¸ ë‚´ì—ì„œ ì‹¤í–‰
        with app.app_context():
            # ë°±ì—… ë””ë ‰í† ë¦¬ ìƒì„±
            backup_dir = create_backup_directory()
        
            # ë°±ì—… ë°ì´í„° ìˆ˜ì§‘
            backup_data, error = get_backup_data()
            if error:
                error_msg = f"ì¼ì¼ ë°±ì—… ë°ì´í„° ìˆ˜ì§‘ ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì¼ì¼', 'failed', error_msg)
                return False
        
            # JSON ë°±ì—… ìƒì„±
            json_path, error = create_json_backup(backup_data, backup_dir, 'daily')
            if error:
                error_msg = f"ì¼ì¼ JSON ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì¼ì¼', 'failed', error_msg)
                return False
        
            # Excel ë°±ì—… ìƒì„±
            excel_path, error = create_excel_backup(backup_data, backup_dir, 'daily')
            if error:
                error_msg = f"ì¼ì¼ Excel ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì¼ì¼', 'failed', error_msg)
                return False
        
            # ë°ì´í„°ë² ì´ìŠ¤ ë°±ì—… ìƒì„±
            db_path, error = create_database_backup(backup_dir, 'daily')
            if error:
                error_msg = f"ì¼ì¼ ë°ì´í„°ë² ì´ìŠ¤ ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì¼ì¼', 'failed', error_msg)
                return False
        
            success_msg = f"ì¼ì¼ ë°±ì—… ì™„ë£Œ: {os.path.basename(json_path)}, {os.path.basename(excel_path)}, {os.path.basename(db_path)}"
            print(f"âœ… {success_msg}")
            create_backup_notification('ì¼ì¼', 'success', success_msg)
            return True
        
    except Exception as e:
        error_msg = f"ì¼ì¼ ë°±ì—… ì‹¤í–‰ ì¤‘ ì˜¤ë¥˜: {str(e)}"
        print(f"âŒ {error_msg}")
        create_backup_notification('ì¼ì¼', 'failed', error_msg)
        return False

def monthly_backup():
    """ì›”ê°„ ë°±ì—… ì‹¤í–‰ (ë§¤ì›” ë§ˆì§€ë§‰ ë‚  23ì‹œ)"""
    try:
        print("ğŸ”„ ì›”ê°„ ë°±ì—… ì‹œì‘...")
        
        # Flask ì•± ì»¨í…ìŠ¤íŠ¸ ë‚´ì—ì„œ ì‹¤í–‰
        with app.app_context():
            # ë°±ì—… ë””ë ‰í† ë¦¬ ìƒì„±
            backup_dir = create_backup_directory()
        
            # ë°±ì—… ë°ì´í„° ìˆ˜ì§‘
            backup_data, error = get_backup_data()
            if error:
                error_msg = f"ì›”ê°„ ë°±ì—… ë°ì´í„° ìˆ˜ì§‘ ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì›”ê°„', 'failed', error_msg)
            return False
        
            # JSON ë°±ì—… ìƒì„±
            json_path, error = create_json_backup(backup_data, backup_dir, 'monthly')
            if error:
                error_msg = f"ì›”ê°„ JSON ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì›”ê°„', 'failed', error_msg)
                return False
        
            # Excel ë°±ì—… ìƒì„±
            excel_path, error = create_excel_backup(backup_data, backup_dir, 'monthly')
            if error:
                error_msg = f"ì›”ê°„ Excel ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì›”ê°„', 'failed', error_msg)
                return False
        
            # ë°ì´í„°ë² ì´ìŠ¤ ë°±ì—… ìƒì„±
            db_path, error = create_database_backup(backup_dir, 'monthly')
            if error:
                error_msg = f"ì›”ê°„ ë°ì´í„°ë² ì´ìŠ¤ ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}"
                print(f"âŒ {error_msg}")
                create_backup_notification('ì›”ê°„', 'failed', error_msg)
                return False
        
            success_msg = f"ì›”ê°„ ë°±ì—… ì™„ë£Œ: {os.path.basename(json_path)}, {os.path.basename(excel_path)}, {os.path.basename(db_path)}"
            print(f"âœ… {success_msg}")
            create_backup_notification('ì›”ê°„', 'success', success_msg)
            return True
        
    except Exception as e:
        error_msg = f"ì›”ê°„ ë°±ì—… ì‹¤í–‰ ì¤‘ ì˜¤ë¥˜: {str(e)}"
        print(f"âŒ {error_msg}")
        create_backup_notification('ì›”ê°„', 'failed', error_msg)
        return False

def run_scheduler():
    """ìŠ¤ì¼€ì¤„ëŸ¬ ì‹¤í–‰ í•¨ìˆ˜"""
    try:
        # ì¼ì¼ ë°±ì—… ìŠ¤ì¼€ì¤„ (ë§¤ì¼ 22ì‹œ)
        schedule.every().day.at("22:00").do(daily_backup)
        
        # ì›”ê°„ ë°±ì—… ì²´í¬ í•¨ìˆ˜ (ë§¤ì¼ 23ì‹œì— ì›”ì˜ ë§ˆì§€ë§‰ ë‚ ì¸ì§€ í™•ì¸)
        def check_monthly_backup():
            now = datetime.now()
            # ì˜¤ëŠ˜ì´ ì›”ì˜ ë§ˆì§€ë§‰ ë‚ ì´ê³  23ì‹œì¸ì§€ í™•ì¸
            last_day_of_month = (now.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            if now.day == last_day_of_month.day and now.hour == 23:
                monthly_backup()
        
        schedule.every().day.at("23:00").do(check_monthly_backup)
        
        print("âœ… ìŠ¤ì¼€ì¤„ ë°±ì—… ì‹œìŠ¤í…œ ì‹œì‘ë¨")
        print("   - ì¼ì¼ ë°±ì—…: ë§¤ì¼ 22:00")
        print("   - ì›”ê°„ ë°±ì—…: ë§¤ì›” ë§ˆì§€ë§‰ ë‚  23:00")
        
        # ìŠ¤ì¼€ì¤„ëŸ¬ ë£¨í”„ ì‹¤í–‰
        while True:
            schedule.run_pending()
            time.sleep(60)  # 1ë¶„ë§ˆë‹¤ ì²´í¬
            
    except Exception as e:
        print(f"âŒ ìŠ¤ì¼€ì¤„ëŸ¬ ì‹¤í–‰ ì¤‘ ì˜¤ë¥˜: {str(e)}")

def start_backup_scheduler():
    """ë°±ê·¸ë¼ìš´ë“œì—ì„œ ìŠ¤ì¼€ì¤„ëŸ¬ ì‹œì‘"""
    try:
        scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
        scheduler_thread.start()
        print("âœ… ë°±ì—… ìŠ¤ì¼€ì¤„ëŸ¬ê°€ ë°±ê·¸ë¼ìš´ë“œì—ì„œ ì‹œì‘ë˜ì—ˆìŠµë‹ˆë‹¤.")
    except Exception as e:
        print(f"âŒ ë°±ì—… ìŠ¤ì¼€ì¤„ëŸ¬ ì‹œì‘ ì‹¤íŒ¨: {str(e)}")

@app.route('/backup/manual', methods=['POST'])
@login_required
def backup_manual():
    """ìˆ˜ë™ ë°±ì—… ì‹¤í–‰"""
    if current_user.role != 'ê°œë°œì':
        flash('ê°œë°œìë§Œ ë°±ì—…ì„ ì‹¤í–‰í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('settings_data'))
    
    try:
        # ë°±ì—… ë””ë ‰í† ë¦¬ ìƒì„±
        backup_dir = create_backup_directory()
        
        # ë°±ì—… ë°ì´í„° ìˆ˜ì§‘
        backup_data, error = get_backup_data()
        if error:
            error_msg = f'ë°±ì—… ë°ì´í„° ìˆ˜ì§‘ ì‹¤íŒ¨: {error}'
            flash(error_msg, 'error')
            create_backup_notification('ìˆ˜ë™', 'failed', error_msg)
            return redirect(url_for('settings_data'))
        
        # JSON ë°±ì—… ìƒì„±
        json_path, error = create_json_backup(backup_data, backup_dir, 'manual')
        if error:
            error_msg = f'JSON ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}'
            flash(error_msg, 'error')
            create_backup_notification('ìˆ˜ë™', 'failed', error_msg)
            return redirect(url_for('settings_data'))
        
        # Excel ë°±ì—… ìƒì„±
        excel_path, error = create_excel_backup(backup_data, backup_dir, 'manual')
        if error:
            error_msg = f'Excel ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}'
            flash(error_msg, 'error')
            create_backup_notification('ìˆ˜ë™', 'failed', error_msg)
            return redirect(url_for('settings_data'))
        
        # ë°ì´í„°ë² ì´ìŠ¤ ë°±ì—… ìƒì„±
        db_path, error = create_database_backup(backup_dir, 'manual')
        if error:
            error_msg = f'ë°ì´í„°ë² ì´ìŠ¤ ë°±ì—… ìƒì„± ì‹¤íŒ¨: {error}'
            flash(error_msg, 'error')
            create_backup_notification('ìˆ˜ë™', 'failed', error_msg)
            return redirect(url_for('settings_data'))
        
        success_msg = f'ë°±ì—…ì´ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤. JSON: {os.path.basename(json_path)}, Excel: {os.path.basename(excel_path)}, DB: {os.path.basename(db_path)}'
        flash(success_msg, 'success')
        create_backup_notification('ìˆ˜ë™', 'success', success_msg)
        return redirect(url_for('settings_data'))
        
    except Exception as e:
        error_msg = f'ë°±ì—… ì‹¤í–‰ ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {str(e)}'
        flash(error_msg, 'error')
        create_backup_notification('ìˆ˜ë™', 'failed', error_msg)
        return redirect(url_for('settings_data'))

@app.route('/backup/list')
@login_required
def backup_list():
    """ë°±ì—… íŒŒì¼ ëª©ë¡ ì¡°íšŒ"""
    if current_user.role != 'ê°œë°œì':
        flash('ê°œë°œìë§Œ ë°±ì—… ëª©ë¡ì„ ì¡°íšŒí•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.', 'error')
        return redirect(url_for('settings_data'))
    
    try:
        backup_dir = create_backup_directory()
        
        # ë°±ì—… íŒŒì¼ ëª©ë¡ ì¡°íšŒ
        backups = []
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith(('.json', '.xlsx', '.db')):
                    file_path = os.path.join(backup_dir, filename)
                    file_stat = os.stat(file_path)
                    
                    # íŒŒì¼ íƒ€ì… ì¶”ì¶œ
                    if 'realtime' in filename:
                        backup_type = 'realtime'
                    elif 'daily' in filename:
                        backup_type = 'daily'
                    elif 'monthly' in filename:
                        backup_type = 'monthly'
                    elif 'manual' in filename:
                        backup_type = 'manual'
                    else:
                        backup_type = 'unknown'
                    
                    # í¬ê¸°ë¥¼ MBë¡œ ë³€í™˜
                    size_mb = round(file_stat.st_size / (1024 * 1024), 2)
                    
                    backups.append({
                        'filename': filename,
                        'type': backup_type,
                        'size_mb': size_mb,
                        'created_at': datetime.fromtimestamp(file_stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                        'filepath': file_path,
                        'restore_safe': True  # ê¸°ë³¸ì ìœ¼ë¡œ ë³µêµ¬ ê°€ëŠ¥ìœ¼ë¡œ ì„¤ì •
                    })
        
        # ìµœì‹  íŒŒì¼ë¶€í„° ì •ë ¬
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        
        return render_template('backup/list.html', backups=backups)
        
    except Exception as e:
        flash(f'ë°±ì—… ëª©ë¡ ì¡°íšŒ ì‹¤íŒ¨: {str(e)}', 'error')
        return redirect(url_for('settings_data'))

@app.route('/backup/status')
@login_required
def backup_status():
    """ë°±ì—… ìƒíƒœ ë° ëª©ë¡ ì¡°íšŒ (JSON API)"""
    if current_user.role != 'ê°œë°œì':
        return jsonify({'error': 'ê°œë°œìë§Œ ì ‘ê·¼í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.'}), 403
    
    try:
        backup_dir = create_backup_directory()
        
        # ë°±ì—… íŒŒì¼ ëª©ë¡ ì¡°íšŒ (ëª¨ë“  í•˜ìœ„ ë””ë ‰í† ë¦¬ í¬í•¨)
        backups = []
        if os.path.exists(backup_dir):
            # ë£¨íŠ¸ ë””ë ‰í† ë¦¬ ê²€ìƒ‰
            for filename in os.listdir(backup_dir):
                if filename.endswith(('.json', '.xlsx', '.db')):
                    file_path = os.path.join(backup_dir, filename)
                    file_stat = os.stat(file_path)
                    
                    # í¬ê¸°ë¥¼ MBë¡œ ë³€í™˜
                    size_mb = round(file_stat.st_size / (1024 * 1024), 2)
                    
                    backups.append({
                        'filename': filename,
                        'size_mb': size_mb,
                        'created_at': datetime.fromtimestamp(file_stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                    })
            
            # í•˜ìœ„ ë””ë ‰í† ë¦¬ë“¤ ê²€ìƒ‰
            subdirs = ['realtime', 'daily', 'monthly', 'database']
            for subdir in subdirs:
                subdir_path = os.path.join(backup_dir, subdir)
                if os.path.exists(subdir_path):
                    for filename in os.listdir(subdir_path):
                        if filename.endswith(('.json', '.xlsx', '.db')):
                            file_path = os.path.join(subdir_path, filename)
                            file_stat = os.stat(file_path)
                            
                            # í¬ê¸°ë¥¼ MBë¡œ ë³€í™˜
                            size_mb = round(file_stat.st_size / (1024 * 1024), 2)
                            
                            backups.append({
                                'filename': f"{subdir}/{filename}",
                                'size_mb': size_mb,
                                'created_at': datetime.fromtimestamp(file_stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                            })
        
        # ìµœì‹  íŒŒì¼ë¶€í„° ì •ë ¬
        backups.sort(key=lambda x: x['created_at'], reverse=True)
        
        return jsonify({
            'backups': backups,
            'total_count': len(backups),
            'backup_dir': backup_dir
        })
        
    except Exception as e:
        return jsonify({'error': f'ë°±ì—… ìƒíƒœ ì¡°íšŒ ì‹¤íŒ¨: {str(e)}'}), 500

if __name__ == '__main__':
    # Firebase ì´ˆê¸°í™”
    initialize_firebase()
    
    # ë°±ì—… ìŠ¤ì¼€ì¤„ëŸ¬ ì‹œì‘
    start_backup_scheduler()
    
    # init_db() ì œê±° - ì„œë²„ ì¬ì‹œì‘ ì‹œ ë°ì´í„° ì´ˆê¸°í™” ë°©ì§€
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)
else:
    # ë°°í¬ëœ í™˜ê²½ì—ì„œë„ ë°ì´í„°ë² ì´ìŠ¤ ì´ˆê¸°í™”
    with app.app_context():
        # Firebase ì´ˆê¸°í™”
        initialize_firebase()
        
        db.create_all()
        # ê¸°ë³¸ ì‚¬ìš©ìê°€ ì—†ìœ¼ë©´ ìƒì„± (í•œ ë²ˆë§Œ) - Firebase ì‚¬ìš© ì‹œ ì„ì‹œ ë¹„í™œì„±í™”
        # if not User.query.filter_by(username='center_head').first():
        #     # init_db() ì œê±° - ì‹¤ì œ ë°ì´í„° ë³´í˜¸
        #     pass
    
    # ë°°í¬ í™˜ê²½ì—ì„œë„ ë°±ì—… ìŠ¤ì¼€ì¤„ëŸ¬ ì‹œì‘
    start_backup_scheduler()