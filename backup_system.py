"""
지역아동센터 학습관리 시스템 - 백업 시스템
"""

import os
import json
import threading
import calendar
import shutil
import time
from datetime import datetime
from flask import Blueprint, request, redirect, url_for, flash, session, jsonify, render_template
from flask_login import login_required, current_user

# 백업 시스템을 위한 import
try:
    import pandas as pd
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    BACKUP_EXCEL_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Excel 백업 기능을 위한 패키지가 설치되지 않았습니다: {e}")
    print("   pip install pandas openpyxl 명령어로 설치하세요.")
    BACKUP_EXCEL_AVAILABLE = False

# Blueprint 생성
backup_bp = Blueprint('backup', __name__, url_prefix='/backup')

# 데이터베이스 모델들은 app.py에서 import
from app import db, Child, DailyPoints, PointsHistory, User, Notification, app

def create_backup_directory():
    """백업 디렉토리 생성"""
    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    # 하위 디렉토리들 생성
    subdirs = ['daily', 'monthly', 'realtime', 'database']
    for subdir in subdirs:
        subdir_path = os.path.join(backup_dir, subdir)
        if not os.path.exists(subdir_path):
            os.makedirs(subdir_path)
    
    return backup_dir

def get_backup_data():
    """백업할 데이터 수집"""
    try:
        # 아동 정보
        children = Child.query.all()
        children_data = []
        for child in children:
            child_dict = {
                'id': child.id,
                'name': child.name,
                'grade': child.grade,
                'cumulative_points': child.cumulative_points,
                'created_at': child.created_at.isoformat() if child.created_at else None
            }
            children_data.append(child_dict)
        
        # 일일 포인트 기록
        daily_points = DailyPoints.query.all()
        daily_points_data = []
        for point in daily_points:
            point_dict = {
                'id': point.id,
                'child_id': point.child_id,
                'date': point.date.isoformat() if point.date else None,
                'korean_points': point.korean_points,
                'math_points': point.math_points,
                'ssen_points': point.ssen_points,
                'reading_points': point.reading_points,
                'total_points': point.total_points,
                'created_by': point.created_by,
                'created_at': point.created_at.isoformat() if point.created_at else None,
                'updated_at': point.updated_at.isoformat() if point.updated_at else None
            }
            daily_points_data.append(point_dict)
        
        # 포인트 히스토리
        points_history = PointsHistory.query.all()
        history_data = []
        for history in points_history:
            history_dict = {
                'id': history.id,
                'child_id': history.child_id,
                'date': history.date.isoformat() if history.date else None,
                'old_korean_points': history.old_korean_points,
                'old_math_points': history.old_math_points,
                'old_ssen_points': history.old_ssen_points,
                'old_reading_points': history.old_reading_points,
                'old_total_points': history.old_total_points,
                'new_korean_points': history.new_korean_points,
                'new_math_points': history.new_math_points,
                'new_ssen_points': history.new_ssen_points,
                'new_reading_points': history.new_reading_points,
                'new_total_points': history.new_total_points,
                'change_type': history.change_type,
                'changed_by': history.changed_by,
                'changed_at': history.changed_at.isoformat() if history.changed_at else None,
                'change_reason': history.change_reason
            }
            history_data.append(history_dict)
        
        # 아동 메모 (ChildNote)
        child_notes = ChildNote.query.all()
        notes_data = []
        for note in child_notes:
            note_dict = {
                'id': note.id,
                'child_id': note.child_id,
                'note': note.note,
                'created_by': note.created_by,
                'created_at': note.created_at.isoformat() if note.created_at else None,
                'updated_at': note.updated_at.isoformat() if note.updated_at else None
            }
            notes_data.append(note_dict)
        
        # 사용자 정보
        users = User.query.all()
        users_data = []
        for user in users:
            user_dict = {
                'id': user.id,
                'username': user.username,
                'name': user.name,
                'role': user.role,
                'created_at': user.created_at.isoformat() if user.created_at else None
            }
            users_data.append(user_dict)
        
        backup_data = {
            'backup_metadata': {
                'backup_id': datetime.now().strftime('%Y-%m-%d_%H-%M-%S'),
                'backup_type': 'manual',
                'timestamp': datetime.now().isoformat(),
                'data_version': '1.0.0',
                'records_count': {
                    'children': len(children_data),
                    'daily_points': len(daily_points_data),
                    'points_history': len(history_data),
                    'child_notes': len(notes_data),
                    'users': len(users_data)
                }
            },
            'children': children_data,
            'daily_points': daily_points_data,
            'points_history': history_data,
            'child_notes': notes_data,
            'users': users_data
        }
        
        return backup_data, None
        
    except Exception as e:
        return None, str(e)

def create_json_backup(backup_data, backup_dir, backup_type='manual'):
    """JSON 형태로 백업 생성"""
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        if backup_type == 'daily':
            filename = f"{datetime.now().strftime('%Y-%m-%d')}_{timestamp.split('_')[1]}.json"
            filepath = os.path.join(backup_dir, 'daily', filename)
        elif backup_type == 'monthly':
            filename = f"{datetime.now().strftime('%Y-%m')}_archive.json"
            filepath = os.path.join(backup_dir, 'monthly', filename)
        else:
            filename = f"{timestamp}.json"
            filepath = os.path.join(backup_dir, 'realtime', filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        return filepath, None
        
    except Exception as e:
        return None, str(e)

def create_excel_backup(backup_data, backup_dir, backup_type='manual'):
    """Excel 형태로 백업 생성"""
    if not BACKUP_EXCEL_AVAILABLE:
        print("❌ Excel 백업을 위한 패키지가 설치되지 않았습니다.")
        return None, "pandas 또는 openpyxl 패키지가 설치되지 않았습니다."
    
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        if backup_type == 'daily':
            filename = f"{datetime.now().strftime('%Y-%m-%d')}_{timestamp.split('_')[1]}.xlsx"
            filepath = os.path.join(backup_dir, 'daily', filename)
        elif backup_type == 'monthly':
            filename = f"{datetime.now().strftime('%Y-%m')}_archive.xlsx"
            filepath = os.path.join(backup_dir, 'monthly', filename)
        else:
            filename = f"{timestamp}.xlsx"
            filepath = os.path.join(backup_dir, 'realtime', filename)
        
        # Excel 워크북 생성
        wb = Workbook()
        
        # 아동 정보 시트
        ws_children = wb.active
        ws_children.title = "아동정보"
        ws_children.append(['ID', '이름', '학년', '누적포인트', '생성일'])
        
        for child in backup_data['children']:
            ws_children.append([
                child['id'],
                child['name'],
                child['grade'],
                child['cumulative_points'],
                child['created_at']
            ])
        
        # 포인트 기록 시트
        ws_points = wb.create_sheet("포인트기록")
        ws_points.append(['ID', '아동ID', '날짜', '국어', '수학', '쎈수학', '독서', '총점', '입력자', '생성일'])
        
        for point in backup_data['daily_points']:
            ws_points.append([
                point['id'],
                point['child_id'],
                point['date'],
                point['korean_points'],
                point['math_points'],
                point['ssen_points'],
                point['reading_points'],
                point['total_points'],
                point['created_by'],
                point['created_at']
            ])
        
        # 포인트 히스토리 시트
        ws_history = wb.create_sheet("포인트변경이력")
        ws_history.append(['ID', '아동ID', '날짜', '변경타입', '변경자', '변경일', '변경사유'])
        
        for history in backup_data['points_history']:
            ws_history.append([
                history['id'],
                history['child_id'],
                history['date'],
                history['change_type'],
                history['changed_by'],
                history['changed_at'],
                history['change_reason']
            ])
        
        # 사용자 정보 시트
        ws_users = wb.create_sheet("사용자정보")
        ws_users.append(['ID', '사용자명', '이름', '권한', '생성일'])
        
        for user in backup_data['users']:
            ws_users.append([
                user['id'],
                user['username'],
                user['name'],
                user['role'],
                user['created_at']
            ])
        
        # 메타데이터 시트
        ws_meta = wb.create_sheet("백업메타데이터")
        meta = backup_data['backup_metadata']
        ws_meta.append(['백업ID', meta['backup_id']])
        ws_meta.append(['백업타입', meta['backup_type']])
        ws_meta.append(['백업시간', meta['timestamp']])
        ws_meta.append(['데이터버전', meta['data_version']])
        ws_meta.append(['아동수', meta['records_count']['children']])
        ws_meta.append(['포인트기록수', meta['records_count']['daily_points']])
        ws_meta.append(['변경이력수', meta['records_count']['points_history']])
        ws_meta.append(['사용자수', meta['records_count']['users']])
        
        # 스타일 적용
        for ws in [ws_children, ws_points, ws_history, ws_users, ws_meta]:
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
        
        # 파일 저장
        wb.save(filepath)
        
        return filepath, None
        
    except Exception as e:
        return None, str(e)

def create_database_backup(backup_dir):
    """데이터베이스 파일 백업"""
    try:
        # 현재 DB 파일 경로
        db_path = os.path.join(os.path.dirname(__file__), 'instance', 'child_center.db')
        
        if not os.path.exists(db_path):
            return None, "데이터베이스 파일을 찾을 수 없습니다"
        
        # 백업 파일명
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        backup_filename = f"{datetime.now().strftime('%Y-%m-%d')}_{timestamp.split('_')[1]}_full.db"
        backup_path = os.path.join(backup_dir, 'database', backup_filename)
        
        # 파일 복사
        shutil.copy2(db_path, backup_path)
        
        return backup_path, None
        
    except Exception as e:
        return None, str(e)

def manual_backup():
    """수동 백업 실행"""
    try:
        # 백업 디렉토리 생성
        backup_dir = create_backup_directory()
        
        # 백업 데이터 수집
        backup_data, error = get_backup_data()
        if error:
            return False, f"데이터 수집 실패: {error}"
        
        # JSON 백업 생성
        json_path, error = create_json_backup(backup_data, backup_dir, 'manual')
        if error:
            return False, f"JSON 백업 생성 실패: {error}"
        
        # Excel 백업 생성
        excel_path, error = create_excel_backup(backup_data, backup_dir, 'manual')
        if error:
            return False, f"Excel 백업 생성 실패: {error}"
        
        # DB 백업 생성
        db_path, error = create_database_backup(backup_dir)
        if error:
            return False, f"DB 백업 생성 실패: {error}"
        
        backup_info = {
            'json_file': os.path.basename(json_path),
            'excel_file': os.path.basename(excel_path),
            'db_file': os.path.basename(db_path),
            'backup_dir': backup_dir,
            'timestamp': datetime.now().isoformat()
        }
        
        return True, backup_info
        
    except Exception as e:
        return False, f"백업 실행 중 오류 발생: {str(e)}"

def realtime_backup(child_id, action_type):
    """실시간 백업 실행 (포인트 입력 시)"""
    try:
        # 백업 디렉토리 생성
        backup_dir = create_backup_directory()
        
        # 백업 데이터 수집
        backup_data, error = get_backup_data()
        if error:
            print(f"실시간 백업 데이터 수집 실패: {error}")
            return False
        
        # JSON 백업 생성
        json_path, error = create_json_backup(backup_data, backup_dir, 'realtime')
        if error:
            print(f"실시간 JSON 백업 생성 실패: {error}")
            return False
        
        # Excel 백업 생성
        excel_path, error = create_excel_backup(backup_data, backup_dir, 'realtime')
        if error:
            print(f"실시간 Excel 백업 생성 실패: {error}")
            return False
        
        print(f"✅ 실시간 백업 완료 - {action_type}: {os.path.basename(json_path)}, {os.path.basename(excel_path)}")
        return True
        
    except Exception as e:
        print(f"❌ 실시간 백업 실행 중 오류: {str(e)}")
        return False

# 백업 관련 라우트들
@backup_bp.route('/manual', methods=['POST'])
@login_required
def backup_manual():
    """수동 백업 실행"""
    if current_user.role != 'admin':
        flash('관리자만 백업을 실행할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    success, result = manual_backup()
    
    if success:
        backup_info = result
        flash(f'백업이 완료되었습니다. JSON: {backup_info["json_file"]}, Excel: {backup_info["excel_file"]}, DB: {backup_info["db_file"]}', 'success')
    else:
        flash(f'백업 실패: {result}', 'error')
    
    return redirect(url_for('dashboard'))

@backup_bp.route('/status')
@login_required
def backup_status():
    """백업 상태 확인"""
    if current_user.role != 'admin':
        return jsonify({'error': '권한이 없습니다'}), 403
    
    try:
        backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
        
        if not os.path.exists(backup_dir):
            return jsonify({'status': 'no_backup_directory'})
        
        # 백업 파일 목록 확인 로직...
        status = {
            'status': 'ok',
            'backup_available': True
        }
        
        return jsonify(status)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@backup_bp.route('/list')
@login_required
def backup_list():
    """백업 파일 목록 조회"""
    if current_user.role != 'admin':
        flash('관리자만 백업 목록을 조회할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # 백업 목록 조회 로직...
        backups = []  # 실제로는 list_available_backups() 호출
        return render_template('backup/list.html', backups=backups)
    except Exception as e:
        flash(f'백업 목록 조회 실패: {str(e)}', 'error')
        return redirect(url_for('dashboard'))
